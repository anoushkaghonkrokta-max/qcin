"""
QCI Notification Engine — single-file Flask application
"""
import csv
import hashlib
import hmac
import io
import json
import logging
import os
import secrets
import smtplib
import psycopg2
import psycopg2.extras
import psycopg2.pool
import struct
import time
import urllib.request
import pytz
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

from cryptography.fernet import Fernet
from flask import (Flask, flash, jsonify, redirect, render_template_string,
                   request, url_for, Response, session)
from markupsafe import escape as _xe
from werkzeug.security import check_password_hash as _check_pw


def h(v):
    """HTML-escape a value for safe injection into f-string HTML."""
    if v is None:
        return ""
    return str(_xe(str(v)))

def generate_password_hash(pw):
    from werkzeug.security import generate_password_hash as _gph
    return _gph(pw, method="pbkdf2:sha256")

def check_password_hash(hsh, pw):
    return _check_pw(hsh, pw)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

app = Flask(__name__)

# ── Secret key (must be set via environment variable on Render) ───────────────
_secret_key_env = os.environ.get("SECRET_KEY")
if not _secret_key_env:
    raise RuntimeError(
        "SECRET_KEY environment variable is not set. "
        "Generate one with: python3 -c \"import secrets; print(secrets.token_hex(32))\" "
        "and add it in the Render project dashboard under Environment → Environment Variables."
    )
app.secret_key = _secret_key_env

DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ── Timezone ─────────────────────────────────────────────────────────────────
_IST = pytz.timezone("Asia/Kolkata")

def now_ist() -> datetime:
    """Return the current datetime in IST (Asia/Kolkata)."""
    return datetime.now(_IST)

# ── Fernet encryption (must be set via environment variable on Render) ────────
_fk_env = os.environ.get("FERNET_KEY")
if not _fk_env:
    raise RuntimeError(
        "FERNET_KEY environment variable is not set. "
        "Generate one with: python3 -c \"from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())\" "
        "and add it in the Render project dashboard under Environment → Environment Variables."
    )
_FERNET_KEY = _fk_env.encode()

_fernet = Fernet(_FERNET_KEY)


def encrypt_str(text: str) -> str:
    return _fernet.encrypt(text.encode()).decode()


def decrypt_str(text: str) -> str:
    try:
        return _fernet.decrypt(text.encode()).decode()
    except Exception:
        return ""


def get_app_setting(key: str, default: str = "") -> str:
    conn = get_db()
    row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row[0] if row else default


def set_app_setting(key: str, value: str):
    conn = get_db()
    conn.execute(
        "INSERT INTO app_settings (key, value) VALUES (?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value)
    )
    conn.commit()
    conn.close()


def log_audit(event_type: str, application_id: str = None, detail: str = "",
              user_name: str = None, board_id: int = None):
    """Write a tamper-evident row to audit_log. Each entry hashes previous entry + payload.
    Never raises — audit failures must not crash the calling request."""
    try:
        conn = get_db()
        timestamp = now_ist().strftime("%Y-%m-%d %H:%M:%S")
        # Hash-chaining: try to read previous hash (column may not exist on older DBs)
        entry_hash = None
        try:
            prev_row = conn.execute(
                "SELECT entry_hash FROM audit_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
            prev_hash = prev_row["entry_hash"] if prev_row and prev_row["entry_hash"] else "GENESIS"
            raw = f"{prev_hash}|{timestamp}|{event_type}|{application_id}|{detail}|{user_name}"
            entry_hash = hashlib.sha256(raw.encode()).hexdigest()
        except Exception:
            pass  # entry_hash column may not exist on older DBs — gracefully skip
        # Try INSERT with entry_hash first; fall back to without if column missing
        try:
            conn.execute(
                "INSERT INTO audit_log "
                "(timestamp, application_id, event_type, detail, user_name, board_id, entry_hash) "
                "VALUES (?,?,?,?,?,?,?)",
                (timestamp, application_id, event_type,
                 detail[:500] if detail else "", user_name, board_id, entry_hash)
            )
        except Exception:
            # Fallback: insert without entry_hash (older schema)
            # PostgreSQL requires a rollback after any failed statement before retrying
            conn.rollback()
            conn.execute(
                "INSERT INTO audit_log "
                "(timestamp, application_id, event_type, detail, user_name, board_id) "
                "VALUES (?,?,?,?,?,?)",
                (timestamp, application_id, event_type,
                 detail[:500] if detail else "", user_name, board_id)
            )
        conn.commit()
        conn.close()
    except Exception as _audit_err:
        log.error("log_audit failed silently: %s", _audit_err)


def _verify_api_key(request) -> dict:
    """Check X-API-Key header. Returns api_key row or None."""
    raw = request.headers.get("X-API-Key", "").strip()
    if not raw:
        return None
    key_hash = hashlib.sha256(raw.encode()).hexdigest()
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM api_keys WHERE key_hash=? AND is_active=1", (key_hash,)
    ).fetchone()
    if row:
        conn.execute("UPDATE api_keys SET last_used=? WHERE id=?",
                     (now_ist().strftime("%Y-%m-%d %H:%M:%S"), row["id"]))
        conn.commit()
    conn.close()
    return dict(row) if row else None


def log_stage_transition(application_id: str, from_stage: str, to_stage: str,
                         changed_by: str = None, board_id: int = None):
    """Record a stage transition in stage_history."""
    conn = get_db()
    conn.execute(
        "INSERT INTO stage_history (timestamp, application_id, from_stage, to_stage, changed_by, board_id) "
        "VALUES (?,?,?,?,?,?)",
        (now_ist().strftime("%Y-%m-%d %H:%M:%S"), application_id, from_stage, to_stage,
         changed_by, board_id)
    )
    conn.commit()
    conn.close()


# ── TOTP (software 2FA — no external library needed) ─────────────────────────
def _hotp(key_bytes: bytes, counter: int) -> int:
    msg = struct.pack(">Q", counter)
    h = hmac.new(key_bytes, msg, hashlib.sha1).digest()
    offset = h[-1] & 0x0F
    return ((h[offset] & 0x7F) << 24 | h[offset+1] << 16 |
            h[offset+2] << 8 | h[offset+3]) % 1_000_000


def totp_now(base32_secret: str) -> str:
    import base64
    try:
        key = base64.b32decode(base32_secret.upper() + "=" * (-len(base32_secret) % 8))
        return f"{_hotp(key, int(time.time()) // 30):06d}"
    except Exception:
        return "000000"


def totp_verify(base32_secret: str, code: str, window: int = 1) -> bool:
    import base64
    try:
        key = base64.b32decode(base32_secret.upper() + "=" * (-len(base32_secret) % 8))
        t = int(time.time()) // 30
        for delta in range(-window, window + 1):
            if f"{_hotp(key, t + delta):06d}" == code.strip():
                return True
        return False
    except Exception:
        return False


def totp_generate_secret() -> str:
    import base64
    return base64.b32encode(secrets.token_bytes(20)).decode().rstrip("=")


def totp_provisioning_uri(secret: str, username: str) -> str:
    import urllib.parse
    padded = secret + "=" * (-len(secret) % 8)
    params = urllib.parse.urlencode({"secret": padded, "issuer": "QCI Engine"})
    return f"otpauth://totp/QCI%20Engine:{urllib.parse.quote(username)}?{params}"


# ── Email queue helpers ───────────────────────────────────────────────────────
def queue_email(programme_name: str, notification_type: str, to_email: str,
                cc_email: str, sender_email: str, sender_password_enc: str,
                ph: dict, smtp_host: str, smtp_port: int,
                application_id: str = None, board_id: int = None,
                stage_name: str = None):
    """Resolve template + placeholders and add to email_queue.
    Per-stage overrides take priority over programme-level templates."""
    conn = get_db()
    # Check per-stage override first
    tmpl = None
    if stage_name:
        tmpl = conn.execute(
            "SELECT subject_line, email_body FROM stage_email_override "
            "WHERE programme_name=? AND stage_name=? AND notification_type=?",
            (programme_name, stage_name, notification_type)
        ).fetchone()
    if not tmpl:
        tmpl = conn.execute(
            "SELECT subject_line, email_body FROM email_templates "
            "WHERE programme_name=? AND notification_type=?",
            (programme_name, notification_type)
        ).fetchone()
    if not tmpl:
        conn.close()
        return
    subj, body = tmpl["subject_line"], tmpl["email_body"]
    for k, v in ph.items():
        subj = subj.replace("{{" + k + "}}", str(v))
        body = body.replace("{{" + k + "}}", str(v))
    conn.execute(
        """INSERT INTO email_queue
           (queued_at, application_id, programme_name, notification_type,
            to_email, cc_email, sender_email, sender_password, smtp_host, smtp_port,
            subject, body, status, board_id)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'pending',?)""",
        (now_ist().strftime("%Y-%m-%d %H:%M:%S"), application_id, programme_name,
         notification_type, to_email, cc_email or "", sender_email, sender_password_enc,
         smtp_host, smtp_port, subj, body, board_id)
    )
    conn.commit()
    conn.close()


def process_email_queue(max_retries: int = 3) -> dict:
    """Process pending email_queue items. Groups by sender credentials to reuse SMTP connections."""
    conn = get_db()
    pending = [dict(r) for r in conn.execute(
        "SELECT * FROM email_queue WHERE status='pending' AND attempts < ?", (max_retries,)
    ).fetchall()]
    conn.close()

    if not pending:
        return {"sent": 0, "failed": 0}

    # Group by (sender_email, smtp_host, smtp_port, sender_password)
    from collections import defaultdict
    groups = defaultdict(list)
    for item in pending:
        key = (item["sender_email"], item["smtp_host"], item["smtp_port"], item.get("sender_password",""))
        groups[key].append(item)

    sent = failed = 0
    conn = get_db()

    for (sender_email, smtp_host, smtp_port, sender_password_enc), items in groups.items():
        pw = decrypt_str(sender_password_enc) if sender_password_enc else ""
        smtp_conn = None
        try:
            if smtp_port == 465:
                smtp_conn = smtplib.SMTP_SSL(smtp_host, 465, timeout=20)
            else:
                smtp_conn = smtplib.SMTP(smtp_host, smtp_port, timeout=20)
                smtp_conn.starttls()
            smtp_conn.login(sender_email, pw)

            for item in items:
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = item["subject"] or ""
                    msg["From"]    = sender_email
                    msg["To"]      = item["to_email"]
                    recipients     = [item["to_email"]]
                    if item.get("cc_email"):
                        for cc in item["cc_email"].split(","):
                            cc = cc.strip()
                            if cc:
                                msg["Cc"] = cc
                                recipients.append(cc)
                    msg.attach(MIMEText(item["body"] or "", "plain"))
                    smtp_conn.sendmail(sender_email, recipients, msg.as_string())
                    conn.execute(
                        "UPDATE email_queue SET status='sent', last_attempt=? WHERE id=?",
                        (now_ist().strftime("%Y-%m-%d %H:%M:%S"), item["id"])
                    )
                    sent += 1
                except Exception as e:
                    attempts = item["attempts"] + 1
                    new_status = "failed" if attempts >= max_retries else "pending"
                    conn.execute(
                        "UPDATE email_queue SET status=?, attempts=?, last_attempt=?, error_msg=? WHERE id=?",
                        (new_status, attempts, now_ist().strftime("%Y-%m-%d %H:%M:%S"),
                         str(e)[:300], item["id"])
                    )
                    failed += 1
        except Exception as conn_err:
            # Connection-level failure — mark all items in group as failed
            log.error("SMTP connection failed for %s: %s", sender_email, conn_err)
            for item in items:
                attempts = item["attempts"] + 1
                new_status = "failed" if attempts >= max_retries else "pending"
                conn.execute(
                    "UPDATE email_queue SET status=?, attempts=?, last_attempt=?, error_msg=? WHERE id=?",
                    (new_status, attempts, now_ist().strftime("%Y-%m-%d %H:%M:%S"),
                     str(conn_err)[:300], item["id"])
                )
                failed += 1
        finally:
            if smtp_conn:
                try:
                    smtp_conn.quit()
                except Exception:
                    pass

    conn.commit()
    conn.close()
    return {"sent": sent, "failed": failed}


def queue_webhook(event_type: str, payload: dict, board_id: int = None):
    """Insert a webhook call into webhook_queue for async batch processing.
    Never blocks the caller — the scheduler drains the queue periodically."""
    url = get_app_setting("webhook_url", "").strip()
    if not url:
        return
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO webhook_queue (queued_at, event_type, payload, url, status, board_id) "
            "VALUES (?,?,?,?,?,?)",
            (now_ist().strftime("%Y-%m-%d %H:%M:%S"), event_type,
             json.dumps(payload), url, "pending", board_id)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        log.warning("queue_webhook failed: %s", e)


def _drain_webhook_queue(batch_size: int = 50):
    """Process pending webhooks in batches. Called by the scheduler, never by a web request."""
    conn = get_db()
    pending = conn.execute(
        "SELECT id, url, event_type, payload FROM webhook_queue WHERE status='pending' "
        "ORDER BY queued_at ASC LIMIT ?", (batch_size,)
    ).fetchall()
    sent = failed = 0
    for row in pending:
        try:
            data = json.dumps({"event": row["event_type"], "data": json.loads(row["payload"])}).encode()
            req = urllib.request.Request(row["url"], data=data,
                                         headers={"Content-Type": "application/json"},
                                         method="POST")
            urllib.request.urlopen(req, timeout=10)
            conn.execute("UPDATE webhook_queue SET status='sent', sent_at=? WHERE id=?",
                         (now_ist().strftime("%Y-%m-%d %H:%M:%S"), row["id"]))
            sent += 1
        except Exception as e:
            conn.execute("UPDATE webhook_queue SET status='failed', sent_at=?, error=? WHERE id=?",
                         (now_ist().strftime("%Y-%m-%d %H:%M:%S"), str(e)[:500], row["id"]))
            failed += 1
    conn.commit()
    conn.close()
    if sent or failed:
        log.info("Webhook drain: %d sent, %d failed", sent, failed)


# Legacy alias — kept so grep/search finds it; routes now use queue_webhook
fire_webhook = queue_webhook


# ── Indian public holidays (2025–2027) ──────────────────────────────────────
_HOLIDAYS = {
    date(2025, 1, 26), date(2025, 3, 17), date(2025, 4, 14), date(2025, 4, 18),
    date(2025, 5, 12), date(2025, 8, 15), date(2025, 10, 2), date(2025, 10, 20),
    date(2025, 11, 5), date(2025, 11, 15), date(2025, 12, 25),
    date(2026, 1, 26), date(2026, 3, 6),  date(2026, 4, 3),  date(2026, 4, 14),
    date(2026, 5, 1),  date(2026, 8, 15), date(2026, 10, 2), date(2026, 10, 9),
    date(2026, 11, 25), date(2026, 12, 25),
    date(2027, 1, 26), date(2027, 3, 25), date(2027, 4, 2),  date(2027, 4, 14),
    date(2027, 8, 15), date(2027, 10, 2), date(2027, 12, 25),
}


def _count_weekdays(d1: date, d2: date) -> int:
    """O(1) count of weekdays (Mon-Fri) in [d1, d2] inclusive."""
    if d1 > d2:
        return 0
    n = (d2 - d1).days + 1
    full_weeks, extra = divmod(n, 7)
    count = full_weeks * 5
    dow = d1.weekday()
    for i in range(extra):
        if (dow + i) % 7 < 5:
            count += 1
    return count


def working_days_elapsed(start: str, end_d=None, hold_days: int = 0, extra_holidays: set = None) -> int:
    """Business days elapsed between start and end (O(1) + O(H) algorithm).

    Returns the count of working days strictly after start up to and including
    end_d, minus hold_days. Weekends (Sat/Sun) and holidays are excluded.

    Complexity: O(1) for weekday math + O(H) for holiday scan, where H is the
    number of holidays (~30 for 3 years). For 10,000 cases this is ~300K simple
    comparisons instead of the millions of date increments the old while-loop did.
    """
    if end_d is None:
        end_d = now_ist().date()
    if not start:
        return 0
    try:
        if isinstance(start, str):
            start = start[:10]
        s = datetime.strptime(start, "%Y-%m-%d").date() if isinstance(start, str) else start
    except (ValueError, TypeError):
        return 0
    if s >= end_d:
        return 0
    all_hols = _HOLIDAYS if extra_holidays is None else _HOLIDAYS | extra_holidays
    weekdays = _count_weekdays(s, end_d)
    hol_on_weekday = sum(1 for hol in all_hols if s <= hol <= end_d and hol.weekday() < 5)
    return max(0, weekdays - 1 - hol_on_weekday - hold_days)


# ── Database helpers ─────────────────────────────────────────────────────────
# Module-level connection pool — survives across requests on a warm Vercel instance.
# min=1, max=3 keeps memory low; Vercel spins up multiple instances under load so
# the effective pool across all instances stays well within Postgres limits.
_db_pool: "psycopg2.pool.ThreadedConnectionPool | None" = None


def _pg_sql(sql: str) -> str:
    """Convert SQLite-style ? placeholders to PostgreSQL %s.
    Skips ? characters inside SQL string literals so that literal question
    marks in VALUES/LIKE clauses are never mis-converted."""
    result = []
    in_single = False
    esc = False
    for ch in sql:
        if esc:
            esc = False
            result.append(ch)
            continue
        if ch == '\\':
            esc = True
            result.append(ch)
            continue
        if ch == "'":
            in_single = not in_single
        elif ch == '?' and not in_single:
            result.append('%s')
            continue
        result.append(ch)
    return ''.join(result)


def _get_pool() -> "psycopg2.pool.ThreadedConnectionPool":
    global _db_pool
    if _db_pool is None or _db_pool.closed:
        if not DATABASE_URL:
            raise RuntimeError(
                "DATABASE_URL environment variable is not set. "
                "Add a Postgres database and set DATABASE_URL "
                "in the Render project dashboard under Environment → Environment Variables."
            )
        _db_pool = psycopg2.pool.ThreadedConnectionPool(1, 3, DATABASE_URL)
    return _db_pool


class DBConn:
    """Thin psycopg2 wrapper that mimics the sqlite3 connection API used throughout this app.

    Handles:
    - ``?`` → ``%s`` placeholder translation
    - ``INTEGER PRIMARY KEY AUTOINCREMENT`` → ``SERIAL PRIMARY KEY`` in DDL
    - ``executescript()`` by splitting on ``;`` and executing each statement
    - sqlite3-style ``conn.execute()`` that returns the cursor
    - Returns connections to the pool on close() instead of closing them
    """

    def __init__(self, pg_conn, pool=None):
        self._conn = pg_conn
        self._pool = pool
        self._cur = pg_conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    def execute(self, sql, params=()):
        pg_sql = _pg_sql(sql)
        if params:
            self._cur.execute(pg_sql, params)
        else:
            self._cur.execute(pg_sql)
        return self._cur

    def executescript(self, sql):
        for stmt in sql.split(";"):
            stmt = stmt.strip()
            if not stmt:
                continue
            pg_stmt = (
                stmt
                .replace("?", "%s")
                .replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
            )
            self._cur.execute(pg_stmt)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        try:
            self._cur.close()
        except Exception:
            pass  # Cursor may be in error state — safe to ignore
        if self._pool:
            try:
                self._conn.rollback()
                self._pool.putconn(self._conn)  # Return healthy connection to pool
            except Exception:
                # Connection is dead (e.g. Render idle timeout killed it).
                # Discard it so the pool can create a fresh one on next getconn().
                try:
                    self._pool.putconn(self._conn, close=True)
                except Exception:
                    pass  # Last resort — give up cleanly
        else:
            try:
                self._conn.close()
            except Exception:
                pass


def get_db() -> DBConn:
    pool = _get_pool()
    pg_conn = pool.getconn()
    pg_conn.autocommit = False
    return DBConn(pg_conn, pool)


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS boards (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            board_name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS programmes (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            programme_name TEXT NOT NULL UNIQUE,
            board_id       INTEGER NOT NULL REFERENCES boards(id)
        );
        CREATE TABLE IF NOT EXISTS programme_config (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            programme_name       TEXT NOT NULL,
            stage_name           TEXT NOT NULL,
            stage_order          INTEGER NOT NULL DEFAULT 0,
            tat_days             INTEGER NOT NULL DEFAULT 0,
            reminder1_day        INTEGER NOT NULL DEFAULT 0,
            reminder2_day        INTEGER NOT NULL DEFAULT 0,
            owner_type           TEXT,
            overdue_interval_days INTEGER NOT NULL DEFAULT 3,
            is_milestone         INTEGER NOT NULL DEFAULT 0,
            is_optional          INTEGER NOT NULL DEFAULT 0,
            sender_email         TEXT,
            sender_password      TEXT,
            smtp_host            TEXT NOT NULL DEFAULT 'smtp.gmail.com',
            smtp_port            INTEGER NOT NULL DEFAULT 587
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role          TEXT NOT NULL DEFAULT 'program_officer',
            full_name     TEXT,
            email         TEXT
        );
        CREATE TABLE IF NOT EXISTS case_tracking (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            application_id       TEXT NOT NULL UNIQUE,
            programme_name       TEXT NOT NULL,
            organisation_name    TEXT NOT NULL,
            current_stage        TEXT NOT NULL,
            stage_start_date     TEXT NOT NULL,
            tat_days             INTEGER NOT NULL DEFAULT 0,
            reminder1_day        INTEGER NOT NULL DEFAULT 0,
            reminder2_day        INTEGER NOT NULL DEFAULT 0,
            owner_type           TEXT,
            action_owner_name    TEXT,
            action_owner_email   TEXT,
            program_officer_email TEXT,
            r1_sent              INTEGER NOT NULL DEFAULT 0,
            r2_sent              INTEGER NOT NULL DEFAULT 0,
            overdue_sent         INTEGER NOT NULL DEFAULT 0,
            overdue_count        INTEGER NOT NULL DEFAULT 0,
            last_overdue_date    TEXT,
            is_milestone         INTEGER NOT NULL DEFAULT 0,
            iteration_count      INTEGER NOT NULL DEFAULT 1,
            escalation_tier      INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS email_templates (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            programme_name    TEXT NOT NULL,
            notification_type TEXT NOT NULL,
            subject_line      TEXT NOT NULL,
            email_body        TEXT NOT NULL,
            UNIQUE(programme_name, notification_type)
        );
        CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp       TEXT NOT NULL,
            application_id  TEXT,
            event_type      TEXT NOT NULL,
            detail          TEXT,
            user_name       TEXT,
            board_id        INTEGER
        );
        CREATE TABLE IF NOT EXISTS stage_history (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp       TEXT NOT NULL,
            application_id  TEXT NOT NULL,
            from_stage      TEXT,
            to_stage        TEXT NOT NULL,
            changed_by      TEXT,
            board_id        INTEGER
        );
        CREATE TABLE IF NOT EXISTS email_queue (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            queued_at       TEXT NOT NULL,
            application_id  TEXT,
            programme_name  TEXT,
            notification_type TEXT NOT NULL,
            to_email        TEXT NOT NULL,
            cc_email        TEXT,
            sender_email    TEXT,
            sender_password TEXT,
            smtp_host       TEXT NOT NULL DEFAULT 'smtp.gmail.com',
            smtp_port       INTEGER NOT NULL DEFAULT 587,
            subject         TEXT,
            body            TEXT,
            status          TEXT NOT NULL DEFAULT 'pending',
            attempts        INTEGER NOT NULL DEFAULT 0,
            last_attempt    TEXT,
            error_msg       TEXT,
            board_id        INTEGER
        );
        CREATE TABLE IF NOT EXISTS scheduler_locks (
            lock_name       TEXT PRIMARY KEY,
            locked_at       TEXT NOT NULL,
            worker_id       TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS holidays (
            id           SERIAL PRIMARY KEY,
            holiday_date TEXT NOT NULL,
            name         TEXT NOT NULL,
            board_id     INTEGER REFERENCES boards(id),
            UNIQUE(holiday_date, board_id)
        );
        CREATE TABLE IF NOT EXISTS user_programme_map (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            programme_id INTEGER NOT NULL REFERENCES programmes(id) ON DELETE CASCADE,
            UNIQUE(user_id, programme_id)
        );
        CREATE TABLE IF NOT EXISTS saved_filters (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            filter_name TEXT NOT NULL,
            filter_json TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(user_id, filter_name)
        );
        CREATE TABLE IF NOT EXISTS stage_email_override (
            id                SERIAL PRIMARY KEY,
            programme_name    TEXT NOT NULL,
            stage_name        TEXT NOT NULL,
            notification_type TEXT NOT NULL,
            subject_line      TEXT NOT NULL,
            email_body        TEXT NOT NULL,
            UNIQUE(programme_name, stage_name, notification_type)
        );
        CREATE TABLE IF NOT EXISTS api_keys (
            id         SERIAL PRIMARY KEY,
            key_hash   TEXT NOT NULL UNIQUE,
            key_prefix TEXT,
            name       TEXT NOT NULL,
            board_id   INTEGER REFERENCES boards(id),
            created_at TEXT NOT NULL,
            last_used  TEXT,
            is_active  INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS escalation_matrix (
            id               SERIAL PRIMARY KEY,
            board_id         INTEGER REFERENCES boards(id),
            days_overdue_min INTEGER NOT NULL DEFAULT 1,
            days_overdue_max INTEGER,
            notify_role      TEXT NOT NULL DEFAULT 'program_head',
            UNIQUE(board_id, days_overdue_min)
        );
        CREATE TABLE IF NOT EXISTS webhook_queue (
            id         SERIAL PRIMARY KEY,
            queued_at  TEXT NOT NULL,
            event_type TEXT NOT NULL,
            payload    TEXT NOT NULL DEFAULT '{}',
            url        TEXT NOT NULL,
            status     TEXT NOT NULL DEFAULT 'pending',
            sent_at    TEXT,
            error      TEXT,
            board_id   INTEGER REFERENCES boards(id)
        );
    """)
    # Migrate existing DBs: add columns if absent (IF NOT EXISTS is safe to re-run)
    for sql in [
        "ALTER TABLE programme_config ADD COLUMN IF NOT EXISTS smtp_host TEXT NOT NULL DEFAULT 'smtp.gmail.com'",
        "ALTER TABLE programme_config ADD COLUMN IF NOT EXISTS smtp_port INTEGER NOT NULL DEFAULT 587",
        "ALTER TABLE programme_config ADD COLUMN IF NOT EXISTS board_id INTEGER",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS board_id INTEGER",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS force_password_reset INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS totp_secret TEXT",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login TEXT",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS board_id INTEGER",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS cc_emails TEXT",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS suppress_until TEXT",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'Active'",
        "ALTER TABLE email_templates ADD COLUMN IF NOT EXISTS board_id INTEGER",
        "ALTER TABLE holidays ADD COLUMN IF NOT EXISTS board_id INTEGER",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS hold_days INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS hold_start_date TEXT",
        "ALTER TABLE audit_log ADD COLUMN IF NOT EXISTS entry_hash TEXT",
        "ALTER TABLE programme_config ADD COLUMN IF NOT EXISTS is_optional INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE programmes ADD COLUMN IF NOT EXISTS tat_days INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE programmes ADD COLUMN IF NOT EXISTS reminder1_days INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE programmes ADD COLUMN IF NOT EXISTS reminder2_days INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE programmes ADD COLUMN IF NOT EXISTS overdue_days INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE programmes ADD COLUMN IF NOT EXISTS notification_emails TEXT",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS escalation_sent INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE api_keys ADD COLUMN IF NOT EXISTS key_prefix TEXT",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS iteration_count INTEGER NOT NULL DEFAULT 1",
        "ALTER TABLE case_tracking ADD COLUMN IF NOT EXISTS escalation_tier INTEGER NOT NULL DEFAULT 0",
    ]:
        try:
            conn.execute(sql)
        except Exception:
            conn.rollback()
    conn.commit()
    conn.close()


# ── Auth helpers ──────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "info")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "info")
            return redirect(url_for("login"))
        if session.get("role") != "super_admin":
            flash("This page requires Super Admin access.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def board_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "info")
            return redirect(url_for("login"))
        if session.get("role") not in ("super_admin", "board_admin"):
            flash("This page requires Board Admin access.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


# ── CSRF protection ───────────────────────────────────────────────────────────
_CSRF_EXEMPT_ENDPOINTS = {
    "login", "logout", "run_check", "run_weekly",
    # AJAX fetch() POSTs — no form body so csrf_token can't be injected by the DOM snippet;
    # these routes all require @login_required so session auth still protects them.
    "save_filter", "delete_filter", "test_smtp",
}
_CSRF_EXEMPT_PREFIXES  = ("/api/",)

@app.before_request
def _csrf_protect():
    """Validate CSRF token on every POST that is not an API or login endpoint."""
    if request.method != "POST":
        return
    endpoint = request.endpoint or ""
    if endpoint in _CSRF_EXEMPT_ENDPOINTS:
        return
    for prefix in _CSRF_EXEMPT_PREFIXES:
        if request.path.startswith(prefix):
            return
    token = request.form.get("csrf_token")
    if token != session.get("csrf_token"):
        flash("Invalid request. Please try again.", "error")
        referrer = request.referrer or url_for("dashboard")
        return redirect(referrer)


# ── Seed data ────────────────────────────────────────────────────────────────
_SEED_STAGES = [
    ("NABH Full Accreditation Hospitals", "Application In Progress", 1, 30, 15, 25, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "Application Fee Pending", 2, 10, 5, 8, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "Application Fee Paid", 3, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "DR Allocated", 4, 3, 1, 2, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "DR In Progress", 5, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "DR NC Response", 6, 20, 10, 15, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "DR NC Review", 7, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "DR Approval by NABH", 8, 3, 1, 2, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "Assessment selection (if applicable)", 9, 5, 2, 4, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA Allocation", 10, 20, 10, 15, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA date Accepted by Assessor", 11, 7, 3, 5, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA Scheduled", 12, 5, 2, 4, "Applicant", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA completed", 13, 3, 1, 2, "Assessor", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA Feedback", 14, 2, 1, 999, "Applicant", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA NC Response 1", 15, 50, 30, 40, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA NC Review 1", 16, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA NC Response 2", 17, 30, 10, 20, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "PA NC Review 2", 18, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "1st Annual Fee Pending", 19, 10, 5, 8, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "1st Annual Fee Paid", 20, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "OA Allocated", 21, 20, 10, 15, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA date Accepted by Assessor", 22, 7, 3, 5, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA Scheduled", 23, 5, 3, 4, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA completed", 24, 3, 1, 2, "Assessor", 2, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA Feedback", 25, 2, 1, 999, "Applicant", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA NC Response 1", 26, 50, 30, 40, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA NC Review 1", 27, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA NC Response 2", 28, 30, 10, 20, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA NC Review 2", 29, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "OA NC Accepted", 30, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "AC Allocated", 31, 15, 7, 10, "Program Officer", 2, 0, 0),
    ("NABH Full Accreditation Hospitals", "AC Document Pending", 32, 90, 30, 60, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "Pending Document Submitted", 33, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "Accredited/ Accredited Renewed", 34, 30, 15, 20, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "2nd Annual Fee Payment", 35, 90, 1, 10, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "2nd Annual Fee Paid", 36, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "SA due", 37, 30, 15, 20, "Program Officer", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA Allocated", 38, 20, 10, 15, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA date Accepted by Assessor", 39, 7, 3, 5, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA Scheduled", 40, 5, 3, 4, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA completed", 41, 3, 1, 2, "Assessor", 2, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA Feedback", 42, 2, 1, 999, "Applicant", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA NC Response 1", 43, 25, 15, 20, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA NC Review 1", 44, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA NC Response 2", 45, 15, 7, 10, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA NC Review 2", 46, 10, 5, 7, "Assessor", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "SA - OA NC Accepted", 47, 0, 0, 0, None, 0, 1, 0),
    ("NABH Full Accreditation Hospitals", "SA - AC Allocated", 48, 15, 7, 10, "Program Officer", 2, 0, 0),
    ("NABH Full Accreditation Hospitals", "Accredited Continued", 49, 30, 15, 20, "Program Officer", 1, 0, 0),
    ("NABH Full Accreditation Hospitals", "3rd Annual Fee Payment", 50, 90, 1, 10, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "4th Annual Fee Payment", 51, 90, 1, 10, "Applicant", 3, 0, 0),
    ("NABH Full Accreditation Hospitals", "Renewal", 52, 90, 30, 60, "Applicant", 3, 0, 0),
]

_SEED_TEMPLATES = [
    ("NABH Full Accreditation Hospitals", "R1",
     "Reminder: {{Stage_Name}} action pending — {{Organisation_Name}}",
     "Dear {{Action_Owner_Name}},\n\nThis is a reminder that your action on the stage '{{Stage_Name}}' "
     "under the {{Programme_Name}} programme is pending.\n\nCase: {{Organisation_Name}}\n"
     "Stage start date: {{Stage_Start_Date}}\nDeadline: {{TAT_Days}} working days from stage start\n"
     "Days remaining: {{Days_Remaining}}\n\nPlease log in to the portal and complete the required action "
     "at your earliest convenience.\n\nRegards,\n{{PO_Name}}\nQuality Council of India"),
    ("NABH Full Accreditation Hospitals", "R2",
     "Action Required: {{Stage_Name}} deadline approaching — {{Organisation_Name}}",
     "Dear {{Action_Owner_Name}},\n\nThe deadline for '{{Stage_Name}}' under {{Programme_Name}} is approaching.\n\n"
     "Case: {{Organisation_Name}}\nDays remaining: {{Days_Remaining}}\n\n"
     "Immediate action is required to avoid a breach of TAT. Please log in to the portal and complete the required action.\n\n"
     "Regards,\n{{PO_Name}}\nQuality Council of India"),
    ("NABH Full Accreditation Hospitals", "Overdue",
     "OVERDUE: {{Stage_Name}} TAT breached — {{Organisation_Name}}",
     "Dear {{Action_Owner_Name}},\n\nThe TAT for '{{Stage_Name}}' under {{Programme_Name}} has been breached.\n\n"
     "Case: {{Organisation_Name}}\nStage start date: {{Stage_Start_Date}}\nTAT: {{TAT_Days}} working days\n\n"
     "This is an urgent matter. Please take immediate action on the portal.\n\n"
     "Regards,\n{{PO_Name}}\nQuality Council of India"),
    ("NABH Full Accreditation Hospitals", "Followup",
     "Follow-up {{Followup_Count}}: {{Stage_Name}} still pending — {{Organisation_Name}}",
     "Dear {{Action_Owner_Name}},\n\nThis is follow-up #{{Followup_Count}} regarding the overdue stage "
     "'{{Stage_Name}}' under {{Programme_Name}}.\n\nCase: {{Organisation_Name}}\n\n"
     "This stage has been overdue. Please take immediate action.\n\n"
     "Regards,\n{{PO_Name}}\nQuality Council of India"),
]


_SEED_BOARDS = ["NABH", "NABL", "NABCB", "NABET"]


def migrate_data():
    """One-time migration: seeds boards, backfills board_id, renames admin→super_admin."""
    conn = get_db()
    # 1. Seed boards
    for b in _SEED_BOARDS:
        try:
            conn.execute("INSERT INTO boards (board_name) VALUES (?)", (b,))
        except Exception:
            conn.rollback()  # PostgreSQL requires rollback after any failed statement
    conn.commit()

    # 2. Rename role admin → super_admin
    conn.execute("UPDATE users SET role='super_admin' WHERE role='admin'")
    conn.commit()

    # 3. Backfill programmes table from programme_config
    distinct_progs = conn.execute(
        "SELECT DISTINCT programme_name FROM programme_config"
    ).fetchall()
    for row in distinct_progs:
        pname = row[0]
        # Already in programmes table? Skip.
        if conn.execute("SELECT id FROM programmes WHERE programme_name=?", (pname,)).fetchone():
            continue
        # Infer board by prefix
        board_row = None
        for b in _SEED_BOARDS:
            if pname.upper().startswith(b):
                board_row = conn.execute("SELECT id FROM boards WHERE board_name=?", (b,)).fetchone()
                break
        if not board_row:
            board_row = conn.execute("SELECT id FROM boards ORDER BY id LIMIT 1").fetchone()
        if board_row:
            try:
                conn.execute("INSERT INTO programmes (programme_name, board_id) VALUES (?,?)",
                             (pname, board_row[0]))
            except Exception:
                conn.rollback()  # PostgreSQL requires rollback after any failed statement
    conn.commit()

    # 4. Backfill board_id on programme_config
    conn.execute("""
        UPDATE programme_config SET board_id = (
            SELECT board_id FROM programmes WHERE programmes.programme_name = programme_config.programme_name
        ) WHERE board_id IS NULL
    """)
    conn.commit()

    # 5. Backfill board_id on case_tracking
    conn.execute("""
        UPDATE case_tracking SET board_id = (
            SELECT board_id FROM programme_config
            WHERE programme_config.programme_name = case_tracking.programme_name LIMIT 1
        ) WHERE board_id IS NULL
    """)
    conn.commit()

    # 6. Backfill board_id on email_templates
    conn.execute("""
        UPDATE email_templates SET board_id = (
            SELECT board_id FROM programmes
            WHERE programmes.programme_name = email_templates.programme_name
        ) WHERE board_id IS NULL
    """)
    conn.commit()

    # 7. Default officer user → NABH board if no board assigned
    nabh = conn.execute("SELECT id FROM boards WHERE board_name='NABH'").fetchone()
    if nabh:
        conn.execute(
            "UPDATE users SET board_id=? WHERE role='program_officer' AND board_id IS NULL",
            (nabh[0],)
        )
    conn.commit()
    conn.close()


def seed_data():
    conn = get_db()
    # Always ensure admin exists; if ADMIN_PASSWORD env var is set, force-reset password
    _admin_pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing_admin = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()
    if not existing_admin:
        conn.execute(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
            ("admin", generate_password_hash(_admin_pw), "super_admin", "Super Administrator"),
        )
        conn.commit()
        log.info("seed_data: admin user created")
    elif os.environ.get("ADMIN_PASSWORD"):
        # Force-reset password if env var explicitly set
        conn.execute(
            "UPDATE users SET password_hash=? WHERE username='admin'",
            (generate_password_hash(_admin_pw),)
        )
        conn.commit()
        log.info("seed_data: admin password reset from ADMIN_PASSWORD env var")

    if conn.execute("SELECT COUNT(*) FROM users WHERE username='officer'").fetchone()[0] == 0:
        nabh = conn.execute("SELECT id FROM boards WHERE board_name='NABH'").fetchone()
        conn.execute(
            "INSERT INTO users (username, password_hash, role, full_name, board_id) VALUES (?,?,?,?,?)",
            ("officer", generate_password_hash("po123"), "program_officer", "Program Officer",
             nabh[0] if nabh else None),
        )
        conn.commit()

    if conn.execute("SELECT COUNT(*) FROM programme_config").fetchone()[0] == 0:
        nabh = conn.execute("SELECT id FROM boards WHERE board_name='NABH'").fetchone()
        nabh_id = nabh[0] if nabh else None
        # Insert programme record
        try:
            conn.execute("INSERT INTO programmes (programme_name, board_id) VALUES (?,?)",
                         ("NABH Full Accreditation Hospitals", nabh_id))
        except Exception:
            conn.rollback()  # PostgreSQL requires rollback after any failed statement
        for row in _SEED_STAGES:
            conn.execute(
                "INSERT INTO programme_config "
                "(programme_name,stage_name,stage_order,tat_days,reminder1_day,reminder2_day,"
                "owner_type,overdue_interval_days,is_milestone,is_optional,board_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (*row, nabh_id),
            )
    if conn.execute("SELECT COUNT(*) FROM email_templates").fetchone()[0] == 0:
        nabh = conn.execute("SELECT id FROM boards WHERE board_name='NABH'").fetchone()
        nabh_id = nabh[0] if nabh else None
        for row in _SEED_TEMPLATES:
            conn.execute(
                "INSERT INTO email_templates (programme_name,notification_type,subject_line,email_body,board_id) VALUES (?,?,?,?,?)",
                (*row, nabh_id),
            )
    conn.commit()
    try:
        _all_boards = conn.execute("SELECT id FROM boards").fetchall()
        for _br in _all_boards:
            _bid = _br[0]
            for _min, _max, _role in [
                (1, 7, "program_head"),
                (8, 30, "board_ceo"),
                (31, None, "board_admin"),
            ]:
                conn.execute(
                    "INSERT INTO escalation_matrix (board_id, days_overdue_min, days_overdue_max, notify_role) "
                    "VALUES (?,?,?,?) ON CONFLICT (board_id, days_overdue_min) DO NOTHING",
                    (_bid, _min, _max, _role)
                )
        conn.commit()
    except Exception as _e:
        conn.rollback()
        log.warning("Escalation matrix seed failed: %s", _e)
    conn.close()


# ── Role constants ────────────────────────────────────────────────────────────
ROLE_META = {
    "super_admin":    ("Super Admin",    "#dbeafe", "#1d4ed8"),
    "board_admin":    ("Board Admin",    "#ede9fe", "#7c3aed"),
    "board_ceo":      ("Board CEO",      "#fef3c7", "#92400e"),
    "program_head":   ("Programme Head", "#dcfce7", "#166534"),
    "program_officer":("Program Officer","#f1f5f9", "#475569"),
}

# ── Board / programme scoping helpers ─────────────────────────────────────────
def user_board_id():
    """Returns board_id from session, or None for super_admin (sees all)."""
    if session.get("role") == "super_admin":
        return None
    return session.get("board_id")


def user_programme_names():
    """Returns list of programme names for program_head; None means 'see all in scope'."""
    if session.get("role") != "program_head":
        return None
    conn = get_db()
    rows = conn.execute(
        """SELECT p.programme_name FROM user_programme_map upm
           JOIN programmes p ON p.id = upm.programme_id
           WHERE upm.user_id=?""",
        (session["user_id"],)
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


# ── Email helpers ─────────────────────────────────────────────────────────────
_DEFAULT_SUBJECTS = {
    "R1":      "Reminder: {{Stage_Name}} action pending — {{Organisation_Name}}",
    "R2":      "Action Required: {{Stage_Name}} deadline approaching — {{Organisation_Name}}",
    "Overdue": "OVERDUE: {{Stage_Name}} TAT breached — {{Organisation_Name}}",
    "Followup":"Follow-up {{Followup_Count}}: {{Stage_Name}} still pending — {{Organisation_Name}}",
}
_DEFAULT_BODIES = {
    "R1":      "Dear {{Action_Owner_Name}},\n\nReminder for stage {{Stage_Name}} under {{Programme_Name}}.\nOrg: {{Organisation_Name}}\nDays remaining: {{Days_Remaining}}\n\nRegards,\n{{PO_Name}}\nQCI",
    "R2":      "Dear {{Action_Owner_Name}},\n\nDeadline approaching for {{Stage_Name}}.\nOrg: {{Organisation_Name}}\nDays remaining: {{Days_Remaining}}\n\nRegards,\n{{PO_Name}}\nQCI",
    "Overdue": "Dear {{Action_Owner_Name}},\n\nTAT breached for {{Stage_Name}}.\nOrg: {{Organisation_Name}}\n\nRegards,\n{{PO_Name}}\nQCI",
    "Followup":"Dear {{Action_Owner_Name}},\n\nFollow-up #{{Followup_Count}} for overdue stage {{Stage_Name}}.\nOrg: {{Organisation_Name}}\n\nRegards,\n{{PO_Name}}\nQCI",
}


def _fill(text: str, ph: dict) -> str:
    for k, v in ph.items():
        text = text.replace("{{" + k + "}}", str(v) if v is not None else "")
    return text


def _get_template(programme: str, ntype: str):
    conn = get_db()
    row = conn.execute(
        "SELECT subject_line, email_body FROM email_templates WHERE programme_name=? AND notification_type=?",
        (programme, ntype),
    ).fetchone()
    conn.close()
    return row


def send_notification(programme: str, ntype: str, to_email: str, cc_email: str,
                      sender_email: str, sender_password: str, ph: dict,
                      smtp_host: str = "smtp.gmail.com", smtp_port: int = 587):
    tmpl = _get_template(programme, ntype)
    subject = _fill(tmpl["subject_line"] if tmpl else _DEFAULT_SUBJECTS[ntype], ph)
    body    = _fill(tmpl["email_body"]    if tmpl else _DEFAULT_BODIES[ntype],    ph)

    if not sender_email or not to_email:
        return False, "Missing sender or recipient email"
    try:
        msg = MIMEMultipart()
        msg["From"]    = sender_email
        msg["To"]      = to_email
        msg["CC"]      = cc_email or ""
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        recipients = [to_email] + ([cc_email] if cc_email else [])
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_host, 465, timeout=15) as s:
                s.login(sender_email, sender_password)
                s.sendmail(sender_email, recipients, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as s:
                s.starttls()
                s.login(sender_email, sender_password)
                s.sendmail(sender_email, recipients, msg.as_string())
        log_audit("email_sent", ph.get("Programme_Name", ""),
                  f"{ntype} to {to_email} | Subject: {subject[:80]}", "system")
        return True, ""
    except Exception as e:
        log_audit("email_error", ph.get("Programme_Name", ""),
                  f"{ntype} to {to_email} | Error: {str(e)[:200]}", "system")
        return False, str(e)


# ── Core daily check ─────────────────────────────────────────────────────────
def run_daily_check(board_id=None) -> dict:
    today = now_ist().date()
    conn  = get_db()
    if board_id is not None:
        cases = [dict(r) for r in conn.execute(
            "SELECT * FROM case_tracking WHERE board_id=?", (board_id,)
        ).fetchall()]
    else:
        cases = [dict(r) for r in conn.execute("SELECT * FROM case_tracking").fetchall()]
    summary = {"r1": 0, "r2": 0, "overdue": 0, "followup": 0,
               "skipped_milestone": 0, "errors": []}

    # Pre-fetch programme_config to avoid N+1 queries per case
    _pc_rows = conn.execute(
        "SELECT programme_name, stage_name, overdue_interval_days, "
        "sender_email, sender_password, smtp_host, smtp_port "
        "FROM programme_config"
    ).fetchall()
    _pc_lookup = {(r["programme_name"], r["stage_name"]): dict(r) for r in _pc_rows}

    # Pre-fetch programme notification_emails for CC injection
    try:
        _prog_notif_rows = conn.execute(
            "SELECT programme_name, notification_emails FROM programmes WHERE notification_emails IS NOT NULL"
        ).fetchall()
        _prog_notif_map = {r["programme_name"]: r["notification_emails"] for r in _prog_notif_rows}
    except Exception:
        _prog_notif_map = {}

    # Pre-fetch programme → programme_head email mapping for escalations
    _ph_rows = conn.execute(
        """SELECT p.programme_name, u.email, u.full_name
           FROM programmes p
           JOIN user_programme_map upm ON upm.programme_id = p.id
           JOIN users u ON u.id = upm.user_id
           WHERE u.role='program_head' AND u.email IS NOT NULL"""
    ).fetchall()
    _ph_map: dict = {}
    for _r in _ph_rows:
        _ph_map.setdefault(_r["programme_name"], []).append(
            {"email": _r["email"], "full_name": _r["full_name"]}
        )

    # Pre-fetch custom DB holidays once for the whole loop
    try:
        _db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                    for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _db_hols = set()

    for case in cases:
        # Skip closed/withdrawn/suspended/on-hold cases
        case_status = case.get("status", "Active")
        if case_status and case_status not in ("Active", None, ""):
            continue
        if case["is_milestone"]:
            summary["skipped_milestone"] += 1
            continue

        elapsed = working_days_elapsed(case["stage_start_date"], today, hold_days=case.get("hold_days", 0), extra_holidays=_db_hols)
        tat     = case["tat_days"]
        r1_day  = case["reminder1_day"]
        r2_day  = case["reminder2_day"]

        cfg = _pc_lookup.get((case["programme_name"], case["current_stage"]))

        overdue_interval = cfg["overdue_interval_days"] if cfg else 3
        sender_email     = cfg["sender_email"] if cfg and cfg["sender_email"] else None
        sender_password  = decrypt_str(cfg["sender_password"]) if cfg and cfg["sender_password"] else ""
        smtp_host        = cfg["smtp_host"] if cfg and cfg["smtp_host"] else "smtp.gmail.com"
        smtp_port        = cfg["smtp_port"] if cfg and cfg["smtp_port"] else 587

        days_remaining = max(0, tat - elapsed)
        ph = {
            "Organisation_Name": case["organisation_name"],
            "Stage_Name":        case["current_stage"],
            "Action_Owner_Name": case["action_owner_name"],
            "Days_Remaining":    days_remaining,
            "TAT_Days":          tat,
            "Stage_Start_Date":  case["stage_start_date"],
            "Programme_Name":    case["programme_name"],
            "Followup_Count":    case["overdue_count"] + 1,
            "PO_Name":           case["program_officer_email"] or "Program Officer",
        }

        # Check suppress_until
        suppress = case.get("suppress_until")
        if suppress:
            try:
                if date.fromisoformat(suppress) >= today:
                    continue
            except ValueError:
                pass

        sender_pw_enc = cfg["sender_password"] if cfg and cfg["sender_password"] else ""
        # Merge case CC emails with programme-level notification_emails
        _prog_extra_cc = _prog_notif_map.get(case["programme_name"], "")
        cc = ";".join(filter(None, [
            case.get("cc_emails") or case.get("program_officer_email") or "",
            _prog_extra_cc
        ]))
        webhook_ph = {
            "application_id": case["application_id"],
            "organisation": case["organisation_name"],
            "programme": case["programme_name"],
            "stage": case["current_stage"],
        }

        def _send(ntype):
            if not sender_email or not case.get("action_owner_email"):
                summary["errors"].append(f"{case['application_id']} {ntype}: missing sender/recipient")
                return False
            queue_email(
                case["programme_name"], ntype,
                case["action_owner_email"], cc,
                sender_email, sender_pw_enc, ph,
                smtp_host, smtp_port,
                case["application_id"], case.get("board_id"),
                stage_name=case["current_stage"]
            )
            fire_webhook(f"notification_{ntype.lower()}", {**webhook_ph, "type": ntype})
            return True

        def _escalate_to_ph():
            """Multi-tier escalation using escalation_matrix table."""
            days_overdue = elapsed - tat
            if days_overdue <= 0:
                return
            # Find applicable tier from matrix
            matrix_rows = conn.execute(
                "SELECT * FROM escalation_matrix WHERE board_id=? AND days_overdue_min <= ? "
                "AND (days_overdue_max IS NULL OR days_overdue_max >= ?) ORDER BY days_overdue_min DESC LIMIT 1",
                (case.get("board_id"), days_overdue, days_overdue)
            ).fetchone()
            if not matrix_rows:
                return
            tier_index = {"program_head": 1, "board_ceo": 2, "board_admin": 3}.get(matrix_rows["notify_role"], 1)
            if case.get("escalation_tier", 0) >= tier_index:
                return  # Already escalated at this or higher tier
            # Find users with the required role scoped to board
            notify_role = matrix_rows["notify_role"]
            target_users = conn.execute(
                "SELECT email, full_name FROM users WHERE role=? AND board_id=? AND email IS NOT NULL",
                (notify_role, case.get("board_id"))
            ).fetchall()
            queued = False
            for t_user in target_users:
                if not t_user["email"]:
                    continue
                esc_ph = {**ph,
                          "Action_Owner_Name": t_user["full_name"] or t_user["email"],
                          "Days_Overdue": days_overdue,
                          "Escalation_Role": notify_role}
                if sender_email:
                    queue_email(
                        case["programme_name"], "Escalation",
                        t_user["email"], cc,
                        sender_email, sender_pw_enc, esc_ph,
                        smtp_host, smtp_port,
                        case["application_id"], case.get("board_id"),
                        stage_name=case["current_stage"]
                    )
                    queued = True
            if queued:
                conn.execute(
                    "UPDATE case_tracking SET escalation_sent=1, escalation_tier=? WHERE id=?",
                    (tier_index, case["id"])
                )
                conn.commit()

        # R1
        if r1_day > 0 and elapsed >= r1_day and not case["r1_sent"]:
            _send("R1")
            conn.execute("UPDATE case_tracking SET r1_sent=1 WHERE id=?", (case["id"],))
            conn.commit()
            summary["r1"] += 1

        # R2
        if r2_day > 0 and elapsed >= r2_day and not case["r2_sent"]:
            _send("R2")
            conn.execute("UPDATE case_tracking SET r2_sent=1 WHERE id=?", (case["id"],))
            conn.commit()
            summary["r2"] += 1

        # First overdue
        if tat > 0 and elapsed >= tat and not case["overdue_sent"]:
            _send("OVERDUE")
            conn.execute(
                "UPDATE case_tracking SET overdue_sent=1, last_overdue_date=? WHERE id=?",
                (today.isoformat(), case["id"]),
            )
            conn.commit()
            summary["overdue"] += 1

        # Followup overdue
        elif tat > 0 and elapsed > tat and case["overdue_sent"] and case["last_overdue_date"]:
            last = datetime.strptime(case["last_overdue_date"], "%Y-%m-%d").date()
            if (today - last).days >= overdue_interval:
                ph["Followup_Count"] = case["overdue_count"] + 1
                _send("FOLLOWUP")
                conn.execute(
                    "UPDATE case_tracking SET overdue_count=overdue_count+1, last_overdue_date=? WHERE id=?",
                    (today.isoformat(), case["id"]),
                )
                conn.commit()
                summary["followup"] += 1

        # Escalation to Programme Head if severely overdue
        if tat > 0 and elapsed > tat and case["overdue_sent"]:
            _escalate_to_ph()

    conn.close()
    # Process queued emails
    q_result = process_email_queue()
    summary["emails_sent"] = q_result["sent"]
    summary["email_failures"] = q_result["failed"]
    return summary


# ── Weekly digest ────────────────────────────────────────────────────────────
def run_weekly_digest():
    """Send a weekly summary email to board_ceo and board_admin users."""
    if get_app_setting("digest_enabled", "1") != "1":
        log.info("Weekly digest is disabled — skipping.")
        return
    conn = get_db()
    today = now_ist().date()
    # Pre-fetch custom DB holidays once for the whole digest
    try:
        _digest_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                           for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _digest_db_hols = set()
    recipients = [dict(r) for r in conn.execute(
        "SELECT u.*, b.board_name FROM users u LEFT JOIN boards b ON b.id=u.board_id "
        "WHERE u.role IN ('board_ceo','board_admin','super_admin') AND u.email IS NOT NULL"
    ).fetchall()]

    for user in recipients:
        bid = user["board_id"]
        if user["role"] == "super_admin":
            cases = [dict(r) for r in conn.execute(
                "SELECT * FROM case_tracking WHERE (status='Active' OR status IS NULL)"
            ).fetchall()]
        else:
            cases = [dict(r) for r in conn.execute(
                "SELECT * FROM case_tracking WHERE board_id=? AND (status='Active' OR status IS NULL)", (bid,)
            ).fetchall()]

        if not cases:
            continue

        total = len(cases)
        overdue = sum(1 for c in cases if not c["is_milestone"] and c["tat_days"] > 0
                      and working_days_elapsed(c["stage_start_date"], today, extra_holidays=_digest_db_hols) >= c["tat_days"])
        at_risk = sum(1 for c in cases if not c["is_milestone"] and c["tat_days"] > 0
                      and working_days_elapsed(c["stage_start_date"], today, extra_holidays=_digest_db_hols) >= c["reminder2_day"]
                      and working_days_elapsed(c["stage_start_date"], today, extra_holidays=_digest_db_hols) < c["tat_days"])
        compliant = total - overdue - at_risk

        body = (f"Weekly QCI Notification Engine Digest — {today.strftime('%d %b %Y')}\n\n"
                f"Board: {user.get('board_name') or 'All Boards'}\n"
                f"{'='*40}\n\n"
                f"Total Active Cases  : {total}\n"
                f"Overdue             : {overdue}\n"
                f"At Risk             : {at_risk}\n"
                f"On Track / Compliant: {compliant}\n\n")

        # Top 5 most overdue
        overdue_cases = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0
                         and working_days_elapsed(c["stage_start_date"], today, extra_holidays=_digest_db_hols) >= c["tat_days"]]
        overdue_cases.sort(key=lambda x: working_days_elapsed(x["stage_start_date"], today, extra_holidays=_digest_db_hols), reverse=True)
        if overdue_cases:
            body += "Top Overdue Cases:\n"
            for c in overdue_cases[:5]:
                days_late = working_days_elapsed(c["stage_start_date"], today, extra_holidays=_digest_db_hols) - c["tat_days"]
                body += f"  • {c['organisation_name']} ({c['application_id']}) — {c['current_stage']} — {days_late}d late\n"
            body += "\n"

        body += "Log in to QCI Notify for full details.\n\nQCI Notification Engine (automated digest)"

        # Get sender credentials from any programme in scope
        cfg = None
        if bid:
            cfg = conn.execute(
                "SELECT sender_email, sender_password, smtp_host, smtp_port "
                "FROM programme_config WHERE board_id=? AND sender_email IS NOT NULL LIMIT 1", (bid,)
            ).fetchone()
        if not cfg:
            cfg = conn.execute(
                "SELECT sender_email, sender_password, smtp_host, smtp_port "
                "FROM programme_config WHERE sender_email IS NOT NULL LIMIT 1"
            ).fetchone()
        if not cfg or not cfg["sender_email"]:
            continue

        try:
            msg = MIMEMultipart()
            msg["From"] = cfg["sender_email"]
            msg["To"] = user["email"]
            msg["Subject"] = f"[QCI Digest] Weekly Summary — {today.strftime('%d %b %Y')}"
            msg.attach(MIMEText(body, "plain"))
            pw = decrypt_str(cfg["sender_password"]) if cfg["sender_password"] else ""
            smtp_host = cfg["smtp_host"] or "smtp.gmail.com"
            smtp_port = cfg["smtp_port"] or 587
            with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as s:
                s.starttls()
                s.login(cfg["sender_email"], pw)
                s.sendmail(cfg["sender_email"], [user["email"]], msg.as_string())
            log.info("Weekly digest sent to %s", user["email"])
        except Exception as e:
            log.warning("Weekly digest failed for %s: %s", user["email"], e)

    # Clean up old audit and email queue records
    try:
        cutoff = (now_ist() - timedelta(days=90)).strftime("%Y-%m-%d")
        # Only purge high-volume low-value logs; retain structural events forever
        conn.execute(
            "DELETE FROM audit_log WHERE timestamp < ? AND event_type IN "
            "('email_sent','scheduled_check','run_check','bulk_advance')",
            (cutoff,)
        )
        conn.execute(
            "DELETE FROM email_queue WHERE queued_at < ? AND status IN ('sent','failed')",
            (cutoff,)
        )
        conn.execute(
            "DELETE FROM webhook_queue WHERE queued_at < ? AND status IN ('sent','failed')",
            (cutoff,)
        )
        conn.commit()
        log.info("Weekly digest cleanup: removed audit/email/webhook records older than %s", cutoff)
    except Exception as e:
        log.warning("Weekly digest cleanup failed: %s", e)

    conn.close()


# ── Case upsert helper ───────────────────────────────────────────────────────
def upsert_case(data: dict) -> str:
    """Insert or update a case. Returns 'created' or 'updated'."""
    conn = get_db()
    try:
        cfg = conn.execute(
            "SELECT tat_days,reminder1_day,reminder2_day,owner_type,is_milestone,overdue_interval_days,board_id,stage_order "
            "FROM programme_config WHERE programme_name=? AND stage_name=?",
            (data["programme_name"], data["stage_name"]),
        ).fetchone()
        if not cfg:
            raise ValueError(f"Stage '{data['stage_name']}' not found in programme '{data['programme_name']}'")

        existing = conn.execute(
            "SELECT id, current_stage FROM case_tracking WHERE application_id=?", (data["application_id"],)
        ).fetchone()

        board_id = cfg["board_id"]
        if existing:
            old_stage = existing["current_stage"]
            new_stage = data["stage_name"]
            _cur_order_row = conn.execute(
                "SELECT stage_order FROM programme_config WHERE programme_name=? AND stage_name=?",
                (data["programme_name"], existing["current_stage"])
            ).fetchone()
            _is_regression = bool(
                _cur_order_row and cfg.get("stage_order") is not None and
                _cur_order_row["stage_order"] is not None and
                cfg["stage_order"] < _cur_order_row["stage_order"]
            )
            conn.execute(
                """UPDATE case_tracking SET
                   programme_name=?, organisation_name=?, current_stage=?, stage_start_date=?,
                   tat_days=?, reminder1_day=?, reminder2_day=?, owner_type=?,
                   action_owner_name=?, action_owner_email=?, program_officer_email=?,
                   r1_sent=0, r2_sent=0, overdue_sent=0, overdue_count=0,
                   last_overdue_date=NULL, is_milestone=?, board_id=?,
                   cc_emails=?, suppress_until=?, escalation_sent=0, escalation_tier=0
                   WHERE application_id=?""",
                (data["programme_name"], data["organisation_name"], data["stage_name"],
                 data["stage_start_date"], cfg["tat_days"], cfg["reminder1_day"], cfg["reminder2_day"],
                 cfg["owner_type"], data["action_owner_name"], data["action_owner_email"],
                 data["program_officer_email"], cfg["is_milestone"], board_id,
                 data.get("cc_emails") or None, data.get("suppress_until") or None,
                 data["application_id"]),
            )
            conn.commit()
            if _is_regression:
                conn.execute(
                    "UPDATE case_tracking SET iteration_count = iteration_count + 1 WHERE application_id=?",
                    (data["application_id"],)
                )
                conn.commit()
            action = "updated"
        else:
            conn.execute(
                """INSERT INTO case_tracking
                   (application_id,programme_name,organisation_name,current_stage,stage_start_date,
                    tat_days,reminder1_day,reminder2_day,owner_type,
                    action_owner_name,action_owner_email,program_officer_email,is_milestone,board_id,
                    cc_emails,suppress_until)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (data["application_id"], data["programme_name"], data["organisation_name"],
                 data["stage_name"], data["stage_start_date"],
                 cfg["tat_days"], cfg["reminder1_day"], cfg["reminder2_day"], cfg["owner_type"],
                 data["action_owner_name"], data["action_owner_email"],
                 data["program_officer_email"], cfg["is_milestone"], board_id,
                 data.get("cc_emails") or None, data.get("suppress_until") or None),
            )
            conn.commit()
            action = "created"
    finally:
        conn.close()

    # Audit/history logging runs outside the connection (uses its own conn)
    if action == "updated":
        old_stage = existing["current_stage"]
        new_stage = data["stage_name"]
        if old_stage != new_stage:
            log_stage_transition(data["application_id"], old_stage, new_stage,
                                 data.get("_changed_by", ""), board_id)
            log_audit("stage_change", data["application_id"],
                      f"{old_stage} → {new_stage}", data.get("_changed_by", ""), board_id)
            if _is_regression:
                log_audit("stage_regression", data["application_id"],
                          f"Stage regressed: {existing['current_stage']} → {data['stage_name']}",
                          data.get("_changed_by", ""), board_id)
            for _skipped in data.get("_skipped_stages", []):
                log_audit("stage_skipped", data["application_id"],
                          f"Optional stage skipped: {_skipped}", data.get("_changed_by", ""), board_id)
        else:
            log_audit("case_updated", data["application_id"],
                      f"Updated (stage: {new_stage})", data.get("_changed_by", ""), board_id)
    else:
        log_stage_transition(data["application_id"], None, data["stage_name"],
                             data.get("_changed_by", ""), board_id)
        log_audit("case_created", data["application_id"],
                  f"New case: {data['stage_name']}", data.get("_changed_by", ""), board_id)
    return action


# ── HTML base template ────────────────────────────────────────────────────────
_BASE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>QCI Notification Engine</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<style>
:root{
  --navy:#003356;--navy-dark:#002240;--navy-light:#e1eef8;
  --accent:#0094ca;--accent-dark:#007baa;--accent-light:#d0f0ff;
  --qci-green:#00984C;--qci-green-light:#d8f5e7;
  --success:#00984C;--success-light:#d8f5e7;
  --warning:#d97706;--warning-light:#fef3c7;
  --danger:#dc2626;--danger-light:#fee2e2;
  --muted:#64748b;--border:#dce5ef;--bg:#f3f6fa;
}
*{font-family:'Inter',sans-serif}
body{background:var(--bg);margin:0}

/* ── Sidebar ── */
#sidebar{
  position:fixed;top:0;left:0;width:220px;height:100vh;
  background:var(--navy-dark);z-index:100;
  display:flex;flex-direction:column;overflow-y:auto;
}
#sidebar .brand{
  padding:20px 20px 16px;border-bottom:1px solid rgba(255,255,255,.08);
}
#sidebar .brand-logo{
  font-size:18px;font-weight:700;color:#fff;letter-spacing:.3px;
  display:flex;align-items:center;gap:8px;text-decoration:none;
}
#sidebar .brand-logo span{color:#11a3d4}
#sidebar .brand-sub{font-size:10px;color:rgba(255,255,255,.4);margin-top:2px;letter-spacing:.5px;text-transform:uppercase}
#sidebar nav{padding:12px 0;flex:1}
#sidebar .nav-section{
  font-size:10px;font-weight:600;color:rgba(255,255,255,.35);
  letter-spacing:1px;text-transform:uppercase;
  padding:12px 20px 4px;
}
#sidebar .nav-link{
  display:flex;align-items:center;gap:10px;
  padding:9px 20px;color:rgba(255,255,255,.7);
  font-size:13.5px;font-weight:500;border-radius:0;
  transition:all .15s;text-decoration:none;
}
#sidebar .nav-link:hover{background:rgba(255,255,255,.07);color:#fff}
#sidebar .nav-link.active{background:var(--accent);color:#fff;border-left:3px solid #fff}
#sidebar .nav-link i{font-size:15px;width:18px;text-align:center}
#sidebar .sidebar-footer{
  padding:14px 20px;border-top:1px solid rgba(255,255,255,.08);
  font-size:11px;color:rgba(255,255,255,.3);
}
#sidebar .sidebar-search{
  padding:8px 14px 0;
}
#sidebar .sidebar-search input{
  width:100%;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.12);
  border-radius:6px;padding:6px 10px;color:#fff;font-size:12px;outline:none;
}
#sidebar .sidebar-search input::placeholder{color:rgba(255,255,255,.35)}
#sidebar .sidebar-search input:focus{background:rgba(255,255,255,.14);border-color:rgba(255,255,255,.3)}

/* ── Dark mode ── */
.dark-mode{
  --bg:#0f172a;--border:#1e293b;--muted:#94a3b8;
}
.dark-mode body{background:var(--bg)}
.dark-mode .card,.dark-mode .card-header,.dark-mode .topbar{background:#1e293b;color:#e2e8f0;border-color:#334155}
.dark-mode .data-table th{background:#1a2942;color:#94a3b8;border-color:#334155}
.dark-mode .data-table td{border-color:#1e293b;color:#cbd5e1}
.dark-mode .data-table tbody tr:hover td{background:#1a2942}
.dark-mode .stat-card{background:#1e293b;border-color:#334155;color:#e2e8f0}
.dark-mode .stat-card .stat-val{color:#e2e8f0}
.dark-mode .form-control,.dark-mode .form-select{background:#0f172a;border-color:#334155;color:#e2e8f0}
.dark-mode .topbar-title{color:#e2e8f0}
.dark-mode .accordion-button{background:#1e293b;color:#e2e8f0}

/* ── Overdue banner ── */
.overdue-banner{
  background:linear-gradient(90deg,#7f1d1d,#dc2626);color:#fff;
  padding:10px 28px;font-size:13px;font-weight:500;
  display:flex;align-items:center;gap:12px;
}
.overdue-banner a{color:#fca5a5;text-decoration:underline}

/* ── Main layout ── */
#main{margin-left:220px;min-height:100vh}
.topbar{
  background:#fff;border-bottom:1px solid var(--border);
  padding:0 28px;height:56px;
  display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:50;
}
.topbar-title{font-size:15px;font-weight:600;color:var(--navy)}
.page-body{padding:24px 28px}

/* ── Cards ── */
.card{border:1px solid var(--border);border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.04)}
.card-header{background:#fff;border-bottom:1px solid var(--border);font-weight:600;font-size:14px;padding:14px 18px;border-radius:10px 10px 0 0 !important}

/* ── Stat cards ── */
.stat-card{border-radius:12px;padding:20px 22px;border:1px solid var(--border);background:#fff;transition:box-shadow .15s}
.stat-card:hover{box-shadow:0 4px 16px rgba(0,0,0,.08)}
.stat-card .stat-icon{width:46px;height:46px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px}
.stat-card .stat-val{font-size:30px;font-weight:700;line-height:1;color:var(--navy);margin-top:12px}
.stat-card .stat-label{font-size:12.5px;color:var(--muted);font-weight:500;margin-top:4px}

/* ── Status pills ── */
.pill{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:999px;font-size:11.5px;font-weight:600;white-space:nowrap}
.pill-ok{background:var(--success-light);color:#065f46}
.pill-warn{background:var(--warning-light);color:#92400e}
.pill-danger{background:var(--danger-light);color:#991b1b}
.pill-muted{background:#f1f5f9;color:var(--muted)}
.pill-milestone{background:#ede9fe;color:#5b21b6}

/* ── Table ── */
.data-table{border-collapse:separate;border-spacing:0;width:100%}
.data-table th{
  background:#f8fafc;font-size:11px;font-weight:600;color:var(--muted);
  text-transform:uppercase;letter-spacing:.5px;
  padding:10px 14px;border-bottom:1px solid var(--border);white-space:nowrap;
}
.data-table td{padding:11px 14px;border-bottom:1px solid #f1f5f9;font-size:13px;vertical-align:middle}
.data-table tr:last-child td{border-bottom:none}
.data-table tbody tr{transition:background .1s}
.data-table tbody tr:hover td{background:#f8fafc}
.data-table td.id-cell{font-weight:600;color:var(--navy);font-size:12.5px}
.row-overdue td{background:#fff5f5}
.row-overdue td:first-child{border-left:3px solid #dc2626}
.row-overdue:hover td{background:#fee2e2 !important}

/* ── Status bar ── */
.tat-bar{height:5px;border-radius:3px;background:#e2e8f0;width:90px;overflow:hidden;margin-top:5px}
.tat-bar-fill{height:100%;border-radius:3px;transition:width .3s}

/* ── Check/cross ── */
.sent-yes{color:var(--success);font-size:15px}
.sent-no{color:#cbd5e1;font-size:15px}

/* ── Forms ── */
.form-card{border-radius:10px;overflow:hidden}
.form-section-title{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--muted);margin-bottom:12px}
.form-label{font-size:13px;font-weight:500;color:#374151;margin-bottom:5px}
.form-control,.form-select{font-size:13.5px;border-color:var(--border);border-radius:7px;padding:8px 12px}
.form-control:focus,.form-select:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,148,202,.12)}
.stage-info-box{background:var(--navy-light);border:1px solid #c7d9f0;border-radius:8px;padding:12px 16px;margin-top:8px;display:none}
.stage-info-box.show{display:block}

/* ── Buttons ── */
.btn{border-radius:7px;font-size:13px;font-weight:500;padding:8px 16px}
.btn-sm{padding:5px 12px;font-size:12.5px}
.btn-primary{background:var(--accent);border-color:var(--accent)}
.btn-primary:hover{background:var(--accent-dark);border-color:var(--accent-dark)}
.btn-navy{background:var(--navy);border-color:var(--navy);color:#fff}
.btn-navy:hover{background:var(--navy-dark);border-color:var(--navy-dark);color:#fff}
.btn-action{padding:3px 10px;font-size:11.5px;border-radius:5px}

/* ── Upload zone ── */
.upload-zone{
  border:2px dashed #cbd5e1;border-radius:10px;padding:36px 20px;
  text-align:center;background:#fafbfc;cursor:pointer;transition:all .2s;
}
.upload-zone:hover,.upload-zone.drag{border-color:var(--accent);background:var(--accent-light)}

/* ── Toast ── */
.toast-container{position:fixed;bottom:24px;right:24px;z-index:9999}
.qci-toast{min-width:300px;border-radius:10px;box-shadow:0 4px 20px rgba(0,0,0,.12)}

/* ── Placeholder chips ── */
.ph-chip{
  display:inline-block;padding:2px 8px;margin:2px;border-radius:4px;
  font-size:11px;font-family:monospace;background:#f1f5f9;color:#475569;
  border:1px solid #e2e8f0;cursor:pointer;transition:background .1s;
}
.ph-chip:hover{background:var(--accent-light);color:var(--accent)}

/* ── Breadcrumb ── */
.page-crumb{font-size:12px;color:var(--muted)}
.page-crumb a{color:var(--muted);text-decoration:none}
.page-crumb a:hover{color:var(--navy)}

/* ── Accordion ── */
.accordion-button:not(.collapsed){background:var(--navy-light);color:var(--navy)}
.accordion-button:focus{box-shadow:0 0 0 3px rgba(0,148,202,.12)}

th{white-space:nowrap}

/* ── Responsive ── */
#sidebar-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:99}
.hamburger{display:none;background:none;border:none;cursor:pointer;padding:4px 8px;color:var(--navy);font-size:20px}
@media(max-width:991px){
  #sidebar{transform:translateX(-100%);transition:transform .25s ease;z-index:200}
  #sidebar.open{transform:translateX(0)}
  #sidebar-overlay.open{display:block}
  #main{margin-left:0}
  .hamburger{display:flex;align-items:center}
  .topbar{padding:0 16px}
  .page-body{padding:16px}
}
@media(max-width:575px){
  .stat-card .stat-val{font-size:24px}
  .page-body{padding:12px 10px}
  .topbar-title{font-size:13px}
  .data-table th,.data-table td{padding:8px 10px;font-size:12px}
  .card-header{font-size:13px;padding:12px 14px}
  .btn-sm{font-size:11px;padding:4px 9px}
  .modal-dialog{margin:10px}
}
@media(min-width:1600px){
  #sidebar{width:240px}
  #main{margin-left:240px}
  .page-body{padding:28px 36px}
  .stat-card .stat-val{font-size:34px}
}
/* ── Responsive (additional) ─────────────────────────────────────────────── */
@media (max-width: 768px) {
  .sidebar { transform: translateX(-260px); transition: transform .25s ease; position: fixed; z-index: 1050; height: 100vh; }
  .sidebar.open { transform: translateX(0); }
  .main-content { margin-left: 0 !important; padding: 16px 12px; }
  .topbar { left: 0 !important; }
  .mobile-menu-btn { display: flex !important; }
  .stat-card { min-width: 140px; }
  .data-table th, .data-table td { padding: 10px 8px; font-size: 12px; }
  .card-header { font-size: 13px; }
  .hide-mobile { display: none !important; }
}
@media (min-width: 769px) {
  .mobile-menu-btn { display: none !important; }
  .sidebar-overlay { display: none !important; }
}
@media (min-width: 1400px) {
  .main-content { max-width: 1600px; }
}
</style>
<meta name="csrf-token" content="{{ session.get('csrf_token', '') }}">
<script>
document.addEventListener('DOMContentLoaded', function() {
  var t = document.querySelector('meta[name="csrf-token"]').content;
  document.querySelectorAll('form[method="post"],form[method="POST"]').forEach(function(f) {
    if (!f.querySelector('input[name="csrf_token"]')) {
      var i = document.createElement('input');
      i.type = 'hidden'; i.name = 'csrf_token'; i.value = t;
      f.appendChild(i);
    }
  });
});
</script>
</head>
<body>

<!-- Sidebar -->
<div id="sidebar">
  <div class="brand">
    <a class="brand-logo" href="/"><i class="bi bi-bell-fill" style="color:#11a3d4"></i> QCI <span>Notify</span></a>
    <div class="brand-sub">Quality Council of India</div>
  </div>
  <div class="sidebar-search">
    <input type="text" id="sidebarSearch" placeholder="&#xF52A; Search cases…"
           onkeydown="if(event.key==='Enter'){window.location='/search?q='+encodeURIComponent(this.value)}">
  </div>
  <nav>
    <div class="nav-section">Main</div>
    <a class="nav-link {{ 'active' if active_page=='dashboard' else '' }}" href="/">
      <i class="bi bi-speedometer2"></i> Dashboard
    </a>
    {% if user_role in ('program_officer', 'program_head') %}
    <a class="nav-link {{ 'active' if active_page=='my_cases' else '' }}" href="/?my_cases=1">
      <i class="bi bi-person-check"></i> My Cases
    </a>
    {% endif %}
    <a class="nav-link {{ 'active' if active_page=='log' else '' }}" href="/log-stage">
      <i class="bi bi-plus-circle"></i> Log Stage Change
    </a>
    <a class="nav-link {{ 'active' if active_page=='bulk' else '' }}" href="/bulk-upload">
      <i class="bi bi-upload"></i> Bulk Upload
    </a>
    <a class="nav-link {{ 'active' if active_page=='bulk_advance' else '' }}" href="/bulk-advance">
      <i class="bi bi-fast-forward-fill"></i> Bulk Advance
    </a>
    <div class="nav-section" style="margin-top:8px">Reports</div>
    {% if user_role == 'board_ceo' %}
    <a class="nav-link {{ 'active' if active_page=='ceo_dashboard' else '' }}" href="/ceo-dashboard">
      <i class="bi bi-speedometer2"></i> CEO Dashboard
    </a>
    {% endif %}
    <a class="nav-link {{ 'active' if active_page=='reports' else '' }}" href="/reports">
      <i class="bi bi-graph-up"></i> Analytics
    </a>
    {% if user_role in ('super_admin', 'board_admin') %}
    <a class="nav-link {{ 'active' if active_page=='assessor_scorecard' else '' }}" href="/assessor-scorecard">
      <i class="bi bi-person-badge"></i> Assessor Scorecard
    </a>
    {% endif %}
    <a class="nav-link {{ 'active' if active_page=='export_excel' else '' }}" href="/export-excel">
      <i class="bi bi-file-earmark-excel"></i> Export Report
    </a>
    <a class="nav-link {{ 'active' if active_page=='search' else '' }}" href="/search">
      <i class="bi bi-search"></i> Search Cases
    </a>
    {% if user_role in ('super_admin', 'board_admin') %}
    <div class="nav-section" style="margin-top:8px">Admin</div>
    <a class="nav-link {{ 'active' if active_page=='settings' else '' }}" href="/settings">
      <i class="bi bi-sliders"></i> Programmes &amp; Stages
    </a>
    <a class="nav-link {{ 'active' if active_page=='templates' else '' }}" href="/templates">
      <i class="bi bi-envelope-paper"></i> Email Templates
    </a>
    <a class="nav-link {{ 'active' if active_page=='email_preview' else '' }}" href="/email-preview">
      <i class="bi bi-envelope-open"></i> Email Preview
    </a>
    <a class="nav-link {{ 'active' if active_page=='queue' else '' }}" href="/email-queue">
      <i class="bi bi-send-check"></i> Email Queue
    </a>
    <a class="nav-link {{ 'active' if active_page=='audit' else '' }}" href="/audit-log">
      <i class="bi bi-journal-text"></i> Audit Log
    </a>
    {% if user_role == 'super_admin' %}
    <a class="nav-link {{ 'active' if active_page=='users' else '' }}" href="/users">
      <i class="bi bi-people"></i> Manage Users
    </a>
    <a class="nav-link {{ 'active' if active_page=='system' else '' }}" href="/system-settings">
      <i class="bi bi-gear"></i> System Settings
    </a>
    <a class="nav-link {{ 'active' if active_page=='api_keys' else '' }}" href="/api-keys">
      <i class="bi bi-key-fill"></i> API Keys
    </a>
    <a class="nav-link" href="/backup">
      <i class="bi bi-download"></i> Backup DB
    </a>
    {% endif %}
    {% endif %}
  </nav>
  <div class="sidebar-footer">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">
      <div style="width:32px;height:32px;border-radius:50%;background:rgba(255,255,255,.1);
                  display:flex;align-items:center;justify-content:center;font-size:14px;color:#fff">
        <i class="bi bi-person-fill"></i>
      </div>
      <div>
        <div style="color:#fff;font-size:13px;font-weight:500">{{ user_name }}</div>
        <div style="color:rgba(255,255,255,.4);font-size:10px;text-transform:uppercase;letter-spacing:.5px">
          {% if user_role=='super_admin' %}Super Admin
          {% elif user_role=='board_admin' %}Board Admin · {{ board_name }}
          {% elif user_role=='board_ceo' %}Board CEO · {{ board_name }}
          {% elif user_role=='program_head' %}Programme Head · {{ board_name }}
          {% else %}Program Officer · {{ board_name }}{% endif %}
        </div>
      </div>
    </div>
    <div style="display:flex;align-items:center;justify-content:space-between">
      <a href="/logout" style="font-size:12px;color:rgba(255,255,255,.4);text-decoration:none;
         display:flex;align-items:center;gap:6px;padding:4px 0">
        <i class="bi bi-box-arrow-left"></i> Sign out
      </a>
      <button onclick="toggleDark()" style="background:none;border:none;cursor:pointer;
              color:rgba(255,255,255,.35);font-size:15px;padding:4px" title="Toggle dark mode">
        <i class="bi bi-moon-fill" id="darkIcon"></i>
      </button>
    </div>
  </div>
</div>

<!-- Main -->
<div id="sidebar-overlay" onclick="closeSidebar()"></div>
<div id="main">
  <div class="topbar">
    <div class="d-flex align-items-center gap-2">
      <button class="hamburger" onclick="openSidebar()" title="Menu">
        <i class="bi bi-list"></i>
      </button>
      <div>
        <div class="topbar-title">{{ page_title }}</div>
        {% if page_crumb %}<div class="page-crumb">{{ page_crumb | safe }}</div>{% endif %}
      </div>
    </div>
    <div class="d-flex align-items-center gap-2">
      {{ topbar_actions | safe }}
    </div>
  </div>

  <!-- Flash toasts -->
  <div class="toast-container">
  {% with msgs = get_flashed_messages(with_categories=true) %}
  {% if msgs %}{% for cat,msg in msgs %}
  <div class="toast qci-toast show align-items-center text-white bg-{{ 'danger' if cat=='error' else ('success' if cat=='success' else 'primary') }} border-0 mb-2" role="alert">
    <div class="d-flex">
      <div class="toast-body">{{ msg }}</div>
      <button type="button" class="btn-close btn-close-white me-2 m-auto" data-bs-dismiss="toast"></button>
    </div>
  </div>
  {% endfor %}{% endif %}{% endwith %}
  </div>

  <div class="page-body">
    {{ content | safe }}
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script>
// Auto-dismiss toasts after 5s
document.querySelectorAll('.toast.show').forEach(function(el){
  setTimeout(function(){ var t=bootstrap.Toast.getOrCreateInstance(el); t.hide(); }, 5000);
});
// Mobile sidebar
function openSidebar(){
  document.getElementById('sidebar').classList.add('open');
  document.getElementById('sidebar-overlay').classList.add('open');
  document.body.style.overflow='hidden';
}
function closeSidebar(){
  document.getElementById('sidebar').classList.remove('open');
  document.getElementById('sidebar-overlay').classList.remove('open');
  document.body.style.overflow='';
}
// Close sidebar on nav link click (mobile)
document.querySelectorAll('#sidebar .nav-link').forEach(function(a){
  a.addEventListener('click', function(){ if(window.innerWidth<992) closeSidebar(); });
});
// Dark mode
function toggleDark(){
  var on = document.documentElement.classList.toggle('dark-mode');
  localStorage.setItem('qci_dark', on ? '1' : '0');
  document.getElementById('darkIcon').className = on ? 'bi bi-sun-fill' : 'bi bi-moon-fill';
}
(function(){
  if(localStorage.getItem('qci_dark')==='1'){
    document.documentElement.classList.add('dark-mode');
    var ic = document.getElementById('darkIcon');
    if(ic) ic.className = 'bi bi-sun-fill';
  }
})();
// Quick advance modal helper
function openQuickAdvance(caseId, appId, progName){
  document.getElementById('qa_case_id').value = caseId;
  document.getElementById('qa_app_label').textContent = appId;
  fetch('/api/stages?programme=' + encodeURIComponent(progName))
    .then(r=>r.json()).then(function(stages){
      var sel = document.getElementById('qa_stage_select');
      sel.innerHTML = stages.map(s=>'<option value="'+s.name+'">'+s.name+'</option>').join('');
    });
  new bootstrap.Modal(document.getElementById('quickAdvanceModal')).show();
}
// Confirm delete
function confirmDelete(url, appId){
  if(confirm('Close case ' + appId + '? This cannot be undone.')){
    window.location = url;
  }
}
</script>
{{ scripts | safe }}
</body></html>"""


def render_page(content: str, scripts: str = "", active_page: str = "",
                page_title: str = "QCI Notification Engine",
                page_crumb: str = "", topbar_actions: str = "") -> str:
    from flask import get_flashed_messages as gfm
    user_role  = session.get("role", "")
    user_name  = session.get("full_name", "")
    board_name = session.get("board_name", "")
    return render_template_string(
        _BASE, content=content, scripts=scripts,
        active_page=active_page, page_title=page_title,
        page_crumb=page_crumb, topbar_actions=topbar_actions,
        get_flashed_messages=gfm,
        user_role=user_role, user_name=user_name, board_name=board_name,
    )


# ── Login page template (no sidebar) ─────────────────────────────────────────
_LOGIN_PAGE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sign In — QCI Notify</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<style>
  *{font-family:'Inter',sans-serif}
  body{background:#f1f5f9;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}
  .login-card{background:#fff;border-radius:16px;padding:40px 44px;width:100%;max-width:400px;
              box-shadow:0 4px 24px rgba(0,0,0,.08);border:1px solid #e2e8f0}
  .brand{text-align:center;margin-bottom:32px}
  .brand-icon{width:56px;height:56px;background:#1a3557;border-radius:14px;
              display:flex;align-items:center;justify-content:center;margin:0 auto 14px;font-size:24px}
  .brand-name{font-size:20px;font-weight:700;color:#1a3557}
  .brand-sub{font-size:12px;color:#94a3b8;margin-top:2px}
  .form-label{font-size:13px;font-weight:500;color:#374151}
  .form-control{border-radius:8px;border-color:#e2e8f0;padding:9px 13px;font-size:14px}
  .form-control:focus{border-color:#2563eb;box-shadow:0 0 0 3px rgba(37,99,235,.1)}
  .btn-login{background:#1a3557;color:#fff;border:none;border-radius:8px;
             padding:11px;font-size:14px;font-weight:600;width:100%}
  .btn-login:hover{background:#0f2035;color:#fff}
  .alert{border-radius:8px;font-size:13px}
  .creds-hint{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
              padding:12px 14px;margin-top:20px;font-size:12px;color:#64748b}
  .creds-hint code{background:#e2e8f0;padding:1px 5px;border-radius:3px;font-size:11px}
</style>
</head>
<body>
<div class="login-card">
  <div class="brand">
    <div class="brand-icon"><i class="bi bi-bell-fill" style="color:#fbbf24"></i></div>
    <div class="brand-name">QCI Notify</div>
    <div class="brand-sub">Quality Council of India</div>
  </div>
  {% with msgs = get_flashed_messages(with_categories=true) %}
  {% if msgs %}{% for cat,msg in msgs %}
  <div class="alert alert-{{ 'danger' if cat=='error' else ('success' if cat=='success' else 'info') }}">
    {{ msg }}</div>
  {% endfor %}{% endif %}{% endwith %}
  <form method="post">
    <div class="mb-3">
      <label class="form-label">Username</label>
      <input type="text" class="form-control" name="username" autofocus required
             placeholder="Enter username" value="{{ prefill_user }}">
    </div>
    <div class="mb-{% if show_totp %}3{% else %}4{% endif %}">
      <label class="form-label">Password</label>
      <input type="password" class="form-control" name="password" required placeholder="••••••••">
    </div>
    {% if show_totp %}
    <div class="mb-4">
      <label class="form-label">Authenticator Code</label>
      <input type="text" class="form-control" name="totp_code" maxlength="6"
             placeholder="6-digit code" autocomplete="one-time-code">
    </div>
    {% endif %}
    <button type="submit" class="btn-login">Sign In</button>
  </form>
  <div class="creds-hint" style="padding:14px 16px">
    <div style="font-size:11px;font-weight:700;letter-spacing:.5px;color:#94a3b8;margin-bottom:10px">DEFAULT CREDENTIALS</div>
    <table style="width:100%;font-size:12px;border-collapse:collapse">
      <thead>
        <tr style="color:#94a3b8;border-bottom:1px solid #e2e8f0">
          <th style="text-align:left;padding-bottom:6px;font-weight:600">Role</th>
          <th style="text-align:left;padding-bottom:6px;font-weight:600">Username</th>
          <th style="text-align:left;padding-bottom:6px;font-weight:600">Password</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:4px 0;color:#1e293b">Super Admin</td><td><code>admin</code></td><td><code>admin123</code></td></tr>
        <tr><td style="padding:4px 0;color:#1e293b">Program Officer</td><td><code>officer</code></td><td><code>po123</code></td></tr>
      </tbody>
    </table>
    <div style="font-size:10px;color:#94a3b8;margin-top:10px;border-top:1px solid #e2e8f0;padding-top:8px">
      Board Admin, Board CEO &amp; Programme Head accounts are created by Super Admin.
    </div>
  </div>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body></html>"""


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            # 2FA check
            if user["totp_secret"]:
                totp_code = request.form.get("totp_code", "").strip()
                if not totp_code:
                    flash("This account has 2FA enabled. Enter your authenticator code.", "info")
                    return render_template_string(_LOGIN_PAGE, get_flashed_messages=gfm,
                                                  show_totp=True, prefill_user=username)
                if not totp_verify(user["totp_secret"], totp_code):
                    flash("Invalid 2FA code.", "error")
                    return render_template_string(_LOGIN_PAGE, get_flashed_messages=gfm,
                                                  show_totp=True, prefill_user=username)
            # Record last login
            conn3 = get_db()
            conn3.execute("UPDATE users SET last_login=? WHERE id=?",
                         (now_ist().strftime("%Y-%m-%d %H:%M:%S"), user["id"]))
            conn3.commit()
            conn3.close()
            session["user_id"]    = user["id"]
            session["username"]   = user["username"]
            session["role"]       = user["role"]
            session["full_name"]  = user["full_name"] or user["username"]
            session["board_id"]   = user["board_id"]
            session["csrf_token"] = secrets.token_hex(32)
            if user["board_id"]:
                conn2 = get_db()
                brow = conn2.execute("SELECT board_name FROM boards WHERE id=?", (user["board_id"],)).fetchone()
                conn2.close()
                session["board_name"] = brow["board_name"] if brow else ""
            else:
                session["board_name"] = ""
            if user["force_password_reset"]:
                flash("Please change your password before continuing.", "info")
                return redirect(url_for("force_pw_reset"))
            flash(f"Welcome back, {session['full_name']}!", "success")
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    from flask import get_flashed_messages as gfm
    return render_template_string(_LOGIN_PAGE, get_flashed_messages=gfm,
                                  show_totp=False, prefill_user="")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been signed out.", "info")
    return redirect(url_for("login"))


@app.route("/force-reset-password", methods=["GET", "POST"])
@login_required
def force_pw_reset():
    if request.method == "POST":
        new_pw = request.form.get("new_password", "").strip()
        confirm_pw = request.form.get("confirm_password", "").strip()
        if len(new_pw) < 8:
            flash("Password must be at least 8 characters.", "error")
        elif new_pw != confirm_pw:
            flash("Passwords do not match.", "error")
        else:
            conn = get_db()
            conn.execute(
                "UPDATE users SET password_hash=?, force_password_reset=0 WHERE id=?",
                (generate_password_hash(new_pw), session["user_id"])
            )
            conn.commit()
            conn.close()
            flash("Password updated. Welcome!", "success")
            return redirect(url_for("dashboard"))
    content = """
<div style="max-width:420px;margin:60px auto">
  <div class="card">
    <div class="card-header" style="background:linear-gradient(135deg,var(--navy),var(--accent));color:#fff;border-radius:10px 10px 0 0">
      <i class="bi bi-shield-lock-fill"></i> Set Your Password
    </div>
    <div class="card-body p-4">
      <p style="font-size:13px;color:#64748b;margin-bottom:20px">
        You are required to set a new password before continuing.
      </p>
      <form method="post">
        <div class="mb-3">
          <label class="form-label">New Password</label>
          <input type="password" class="form-control" name="new_password" minlength="8" required>
          <div style="font-size:11px;color:#94a3b8;margin-top:4px">Minimum 8 characters</div>
        </div>
        <div class="mb-4">
          <label class="form-label">Confirm Password</label>
          <input type="password" class="form-control" name="confirm_password" required>
        </div>
        <button class="btn btn-primary w-100">Set Password &amp; Continue</button>
      </form>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="", page_title="Set New Password")


@app.route("/users", methods=["GET", "POST"])
@board_admin_required
def manage_users():
    _is_super = session.get("role") == "super_admin"
    _caller_board_id = user_board_id()
    conn = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            try:
                role = request.form["role"]
                # board_admin cannot create super_admin accounts
                if not _is_super and role == "super_admin":
                    flash("Permission denied: cannot create super_admin accounts.", "error")
                    conn.close()
                    return redirect(url_for("manage_users"))
                board_id = request.form.get("board_id") or None
                if board_id:
                    board_id = int(board_id)
                # board_admin can only create users in their own board
                if not _is_super and _caller_board_id is not None:
                    board_id = _caller_board_id
                conn.execute(
                    "INSERT INTO users (username, password_hash, role, full_name, email, board_id) VALUES (?,?,?,?,?,?)",
                    (request.form["username"].strip(),
                     generate_password_hash(request.form["password"]),
                     role,
                     request.form["full_name"].strip(),
                     request.form["email"].strip(),
                     board_id),
                )
                conn.commit()
                new_user = conn.execute(
                    "SELECT id FROM users WHERE username=?", (request.form["username"].strip(),)
                ).fetchone()
                if new_user and role in ("program_head", "program_officer", "board_ceo"):
                    prog_ids = request.form.getlist("programme_ids")
                    for pid in prog_ids:
                        try:
                            conn.execute(
                                "INSERT INTO user_programme_map (user_id, programme_id) VALUES (?,?) ON CONFLICT (user_id, programme_id) DO NOTHING",
                                (new_user["id"], int(pid))
                            )
                        except Exception:
                            pass
                    conn.commit()
                flash(f"User '{request.form['username']}' created.", "success")
            except Exception as e:
                flash(f"Error: {e}", "error")
        elif action == "delete":
            uid = request.form.get("user_id")
            if str(uid) == str(session["user_id"]):
                flash("You cannot delete your own account.", "error")
            else:
                conn.execute("DELETE FROM users WHERE id=?", (uid,))
                conn.commit()
                flash("User deleted.", "success")
        elif action == "reset_password":
            uid = request.form.get("user_id")
            new_pw = request.form.get("new_password", "").strip()
            if new_pw:
                conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                             (generate_password_hash(new_pw), uid))
                conn.commit()
                flash("Password updated.", "success")
        elif action == "remap_ph":
            uid = request.form.get("remap_user_id")
            if uid:
                conn.execute("DELETE FROM user_programme_map WHERE user_id=?", (uid,))
                prog_ids = request.form.getlist("programme_ids")
                for pid in prog_ids:
                    try:
                        conn.execute(
                            "INSERT INTO user_programme_map (user_id, programme_id) VALUES (?,?) ON CONFLICT (user_id, programme_id) DO NOTHING",
                            (int(uid), int(pid))
                        )
                    except Exception:
                        pass
                conn.commit()
                flash("Programme mapping updated.", "success")

    if _is_super:
        users = [dict(r) for r in conn.execute(
            """SELECT u.*, b.board_name FROM users u
               LEFT JOIN boards b ON b.id = u.board_id
               ORDER BY u.role, u.username"""
        ).fetchall()]
    else:
        users = [dict(r) for r in conn.execute(
            """SELECT u.*, b.board_name FROM users u
               LEFT JOIN boards b ON b.id = u.board_id
               WHERE u.board_id=?
               ORDER BY u.role, u.username""",
            (_caller_board_id,)
        ).fetchall()]
    boards = [dict(r) for r in conn.execute("SELECT * FROM boards ORDER BY board_name").fetchall()]
    if _is_super:
        all_programmes = [dict(r) for r in conn.execute(
            "SELECT p.*, b.board_name FROM programmes p JOIN boards b ON b.id=p.board_id ORDER BY b.board_name, p.programme_name"
        ).fetchall()]
    else:
        all_programmes = [dict(r) for r in conn.execute(
            "SELECT p.*, b.board_name FROM programmes p JOIN boards b ON b.id=p.board_id WHERE p.board_id=? ORDER BY p.programme_name",
            (_caller_board_id,)
        ).fetchall()]

    # Build programme map per user (for program_head display)
    ph_prog_map = {}
    for row in conn.execute(
        """SELECT upm.user_id, p.programme_name
           FROM user_programme_map upm JOIN programmes p ON p.id=upm.programme_id"""
    ).fetchall():
        ph_prog_map.setdefault(row["user_id"], []).append(row["programme_name"])

    conn.close()

    rows = ""
    for u in users:
        rl, bg, fg = ROLE_META.get(u["role"], (u["role"], "#f1f5f9", "#475569"))
        role_pill = f'<span class="pill" style="background:{bg};color:{fg}">{rl}</span>'
        board_cell = u.get("board_name") or '<span style="color:#94a3b8">—</span>'
        is_self = u["id"] == session["user_id"]
        _REMAP_ROLES = {"program_officer", "program_head", "board_ceo"}
        prog_cell = ""
        if u["role"] in _REMAP_ROLES:
            progs = ph_prog_map.get(u["id"], [])
            if progs:
                prog_cell = " ".join(
                    f'<span style="background:#ede9fe;color:#7c3aed;font-size:10px;padding:1px 6px;border-radius:10px;white-space:nowrap">{p}</span>'
                    for p in progs
                )
            else:
                prog_cell = '<span style="color:#f59e0b;font-size:11px">None mapped</span>'
        last_login_cell = u.get("last_login") or '<span style="color:#94a3b8;font-size:11px">Never</span>'
        _uid = u["id"]
        _uname = u["username"]
        _urole = u["role"]
        remap_btn = (
            '<button class="btn btn-sm btn-outline-primary" style="font-size:12px;padding:3px 8px" title="Map Programmes" '
            f'onclick="showRemapPh({json.dumps(str(_uid))}, {json.dumps(str(_uname))}, {json.dumps(str(_urole))})">'
            '<i class="bi bi-diagram-3"></i></button>'
            if _urole in _REMAP_ROLES else ""
        )
        del_btn = (
            '<button class="btn btn-sm btn-outline-danger" style="font-size:12px;padding:3px 8px" title="Delete User" '
            f'onclick="confirmDeleteUser({_uid}, {json.dumps(_uname)})">'
            '<i class="bi bi-trash"></i></button>'
            if not is_self else ""
        )
        rows += f"""<tr>
          <td style="font-weight:600">{u['username']}</td>
          <td>{u['full_name'] or '—'}</td>
          <td>{u['email'] or '—'}</td>
          <td>{role_pill}</td>
          <td>{board_cell}</td>
          <td style="font-size:12px;color:#475569;max-width:180px">{prog_cell or '<span style="color:#94a3b8">—</span>'}</td>
          <td style="font-size:11px;color:#64748b">{last_login_cell}</td>
          <td>
            <div class="d-flex gap-1 align-items-center" style="white-space:nowrap">
              <button class="btn btn-sm btn-outline-secondary" style="font-size:12px;padding:3px 8px" title="Reset Password"
                onclick="showResetPw({u['id']}, {json.dumps(u['username'])})">
                <i class="bi bi-key"></i></button>
              {remap_btn}
              {del_btn}
            </div>
          </td>
        </tr>"""

    board_opts = '<option value="">— None —</option>' + "".join(
        f'<option value="{b["id"]}">{b["board_name"]}</option>' for b in boards
    )
    prog_opts_by_board = "".join(
        f'<option value="{p["id"]}" data-board="{p["board_id"]}">[{p["board_name"]}] {p["programme_name"]}</option>'
        for p in all_programmes
    )

    content = f"""
<div class="row g-4">
  <div class="col-lg-8">
    <div class="card">
      <div class="card-header d-flex justify-content-between align-items-center flex-wrap gap-2">
        <span>All Users</span>
        <div class="d-flex gap-2 align-items-center">
          <input type="text" id="userSearch" class="form-control form-control-sm"
                 placeholder="Search users..." oninput="filterUsers(this.value)"
                 style="max-width:200px;font-size:12px">
          <a href="/bulk-users" class="btn btn-sm btn-outline-primary">
            <i class="bi bi-upload"></i> Bulk Import
          </a>
        </div>
      </div>
      <div style="overflow-x:auto">
        <table class="data-table" id="userTable">
          <thead><tr><th>Username</th><th>Full Name</th><th>Email</th><th>Role</th><th>Board</th><th>Programmes</th><th>Last Login</th><th>Actions</th></tr></thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="col-lg-4">
    <div class="card">
      <div class="card-header"><i class="bi bi-person-plus" style="color:#059669"></i> Add New User</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="add">
          <div class="mb-3">
            <label class="form-label">Username</label>
            <input type="text" class="form-control" name="username" required>
          </div>
          <div class="mb-3">
            <label class="form-label">Full Name</label>
            <input type="text" class="form-control" name="full_name">
          </div>
          <div class="mb-3">
            <label class="form-label">Email</label>
            <input type="email" class="form-control" name="email">
          </div>
          <div class="mb-3">
            <label class="form-label">Role</label>
            <select class="form-select" name="role" id="roleSelect" onchange="toggleRoleFields()">
              <option value="program_officer">Program Officer</option>
              <option value="program_head">Programme Head</option>
              <option value="board_ceo">Board CEO</option>
              <option value="board_admin">Board Admin</option>
              <option value="super_admin">Super Admin</option>
            </select>
          </div>
          <div class="mb-3" id="boardField">
            <label class="form-label">Board <span style="color:#94a3b8;font-size:11px">(required for all except Super Admin)</span></label>
            <select class="form-select" name="board_id" id="boardSelect" onchange="filterProgsByBoard()">
              {board_opts}
            </select>
          </div>
          <div class="mb-3" id="progField" style="display:none">
            <label class="form-label">Programmes <span style="color:#94a3b8;font-size:11px">(hold Ctrl/Cmd for multiple)</span></label>
            <select class="form-select" name="programme_ids" id="progSelect" multiple size="5">
              {prog_opts_by_board}
            </select>
            <div style="font-size:11px;color:#94a3b8;margin-top:4px">Only programmes for the selected board are shown.</div>
          </div>
          <div class="mb-3">
            <label class="form-label">Password</label>
            <input type="password" class="form-control" name="password" required>
          </div>
          <button type="submit" class="btn btn-primary w-100">Create User</button>
        </form>
      </div>
    </div>
  </div>
</div>

<!-- Reset password modal -->
<div class="modal fade" id="resetPwModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered modal-sm">
    <div class="modal-content">
      <div class="modal-header border-0"><h6 class="modal-title">Reset Password — <span id="rpUser"></span></h6></div>
      <form method="post">
        <input type="hidden" name="action" value="reset_password">
        <input type="hidden" name="user_id" id="rpUserId">
        <div class="modal-body pt-0">
          <input type="password" class="form-control" name="new_password" placeholder="New password" required>
        </div>
        <div class="modal-footer border-0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button class="btn btn-sm btn-primary" type="submit">Update</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Programme Head Re-mapping Modal -->
<div class="modal fade" id="remapPhModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header border-0">
        <h6 class="modal-title"><i class="bi bi-diagram-3-fill" style="color:#166534"></i> Re-map Programmes — <span id="remapPhUser"></span></h6>
        <button class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <form method="post">
        <input type="hidden" name="action" value="remap_ph">
        <input type="hidden" name="remap_user_id" id="remapPhUserId">
        <div class="modal-body pt-0">
          <p style="font-size:12px;color:#64748b">Select all programmes this user should have access to (hold Ctrl/Cmd for multiple):</p>
          <select class="form-select" name="programme_ids" id="remapProgSelect" multiple size="7">
            {prog_opts_by_board}
          </select>
        </div>
        <div class="modal-footer border-0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button class="btn btn-sm btn-success" type="submit">Save Mapping</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Delete user modal — form lives outside -->
<div class="modal fade" id="delUserModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered modal-sm">
    <div class="modal-content">
      <div class="modal-body text-center p-4">
        <i class="bi bi-person-x" style="font-size:32px;color:#dc2626"></i>
        <div style="font-weight:600;margin-top:10px">Delete <span id="delUserName"></span>?</div>
        <div style="font-size:12px;color:#94a3b8;margin-top:6px">This cannot be undone.</div>
      </div>
      <div class="modal-footer border-0 justify-content-center gap-2">
        <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
        <button type="button" class="btn btn-sm btn-danger" onclick="submitDeleteUser()">Delete</button>
      </div>
    </div>
  </div>
</div>
<!-- Hidden delete form — outside modal so Cancel cannot accidentally submit it -->
<form method="post" id="deleteUserForm" style="display:none">
  <input type="hidden" name="action" value="delete">
  <input type="hidden" name="user_id" id="delUserId">
</form>
"""
    scripts = """<script>
var _PROG_ROLES = ['program_officer', 'program_head', 'board_ceo'];

function showResetPw(id, name){
  document.getElementById('rpUserId').value = id;
  document.getElementById('rpUser').textContent = name;
  new bootstrap.Modal(document.getElementById('resetPwModal')).show();
}
function confirmDeleteUser(id, name){
  document.getElementById('delUserId').value = id;
  document.getElementById('delUserName').textContent = name;
  new bootstrap.Modal(document.getElementById('delUserModal')).show();
}
function submitDeleteUser(){
  document.getElementById('deleteUserForm').submit();
}
function filterProgsByBoard(){
  var bid = document.getElementById('boardSelect').value;
  var opts = document.getElementById('progSelect').options;
  for(var i=0; i<opts.length; i++){
    var show = !bid || opts[i].dataset.board == bid;
    opts[i].style.display = show ? '' : 'none';
    if(bid && opts[i].dataset.board != bid) opts[i].selected = false;
  }
}
function toggleRoleFields(){
  var role = document.getElementById('roleSelect').value;
  document.getElementById('boardField').style.display = role === 'super_admin' ? 'none' : '';
  var showProg = _PROG_ROLES.indexOf(role) >= 0;
  document.getElementById('progField').style.display = showProg ? '' : 'none';
  if(showProg) filterProgsByBoard();
}
toggleRoleFields();
function showRemapPh(id, name, role){
  document.getElementById('remapPhUserId').value = id;
  document.getElementById('remapPhUser').textContent = name + ' (' + role + ')';
  new bootstrap.Modal(document.getElementById('remapPhModal')).show();
}
function filterUsers(q){
  q = q.toLowerCase();
  document.querySelectorAll('#userTable tbody tr').forEach(function(row){
    row.style.display = row.textContent.toLowerCase().includes(q) ? '' : 'none';
  });
}
</script>"""
    return render_page(content, scripts, active_page="users", page_title="Manage Users")


@app.route("/bulk-users", methods=["GET", "POST"])
@admin_required
def bulk_users():
    """Bulk import users from CSV. Columns: username, full_name, email, role, board_name, password"""
    errors = []
    created = 0
    if request.method == "POST":
        f = request.files.get("csv_file")
        if not f:
            flash("No file uploaded.", "error")
            return redirect(url_for("bulk_users"))
        conn = get_db()
        boards = {r["board_name"].lower(): r["id"] for r in conn.execute("SELECT id, board_name FROM boards").fetchall()}
        try:
            fname_u = (f.filename or "").lower()
            if fname_u.endswith(".xlsx") or fname_u.endswith(".xls"):
                if not HAS_XLSX:
                    flash("openpyxl not installed — cannot read Excel files.", "error")
                    return redirect(url_for("bulk_users"))
                wb_u = openpyxl.load_workbook(io.BytesIO(f.read()))
                ws_u = wb_u.active
                all_u_rows = list(ws_u.iter_rows(values_only=True))
                if not all_u_rows:
                    flash("The uploaded file is empty.", "error")
                    return redirect(url_for("bulk_users"))
                hdrs_u = [str(h).strip() if h else "" for h in all_u_rows[0]]
                row_dicts_u = [
                    {hdrs_u[j]: (str(v).strip() if v is not None else "")
                     for j, v in enumerate(r) if j < len(hdrs_u)}
                    for r in all_u_rows[1:]
                ]
                reader_iter = iter(row_dicts_u)
            else:
                content_bytes = f.read().decode("utf-8-sig")
                reader_iter = csv.DictReader(io.StringIO(content_bytes))
            for i, row in enumerate(reader_iter, 2):
                uname = (row.get("username") or "").strip()
                if not uname:
                    errors.append(f"Row {i}: username missing")
                    continue
                role = (row.get("role") or "program_officer").strip()
                if role not in ROLE_META:
                    errors.append(f"Row {i} ({uname}): invalid role '{role}'")
                    continue
                board_name = (row.get("board_name") or "").strip().lower()
                board_id = boards.get(board_name) if board_name else None
                raw_pw = (row.get("password") or "").strip()
                if not raw_pw:
                    raw_pw = "Welcome@123"
                    force_reset = 1
                else:
                    force_reset = 0
                try:
                    conn.execute(
                        "INSERT INTO users (username, password_hash, role, full_name, email, board_id, force_password_reset) VALUES (?,?,?,?,?,?,?)",
                        (uname, generate_password_hash(raw_pw), role,
                         (row.get("full_name") or "").strip(),
                         (row.get("email") or "").strip() or None,
                         board_id, force_reset)
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {i} ({uname}): {e}")
            conn.commit()
        except Exception as e:
            errors.append(f"File error: {e}")
        conn.close()
        if created:
            flash(f"{created} user(s) imported successfully.", "success")
        for err in errors[:10]:
            flash(err, "error")
        return redirect(url_for("bulk_users"))

    # Build CSV template download
    template_csv = "username,full_name,email,role,board_name,password\njohn_doe,John Doe,john@org.com,program_officer,NABH,changeme123\n"
    content = f"""
<div class="row g-4">
  <div class="col-lg-7">
    <div class="card">
      <div class="card-header"><i class="bi bi-upload" style="color:var(--accent)"></i> Bulk User Import</div>
      <div class="card-body p-4">
        <p style="font-size:13px;color:#64748b">Upload a <strong>CSV or Excel (.xlsx)</strong> with columns: <code>username</code>, <code>full_name</code>, <code>email</code>, <code>role</code>, <code>board_name</code>, <code>password</code>.</p>
        <form method="post" enctype="multipart/form-data">
          <div class="upload-zone mb-3" onclick="document.getElementById('userCsvFile').click()">
            <i class="bi bi-file-earmark-person" style="font-size:32px;color:#94a3b8"></i>
            <div style="font-size:13px;color:#64748b;margin-top:8px">Click to select CSV or Excel file</div>
            <div id="csvFileName" style="font-size:12px;color:var(--accent);margin-top:4px"></div>
          </div>
          <input type="file" id="userCsvFile" name="csv_file" accept=".csv,.xlsx,.xls" style="display:none"
                 onchange="document.getElementById('csvFileName').textContent=this.files[0].name">
          <button type="submit" class="btn btn-primary w-100">
            <i class="bi bi-upload"></i> Import Users
          </button>
        </form>
      </div>
    </div>
  </div>
  <div class="col-lg-5">
    <div class="card">
      <div class="card-header">Valid Roles &amp; Boards</div>
      <div class="card-body p-4" style="font-size:13px">
        <strong>Roles:</strong> <code>super_admin</code>, <code>board_admin</code>, <code>board_ceo</code>, <code>program_head</code>, <code>program_officer</code><br><br>
        <strong>Boards:</strong> NABH, NABL, NABCB, NABET (or any board you have added)<br><br>
        <strong>Notes:</strong>
        <ul style="font-size:12px;color:#64748b">
          <li>If password is blank, a random password is set. Ask users to reset on first login.</li>
          <li>Duplicate usernames will be skipped with an error.</li>
          <li>Programme mappings for programme_head must be done separately via Remap.</li>
        </ul>
        <a href="data:text/csv;charset=utf-8,{template_csv}" download="user_import_template.csv"
           class="btn btn-sm btn-outline-secondary w-100 mt-2">
          <i class="bi bi-download"></i> Download Template CSV
        </a>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="users", page_title="Bulk User Import",
                       page_crumb='<a href="/users">Manage Users</a> / Bulk Import')


@app.route("/ceo-dashboard")
@login_required
def ceo_dashboard():
    """Board CEO / super_admin high-level KPI dashboard."""
    if session.get("role") not in ("board_ceo", "super_admin", "board_admin"):
        return redirect(url_for("dashboard"))
    conn = get_db()
    bid = user_board_id()
    today = now_ist().date()

    if bid is not None:
        cases = [dict(r) for r in conn.execute(
            "SELECT * FROM case_tracking WHERE board_id=? AND (status='Active' OR status IS NULL)", (bid,)
        ).fetchall()]
        programmes = [dict(r) for r in conn.execute(
            "SELECT * FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
    else:
        cases = [dict(r) for r in conn.execute(
            "SELECT * FROM case_tracking WHERE (status='Active' OR status IS NULL)"
        ).fetchall()]
        programmes = [dict(r) for r in conn.execute(
            "SELECT p.*, b.board_name FROM programmes p JOIN boards b ON b.id=p.board_id ORDER BY b.board_name, p.programme_name"
        ).fetchall()]
    try:
        _ceo_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                        for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _ceo_db_hols = set()
    conn.close()

    for c in cases:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, extra_holidays=_ceo_db_hols)

    total_cases = len(cases)
    overdue = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0 and c["days_elapsed"] >= c["tat_days"]]
    at_risk  = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0
                and c["days_elapsed"] >= c["reminder2_day"] and c["days_elapsed"] < c["tat_days"]]
    sla_eligible = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0]
    sla_within = len([c for c in sla_eligible if c["days_elapsed"] < c["tat_days"]])
    sla_pct = int(sla_within / len(sla_eligible) * 100) if sla_eligible else 100
    sla_color = "#00984C" if sla_pct >= 80 else ("#d97706" if sla_pct >= 60 else "#dc2626")
    avg_elapsed = int(sum(c["days_elapsed"] for c in sla_eligible) / len(sla_eligible)) if sla_eligible else 0

    # Per-programme SLA scorecard
    prog_sla = {}
    for c in cases:
        pn = c["programme_name"]
        if pn not in prog_sla:
            prog_sla[pn] = {"total": 0, "overdue": 0, "at_risk": 0, "on_track": 0, "eligible": 0, "within": 0}
        prog_sla[pn]["total"] += 1
        if c["is_milestone"] or c["tat_days"] == 0:
            continue
        prog_sla[pn]["eligible"] += 1
        if c["days_elapsed"] >= c["tat_days"]:
            prog_sla[pn]["overdue"] += 1
        elif c["days_elapsed"] >= c["reminder2_day"]:
            prog_sla[pn]["at_risk"] += 1
        else:
            prog_sla[pn]["on_track"] += 1
            prog_sla[pn]["within"] += 1
        if c["days_elapsed"] < c["tat_days"]:
            prog_sla[pn]["within"] += 0  # already counted above via on_track

    # Fix within count (should be on_track)
    for pn in prog_sla:
        prog_sla[pn]["within"] = prog_sla[pn]["on_track"]

    scorecard_rows = ""
    chart_labels = []
    chart_sla = []
    chart_overdue = []
    for pn, ps in sorted(prog_sla.items()):
        pct = int(ps["within"] / ps["eligible"] * 100) if ps["eligible"] else 100
        pct_color = "#00984C" if pct >= 80 else ("#d97706" if pct >= 60 else "#dc2626")
        chart_labels.append(pn[:30])
        chart_sla.append(pct)
        chart_overdue.append(ps["overdue"])
        scorecard_rows += f"""<tr>
  <td style="font-weight:600;font-size:12.5px">{pn}</td>
  <td class="text-center">{ps['total']}</td>
  <td class="text-center" style="color:#00984C;font-weight:600">{ps['on_track']}</td>
  <td class="text-center" style="color:#d97706;font-weight:600">{ps['at_risk']}</td>
  <td class="text-center" style="color:#dc2626;font-weight:600">{ps['overdue']}</td>
  <td>
    <div class="d-flex align-items-center gap-2">
      <div style="flex:1;height:8px;background:#f1f5f9;border-radius:4px;overflow:hidden">
        <div style="width:{pct}%;height:100%;background:{pct_color};border-radius:4px"></div>
      </div>
      <span style="font-size:12px;font-weight:700;color:{pct_color};width:36px">{pct}%</span>
    </div>
  </td>
</tr>"""

    # Monthly trend (last 6 months cases created)
    # Use stage_history first_entry as proxy for case creation
    monthly_labels = []
    monthly_new = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        monthly_labels.append(d.strftime("%b %Y"))
        ym = d.strftime("%Y-%m")
        cnt = sum(1 for c in cases if c["stage_start_date"] and c["stage_start_date"][:7] == ym)
        monthly_new.append(cnt)

    # Assessor delay analysis
    owner_delay = {}
    for c in cases:
        if not c["is_milestone"] and c["tat_days"] > 0 and c["days_elapsed"] >= c["tat_days"]:
            ot = c.get("owner_type") or "Unassigned"
            owner_delay[ot] = owner_delay.get(ot, 0) + 1

    delay_rows = "".join(
        f'<tr><td style="font-weight:500">{ot}</td>'
        f'<td class="text-center"><span class="pill pill-danger">{cnt}</span></td></tr>'
        for ot, cnt in sorted(owner_delay.items(), key=lambda x: -x[1])
    ) or '<tr><td colspan="2" style="text-align:center;color:#94a3b8;padding:20px">No overdue cases</td></tr>'

    chart_labels_json = json.dumps(chart_labels)
    chart_sla_json = json.dumps(chart_sla)
    chart_overdue_json = json.dumps(chart_overdue)
    monthly_labels_json = json.dumps(monthly_labels)
    monthly_new_json = json.dumps(monthly_new)

    content = f"""
<div class="row g-3 mb-4">
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between">
        <div><div class="stat-val">{total_cases}</div><div class="stat-label">Total Active Cases</div></div>
        <div class="stat-icon" style="background:#e1eef8;color:#003356"><i class="bi bi-folder2-open"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between">
        <div><div class="stat-val" style="color:{sla_color}">{sla_pct}%</div><div class="stat-label">Overall SLA Compliance</div></div>
        <div class="stat-icon" style="background:{sla_color}15;color:{sla_color}"><i class="bi bi-shield-check"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between">
        <div><div class="stat-val" style="color:#dc2626">{len(overdue)}</div><div class="stat-label">Overdue Cases</div></div>
        <div class="stat-icon" style="background:#fee2e2;color:#dc2626"><i class="bi bi-exclamation-triangle-fill"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between">
        <div><div class="stat-val" style="color:var(--accent)">{avg_elapsed}d</div><div class="stat-label">Avg Days in Stage</div></div>
        <div class="stat-icon" style="background:#d0f0ff;color:var(--accent)"><i class="bi bi-clock-history"></i></div>
      </div>
    </div>
  </div>
</div>

<div class="row g-3 mb-4">
  <div class="col-lg-7">
    <div class="card">
      <div class="card-header"><i class="bi bi-shield-fill-check" style="color:var(--accent)"></i> Per-Programme SLA Scorecard</div>
      <div style="overflow-x:auto">
        <table class="data-table">
          <thead><tr><th>Programme</th><th style="text-align:center">Cases</th>
            <th style="text-align:center">On Track</th><th style="text-align:center">At Risk</th>
            <th style="text-align:center">Overdue</th><th>SLA %</th></tr></thead>
          <tbody>{scorecard_rows if scorecard_rows else '<tr><td colspan="6" style="text-align:center;color:#94a3b8;padding:20px">No data</td></tr>'}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="col-lg-5">
    <div class="card h-100">
      <div class="card-header"><i class="bi bi-bar-chart-fill" style="color:#7c3aed"></i> Overdue by Owner Type</div>
      <div class="card-body">
        <table class="data-table">
          <thead><tr><th>Owner Type</th><th style="text-align:center">Overdue Cases</th></tr></thead>
          <tbody>{delay_rows}</tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<div class="row g-3">
  <div class="col-lg-8">
    <div class="card">
      <div class="card-header"><i class="bi bi-bar-chart-line" style="color:var(--accent)"></i> Programme SLA Compliance (%)</div>
      <div class="card-body"><canvas id="slaChart" height="180"></canvas></div>
    </div>
  </div>
  <div class="col-lg-4">
    <div class="card">
      <div class="card-header"><i class="bi bi-graph-up" style="color:#00984C"></i> Monthly New Cases (6 mo)</div>
      <div class="card-body"><canvas id="trendChart" height="180"></canvas></div>
    </div>
  </div>
</div>
"""
    scripts = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script>
(function(){{
  var labels = {chart_labels_json};
  var slaData = {chart_sla_json};
  var colors = slaData.map(v => v>=80?'#00984C':v>=60?'#d97706':'#dc2626');
  new Chart(document.getElementById('slaChart'), {{
    type: 'bar',
    data: {{ labels: labels, datasets: [{{
      label: 'SLA %', data: slaData, backgroundColor: colors, borderRadius: 5
    }}]}},
    options: {{ plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ min:0, max:100 }} }},
      responsive: true }}
  }});
  var mLabels = {monthly_labels_json};
  var mData = {monthly_new_json};
  new Chart(document.getElementById('trendChart'), {{
    type: 'line',
    data: {{ labels: mLabels, datasets: [{{
      label: 'New Cases', data: mData, borderColor:'#0094ca', backgroundColor:'rgba(0,148,202,.1)',
      fill:true, tension:0.3, pointRadius:4
    }}]}},
    options: {{ plugins: {{ legend: {{ display: false }} }}, responsive: true }}
  }});
}})();
</script>"""
    return render_page(content, scripts, active_page="ceo_dashboard", page_title="CEO Dashboard")


@app.route("/")
@login_required
def dashboard():
    conn = get_db()
    prog_filter   = request.args.get("programme", "")
    owner_filter  = request.args.get("owner_type", "")
    my_cases      = request.args.get("my_cases", "")
    sort          = request.args.get("sort", "elapsed_desc")
    status_filter = request.args.get("status", "Active")  # default: only Active

    bid = user_board_id()
    ph_progs = user_programme_names()  # None unless program_head
    if bid is not None:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
        base_q = "SELECT * FROM case_tracking WHERE board_id=?"
        base_params = [bid]
    else:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes ORDER BY programme_name"
        ).fetchall()]
        base_q = "SELECT * FROM case_tracking WHERE 1=1"
        base_params = []

    # Restrict program_head to their mapped programmes
    if ph_progs is not None:
        programmes = [p for p in programmes if p in ph_progs]
        if ph_progs:
            placeholders = ",".join("?" * len(ph_progs))
            base_q += f" AND programme_name IN ({placeholders})"
            base_params.extend(ph_progs)
        else:
            base_q += " AND 1=0"  # no programmes mapped → no cases

    if status_filter and status_filter != "All":
        base_q += " AND (status=? OR (status IS NULL AND ?='Active'))"
        base_params.extend([status_filter, status_filter])
    if prog_filter:
        base_q += " AND programme_name=?"
        base_params.append(prog_filter)
    if owner_filter:
        base_q += " AND owner_type=?"
        base_params.append(owner_filter)
    if my_cases:
        po_email = conn.execute(
            "SELECT email FROM users WHERE id=?", (session["user_id"],)
        ).fetchone()
        if po_email and po_email["email"]:
            base_q += " AND program_officer_email=?"
            base_params.append(po_email["email"])
        else:
            # No email on this account — return zero results with a flash hint
            base_q += " AND 1=0"
            flash("Your account has no email address set. Add one in your profile so 'My Cases' can filter correctly.", "info")

    # ── Stat card counts — lightweight query on full filtered set ──
    stat_q = f"SELECT stage_start_date, tat_days, reminder2_day, is_milestone FROM ({base_q}) _sub"
    stat_rows = [dict(r) for r in conn.execute(stat_q, base_params).fetchall()]

    today = now_ist().date()
    n_total    = len(stat_rows)
    n_overdue  = 0
    n_at_risk  = 0
    n_milestone= 0
    for _s in stat_rows:
        if _s["is_milestone"]:
            n_milestone += 1
        elif _s["tat_days"] > 0:
            _el = working_days_elapsed(_s["stage_start_date"], today)
            if _el >= _s["tat_days"]:
                n_overdue += 1
            elif _el >= _s["reminder2_day"]:
                n_at_risk += 1
    n_ok = n_total - n_overdue - n_at_risk - n_milestone

    # ── Pagination — fetch only current page from the DB ──
    PAGE_SIZE   = 50
    page        = max(1, int(request.args.get("page", 1)))
    total_count = n_total
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = min(page, total_pages)

    # SQL sort (calendar-day approximation; exact business days computed for page rows only)
    if sort == "elapsed_desc":
        order_clause = "ORDER BY stage_start_date ASC"
    elif sort == "elapsed_asc":
        order_clause = "ORDER BY stage_start_date DESC"
    elif sort in ("app_desc", "app_asc"):
        order_clause = f"ORDER BY application_id {'DESC' if sort.endswith('desc') else 'ASC'}"
    elif sort in ("org_desc", "org_asc"):
        order_clause = f"ORDER BY organisation_name {'DESC' if sort.endswith('desc') else 'ASC'}"
    else:
        order_clause = "ORDER BY stage_start_date ASC"

    page_q      = f"{base_q} {order_clause} LIMIT ? OFFSET ?"
    page_params = base_params + [PAGE_SIZE, (page - 1) * PAGE_SIZE]
    page_cases  = [dict(r) for r in conn.execute(page_q, page_params).fetchall()]
    conn.close()

    # Exact business-day elapsed only for the 50 rows on this page
    for c in page_cases:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, hold_days=c.get("hold_days", 0))

    cases = page_cases  # used by rows_html loop below

    rows_html = ""
    for c in page_cases:
        elapsed = c["days_elapsed"]
        tat     = c["tat_days"]

        if c["is_milestone"]:
            status_pill = '<span class="pill pill-milestone"><i class="bi bi-flag-fill"></i> Milestone</span>'
            tr_cls = ""
            bar_html = ""
        elif tat > 0 and elapsed >= tat:
            status_pill = f'<span class="pill pill-danger"><i class="bi bi-exclamation-triangle-fill"></i> Overdue · {elapsed}d</span>'
            tr_cls = "row-overdue"
            pct = 100
            bar_html = f'<div class="tat-bar mt-1"><div class="tat-bar-fill" style="width:{pct}%;background:#dc2626"></div></div>'
        elif tat > 0 and elapsed >= c["reminder2_day"]:
            status_pill = f'<span class="pill pill-warn"><i class="bi bi-clock"></i> At Risk · {elapsed}d</span>'
            tr_cls = ""
            pct = min(99, int(elapsed / tat * 100)) if tat else 0
            bar_html = f'<div class="tat-bar mt-1"><div class="tat-bar-fill" style="width:{pct}%;background:#d97706"></div></div>'
        else:
            status_pill = f'<span class="pill pill-ok"><i class="bi bi-check-circle"></i> On Track · {elapsed}d</span>'
            tr_cls = ""
            pct = min(80, int(elapsed / tat * 100)) if tat else 0
            bar_html = f'<div class="tat-bar mt-1"><div class="tat-bar-fill" style="width:{pct}%;background:#00984C"></div></div>'

        def yn(v):
            return '<i class="bi bi-check-circle-fill sent-yes"></i>' if v else '<i class="bi bi-circle sent-no"></i>'

        owner_type_badge = ""
        if c.get("owner_type"):
            colors = {"Applicant": "#0094ca", "Assessor": "#7c3aed", "Program Officer": "#00984C"}
            bg = colors.get(c["owner_type"], "#64748b")
            owner_type_badge = f'<span style="font-size:10px;padding:1px 7px;border-radius:4px;background:{bg}22;color:{bg};font-weight:600">{c["owner_type"]}</span> '

        case_st = c.get("status") or "Active"
        st_colors = {"Active": ("#00984C", "#d8f5e7"), "On Hold": ("#7c3aed", "#ede9fe"),
                     "Closed": ("#64748b", "#f1f5f9"),
                     "Withdrawn": ("#d97706", "#fef3c7"), "Suspended": ("#dc2626", "#fee2e2")}
        st_fg, st_bg = st_colors.get(case_st, ("#64748b", "#f1f5f9"))
        case_status_badge = f'<span style="font-size:10px;padding:1px 7px;border-radius:4px;background:{st_bg};color:{st_fg};font-weight:600">{case_st}</span>'
        _app_id_h  = h(c['application_id'])
        _org_h     = h(c['organisation_name'])
        _prog_h    = h(c['programme_name'])
        _stage_h   = h(c['current_stage'])
        _owner_h   = h(c['action_owner_name']) if c['action_owner_name'] else '—'
        _email_h   = h(c['action_owner_email']) if c['action_owner_email'] else ''
        _sdate_h   = h(c['stage_start_date'])
        _case_st_h = h(case_st)
        rows_html += f"""<tr class="{tr_cls}">
          <td class="id-cell">{_app_id_h}<br>{case_status_badge}</td>
          <td><div style="font-weight:500;color:#1e293b">{_org_h}</div>
              <div style="font-size:11px;color:#94a3b8">{_prog_h}</div></td>
          <td><div style="font-size:13px">{_stage_h}</div>
              <div style="font-size:11px;color:#94a3b8">{owner_type_badge}Start: {_sdate_h}</div></td>
          <td><div>{status_pill}</div>{bar_html}</td>
          <td class="text-center">{yn(c['r1_sent'])}</td>
          <td class="text-center">{yn(c['r2_sent'])}</td>
          <td class="text-center">{yn(c['overdue_sent'])}</td>
          <td class="text-center" style="font-weight:600;color:{'#dc2626' if c['overdue_count'] else '#94a3b8'}">{c['overdue_count'] or '—'}</td>
          <td><div style="font-size:12.5px">{_owner_h}</div>
              <div style="font-size:11px;color:#94a3b8">{_email_h}</div></td>
          <td style="white-space:nowrap">
            <a href="/case-history/{_app_id_h}" class="btn btn-sm btn-action btn-outline-secondary me-1"
               title="History"><i class="bi bi-clock-history"></i></a>
            <button class="btn btn-sm btn-action btn-outline-success me-1" title="Quick Advance"
               onclick="openQuickAdvance({c['id']}, {json.dumps(c['application_id'])}, {json.dumps(c['programme_name'])})">
               <i class="bi bi-arrow-right-circle"></i></button>
            <a href="/edit-case/{c['id']}" class="btn btn-sm btn-action btn-outline-primary me-1">Edit</a>
            <button class="btn btn-sm btn-action btn-outline-warning me-1" title="Change Status"
               onclick="openStatusModal({c['id']}, {json.dumps(c['application_id'])}, {json.dumps(case_st)})">
               <i class="bi bi-toggles"></i></button>
            <button class="btn btn-sm btn-action btn-outline-danger"
              onclick="confirmDelete('/delete-case/{c['id']}', {json.dumps(c['application_id'])})">Delete</button>
          </td>
        </tr>"""

    opt_programmes = "".join(
        f'<option value="{p}" {"selected" if p==prog_filter else ""}>{p}</option>'
        for p in programmes
    )

    stat_cards = f"""
<div class="row g-3 mb-4">
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between align-items-start">
        <div>
          <div class="stat-val">{n_total}</div>
          <div class="stat-label">Total Active Cases</div>
        </div>
        <div class="stat-icon" style="background:#e1eef8;color:#003356"><i class="bi bi-folder2-open"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between align-items-start">
        <div>
          <div class="stat-val" style="color:#00984C">{n_ok}</div>
          <div class="stat-label">On Track</div>
        </div>
        <div class="stat-icon" style="background:#d8f5e7;color:#00984C"><i class="bi bi-check-circle-fill"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between align-items-start">
        <div>
          <div class="stat-val" style="color:#d97706">{n_at_risk}</div>
          <div class="stat-label">At Risk</div>
        </div>
        <div class="stat-icon" style="background:#fef3c7;color:#d97706"><i class="bi bi-clock-fill"></i></div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="stat-card">
      <div class="d-flex justify-content-between align-items-start">
        <div>
          <div class="stat-val" style="color:#dc2626">{n_overdue}</div>
          <div class="stat-label">Overdue</div>
        </div>
        <div class="stat-icon" style="background:#fee2e2;color:#dc2626"><i class="bi bi-exclamation-triangle-fill"></i></div>
      </div>
    </div>
  </div>
</div>"""

    empty_html = """
<div style="text-align:center;padding:60px 20px;color:#94a3b8">
  <i class="bi bi-inbox" style="font-size:48px;display:block;margin-bottom:12px"></i>
  <div style="font-size:15px;font-weight:500">No cases found</div>
  <div style="font-size:13px;margin-top:4px">Log a stage change or upload a CSV to get started.</div>
  <a href="/log-stage" class="btn btn-primary mt-3 btn-sm">Log a Stage</a>
</div>"""

    # ── Owner-type pendency breakdown ──
    owner_counts = {}
    owner_overdue = {}
    for c in cases:
        ot = c.get("owner_type") or "Unassigned"
        owner_counts[ot] = owner_counts.get(ot, 0) + 1
        if not c["is_milestone"] and c["tat_days"] > 0 and c["days_elapsed"] >= c["tat_days"]:
            owner_overdue[ot] = owner_overdue.get(ot, 0) + 1

    OWNER_META = {
        "Applicant":       ("#0094ca", "bi-person-fill"),
        "Assessor":        ("#7c3aed", "bi-clipboard-check"),
        "Program Officer": ("#00984C", "bi-headset"),
        "Unassigned":      ("#94a3b8", "bi-question-circle"),
    }

    owner_cards_html = ""
    for ot in ["Applicant", "Assessor", "Program Officer", "Unassigned"]:
        cnt = owner_counts.get(ot, 0)
        if cnt == 0 and ot == "Unassigned":
            continue
        od = owner_overdue.get(ot, 0)
        color, icon = OWNER_META.get(ot, ("#94a3b8", "bi-circle"))
        od_badge = f'<span style="font-size:11px;padding:2px 8px;border-radius:6px;background:#fee2e2;color:#dc2626;font-weight:600">{od} overdue</span>' if od else '<span style="font-size:11px;color:#94a3b8">0 overdue</span>'
        pct = int(cnt / n_total * 100) if n_total else 0
        owner_cards_html += f"""<div class="col-6 col-lg-3">
  <div class="stat-card" style="border-left:4px solid {color}">
    <div class="d-flex justify-content-between align-items-start">
      <div>
        <div class="stat-val" style="font-size:24px;color:{color}">{cnt}</div>
        <div class="stat-label">{ot}</div>
        <div style="margin-top:6px">{od_badge}</div>
      </div>
      <div class="stat-icon" style="background:{color}15;color:{color}"><i class="bi {icon}"></i></div>
    </div>
    <div class="tat-bar mt-2" style="width:100%"><div class="tat-bar-fill" style="width:{pct}%;background:{color}"></div></div>
    <div style="font-size:10px;color:#94a3b8;margin-top:2px">{pct}% of total</div>
  </div>
</div>"""

    # ── Programme-wise summary for analytics ──
    prog_summary = {}
    for c in cases:
        pn = c["programme_name"]
        if pn not in prog_summary:
            prog_summary[pn] = {"total": 0, "overdue": 0, "on_track": 0, "at_risk": 0}
        prog_summary[pn]["total"] += 1
        if c["is_milestone"]:
            continue
        if c["tat_days"] > 0 and c["days_elapsed"] >= c["tat_days"]:
            prog_summary[pn]["overdue"] += 1
        elif c["tat_days"] > 0 and c["days_elapsed"] >= c["reminder2_day"]:
            prog_summary[pn]["at_risk"] += 1
        else:
            prog_summary[pn]["on_track"] += 1

    prog_rows_html = ""
    for pn, ps in sorted(prog_summary.items()):
        health = ps["on_track"] / ps["total"] * 100 if ps["total"] else 0
        prog_rows_html += f"""<tr>
  <td style="font-weight:600;font-size:12.5px">{pn}</td>
  <td class="text-center">{ps['total']}</td>
  <td class="text-center" style="color:#00984C;font-weight:600">{ps['on_track']}</td>
  <td class="text-center" style="color:#d97706;font-weight:600">{ps['at_risk']}</td>
  <td class="text-center" style="color:#dc2626;font-weight:600">{ps['overdue']}</td>
  <td>
    <div class="tat-bar" style="width:100%;height:7px">
      <div class="tat-bar-fill" style="width:{health:.0f}%;background:#00984C"></div>
    </div>
    <div style="font-size:10px;color:#94a3b8">{health:.0f}% healthy</div>
  </td>
</tr>"""

    # ── Ageing analysis ──
    age_buckets = {"0-7 days": 0, "8-15 days": 0, "16-30 days": 0, "31-60 days": 0, "60+ days": 0}
    for c in cases:
        e = c["days_elapsed"]
        if e <= 7:
            age_buckets["0-7 days"] += 1
        elif e <= 15:
            age_buckets["8-15 days"] += 1
        elif e <= 30:
            age_buckets["16-30 days"] += 1
        elif e <= 60:
            age_buckets["31-60 days"] += 1
        else:
            age_buckets["60+ days"] += 1

    AGE_COLORS = {"0-7 days": "#00984C", "8-15 days": "#0094ca", "16-30 days": "#d97706", "31-60 days": "#f93822", "60+ days": "#dc2626"}
    age_bars_html = ""
    for lbl, cnt in age_buckets.items():
        pct = int(cnt / n_total * 100) if n_total else 0
        clr = AGE_COLORS[lbl]
        age_bars_html += f"""<div class="d-flex align-items-center gap-2 mb-2">
  <div style="width:70px;font-size:12px;font-weight:500;color:#475569">{lbl}</div>
  <div style="flex:1;height:20px;background:#f1f5f9;border-radius:4px;overflow:hidden">
    <div style="width:{pct}%;height:100%;background:{clr};border-radius:4px;min-width:{('24px' if cnt else '0')}"></div>
  </div>
  <div style="width:36px;text-align:right;font-size:12px;font-weight:600;color:#475569">{cnt}</div>
</div>"""

    # ── SLA compliance metric ──
    sla_eligible = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0]
    sla_within = sum(1 for c in sla_eligible if c["days_elapsed"] < c["tat_days"])
    sla_pct = int(sla_within / len(sla_eligible) * 100) if sla_eligible else 100
    sla_color = "#00984C" if sla_pct >= 80 else ("#d97706" if sla_pct >= 60 else "#dc2626")

    # ── Notification coverage ──
    notif_eligible = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0]
    r1_coverage = sum(1 for c in notif_eligible if c["r1_sent"]) if notif_eligible else 0
    r1_pct = int(r1_coverage / len(notif_eligible) * 100) if notif_eligible else 0

    analytics_section = f"""
<div class="row g-3 mb-4">
  <div class="col-12">
    <div style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--muted);margin-bottom:8px">
      <i class="bi bi-people-fill"></i> Pendency by Owner Type
    </div>
  </div>
  {owner_cards_html}
</div>

<div class="row g-3 mb-4">
  <div class="col-lg-7">
    <div class="card">
      <div class="card-header"><i class="bi bi-bar-chart-fill" style="color:var(--accent)"></i> Programme Health</div>
      <div style="overflow-x:auto">
        <table class="data-table">
          <thead><tr>
            <th>Programme</th>
            <th style="text-align:center">Total</th>
            <th style="text-align:center">On Track</th>
            <th style="text-align:center">At Risk</th>
            <th style="text-align:center">Overdue</th>
            <th>Health</th>
          </tr></thead>
          <tbody>{prog_rows_html if prog_rows_html else '<tr><td colspan="6" style="text-align:center;color:#94a3b8;padding:20px">No data</td></tr>'}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="col-lg-5">
    <div class="card mb-3">
      <div class="card-header"><i class="bi bi-hourglass-split" style="color:#d97706"></i> Case Ageing</div>
      <div class="card-body">
        {age_bars_html}
      </div>
    </div>
    <div class="card">
      <div class="card-header"><i class="bi bi-shield-check" style="color:{sla_color}"></i> SLA Compliance</div>
      <div class="card-body text-center" style="padding:20px">
        <div style="font-size:48px;font-weight:700;color:{sla_color};line-height:1">{sla_pct}%</div>
        <div style="font-size:12px;color:#64748b;margin-top:4px">of cases within TAT</div>
        <div class="tat-bar mt-3 mx-auto" style="width:80%;height:8px">
          <div class="tat-bar-fill" style="width:{sla_pct}%;background:{sla_color}"></div>
        </div>
        <div class="d-flex justify-content-between mt-2" style="font-size:11px;color:#94a3b8;width:80%;margin:0 auto">
          <span>{sla_within} within TAT</span>
          <span>{len(sla_eligible) - sla_within} breached</span>
        </div>
        <hr style="margin:16px 0;border-color:#f1f5f9">
        <div class="d-flex justify-content-center gap-4">
          <div>
            <div style="font-size:20px;font-weight:700;color:var(--accent)">{r1_pct}%</div>
            <div style="font-size:11px;color:#94a3b8">R1 Coverage</div>
          </div>
          <div>
            <div style="font-size:20px;font-weight:700;color:var(--navy)">{len(sla_eligible)}</div>
            <div style="font-size:11px;color:#94a3b8">TAT-tracked</div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>
"""

    # Overdue banner
    overdue_banner_html = ""
    if n_overdue > 0 and not prog_filter and not owner_filter and not my_cases:
        overdue_banner_html = f"""
<div class="overdue-banner">
  <i class="bi bi-exclamation-triangle-fill" style="font-size:18px"></i>
  <span><strong>{n_overdue} case{"s" if n_overdue!=1 else ""} overdue</strong> — immediate attention required.</span>
  <a href="/?owner_type=Applicant">View Applicant</a>
  <a href="/?owner_type=Assessor">View Assessor</a>
  <a href="/?owner_type=Program+Officer">View PO</a>
</div>"""

    # Load saved filters for current user
    conn2 = get_db()
    saved_filters_list = [dict(r) for r in conn2.execute(
        "SELECT id, filter_name, filter_json FROM saved_filters WHERE user_id=? ORDER BY filter_name",
        (session["user_id"],)
    ).fetchall()]
    conn2.close()
    saved_filter_opts = "".join(
        f'<option value="{h(f["filter_json"])}">{h(f["filter_name"])} '
        f'<a data-id="{f["id"]}" onclick="deleteFilter({f["id"]})">×</a></option>'
        for f in saved_filters_list
    )
    saved_filter_btns = "".join(
        f'<a href="/?{json.loads(f["filter_json"]).get("qs","")}" '
        f'class="btn btn-sm btn-outline-secondary" style="font-size:12px">'
        f'<i class="bi bi-bookmark-fill" style="color:#7c3aed"></i> {f["filter_name"]}</a>'
        for f in saved_filters_list
    )

    # Build pagination HTML
    if total_pages > 1:
        prev_page = max(1, page - 1)
        next_page = min(total_pages, page + 1)
        def page_url(p):
            args = request.args.to_dict()
            args["page"] = p
            return url_for("dashboard", **args)
        _start_idx = (page - 1) * PAGE_SIZE
        _pagination_html = f'<div class="d-flex justify-content-between align-items-center mt-3 px-1"><div style="font-size:12px;color:#94a3b8">Showing {_start_idx+1}–{min(_start_idx+PAGE_SIZE, total_count)} of {total_count} cases</div><nav><ul class="pagination pagination-sm mb-0"><li class="page-item {"disabled" if page==1 else ""}"><a class="page-link" href="{page_url(prev_page)}">&#8249; Prev</a></li>'
        for _p in range(max(1, page-2), min(total_pages+1, page+3)):
            _pagination_html += f'<li class="page-item {"active" if _p==page else ""}"><a class="page-link" href="{page_url(_p)}">{_p}</a></li>'
        _pagination_html += f'<li class="page-item {"disabled" if page==total_pages else ""}"><a class="page-link" href="{page_url(next_page)}">Next &#8250;</a></li></ul></nav></div>'
    else:
        _pagination_html = ""

    content = f"""
{overdue_banner_html}
{stat_cards}
{analytics_section}

<div class="card">
  <div class="card-header d-flex justify-content-between align-items-center flex-wrap gap-2">
    <span>Cases <span style="font-size:12px;font-weight:400;color:#94a3b8;margin-left:6px">{total_count} total</span>
    {"&nbsp;" + saved_filter_btns if saved_filter_btns else ""}
    </span>
    <form method="get" class="d-flex gap-2 align-items-center flex-wrap" id="filterForm">
      <select name="status" class="form-select form-select-sm" style="width:auto">
        <option value="Active" {"selected" if status_filter=="Active" else ""}>Active</option>
        <option value="On Hold" {"selected" if status_filter=="On Hold" else ""}>On Hold</option>
        <option value="Closed" {"selected" if status_filter=="Closed" else ""}>Closed</option>
        <option value="Withdrawn" {"selected" if status_filter=="Withdrawn" else ""}>Withdrawn</option>
        <option value="Suspended" {"selected" if status_filter=="Suspended" else ""}>Suspended</option>
        <option value="All" {"selected" if status_filter=="All" else ""}>All Statuses</option>
      </select>
      <select name="programme" class="form-select form-select-sm" style="width:auto">
        <option value="">All Programmes</option>
        {opt_programmes}
      </select>
      <select name="owner_type" class="form-select form-select-sm" style="width:auto">
        <option value="">All Owners</option>
        <option value="Applicant" {"selected" if owner_filter=="Applicant" else ""}>Applicant</option>
        <option value="Assessor" {"selected" if owner_filter=="Assessor" else ""}>Assessor</option>
        <option value="Program Officer" {"selected" if owner_filter=="Program Officer" else ""}>Program Officer</option>
      </select>
      <select name="sort" class="form-select form-select-sm" style="width:auto">
        <option value="elapsed_desc" {"selected" if sort=="elapsed_desc" else ""}>Days Elapsed ↓</option>
        <option value="elapsed_asc"  {"selected" if sort=="elapsed_asc" else ""}>Days Elapsed ↑</option>
        <option value="app_asc"      {"selected" if sort=="app_asc" else ""}>Application ID</option>
        <option value="org_asc"      {"selected" if sort=="org_asc" else ""}>Organisation</option>
      </select>
      <button class="btn btn-sm btn-primary" type="submit">Filter</button>
      <button type="button" class="btn btn-sm btn-outline-secondary" onclick="saveCurrentFilter()" title="Save this filter">
        <i class="bi bi-bookmark-plus"></i>
      </button>
      {"" if not (prog_filter or owner_filter or status_filter not in ("Active","")) else '<a class="btn btn-sm btn-outline-secondary" href="/">Clear</a>'}
    </form>
  </div>
  <div style="overflow-x:auto">
    <table class="data-table">
      <thead>
        <tr>
          <th>Application ID</th>
          <th>Organisation / Programme</th>
          <th>Stage</th>
          <th>Status</th>
          <th style="text-align:center">R1</th>
          <th style="text-align:center">R2</th>
          <th style="text-align:center">Overdue</th>
          <th style="text-align:center">Follow-ups</th>
          <th>Action Owner</th>
          <th>Actions</th>
        </tr>
      </thead>
      <tbody>
        {rows_html if rows_html else f'<tr><td colspan="10">{empty_html}</td></tr>'}
      </tbody>
    </table>
  </div>
  {_pagination_html}
</div>

<!-- Delete confirmation modal -->
<div class="modal fade" id="deleteModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered modal-sm">
    <div class="modal-content">
      <div class="modal-body text-center p-4">
        <i class="bi bi-trash3" style="font-size:32px;color:#dc2626"></i>
        <div style="font-weight:600;margin-top:10px">Close Case?</div>
        <div style="font-size:13px;color:#64748b;margin-top:4px" id="deleteModalMsg"></div>
      </div>
      <div class="modal-footer border-0 pt-0 justify-content-center gap-2">
        <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
        <a id="deleteConfirmBtn" href="#" class="btn btn-sm btn-danger">Close Case</a>
      </div>
    </div>
  </div>
</div>

<!-- Quick Advance modal -->
<div class="modal fade" id="quickAdvanceModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header border-0 pb-0">
        <h6 class="modal-title"><i class="bi bi-arrow-right-circle-fill" style="color:#059669"></i>
          Quick Stage Advance — <span id="qa_app_label" style="color:#64748b;font-weight:400"></span></h6>
        <button class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <form method="post" action="/quick-advance">
        <input type="hidden" name="case_id" id="qa_case_id">
        <div class="modal-body">
          <div class="mb-3">
            <label class="form-label">Move to Stage</label>
            <select class="form-select" id="qa_stage_select" name="target_stage" required>
              <option value="">Loading…</option>
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label">Start Date <span style="color:#94a3b8;font-size:11px">(new stage)</span></label>
            <input type="date" class="form-control" name="new_start_date"
                   value="{date.today().strftime('%Y-%m-%d')}">
          </div>
          <div class="mb-1">
            <label class="form-label">Suppress Notifications Until <span style="color:#94a3b8;font-size:11px">(optional)</span></label>
            <input type="date" class="form-control" name="suppress_until">
          </div>
        </div>
        <div class="modal-footer border-0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button class="btn btn-sm btn-success" type="submit">
            <i class="bi bi-arrow-right-circle"></i> Advance Stage
          </button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Case Status Change Modal -->
<div class="modal fade" id="statusModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered modal-sm">
    <div class="modal-content">
      <div class="modal-header border-0">
        <h6 class="modal-title"><i class="bi bi-toggles" style="color:#7c3aed"></i> Change Case Status</h6>
        <button class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <form method="post" action="/update-case-status">
        <input type="hidden" name="application_id" id="sc_case_id">
        <div class="modal-body pt-0">
          <div style="font-size:13px;color:#64748b;margin-bottom:12px">Case: <strong id="sc_app_id"></strong></div>
          <select class="form-select" name="status" id="sc_status_select">
            <option value="Active">Active</option>
            <option value="On Hold">&#9646;&#9646; On Hold (TAT paused)</option>
            <option value="Closed">Closed</option>
            <option value="Withdrawn">Withdrawn</option>
            <option value="Suspended">Suspended</option>
          </select>
        </div>
        <div class="modal-footer border-0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button class="btn btn-sm btn-primary" type="submit">Update Status</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Save Filter Modal -->
<div class="modal fade" id="saveFilterModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered modal-sm">
    <div class="modal-content">
      <div class="modal-header border-0">
        <h6 class="modal-title"><i class="bi bi-bookmark-plus" style="color:#7c3aed"></i> Save Filter</h6>
        <button class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body pt-0">
        <label class="form-label">Filter name</label>
        <input type="text" class="form-control" id="filterNameInput" placeholder="e.g. My NABH Overdue">
      </div>
      <div class="modal-footer border-0">
        <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
        <button class="btn btn-sm btn-primary" onclick="confirmSaveFilter()">Save</button>
      </div>
    </div>
  </div>
</div>

<!-- Run check result toast -->
<div id="runToast" class="toast qci-toast" role="alert" style="position:fixed;bottom:24px;right:24px;z-index:9999;min-width:320px">
  <div class="toast-header">
    <i class="bi bi-play-circle-fill text-success me-2"></i>
    <strong class="me-auto">Check Complete</strong>
    <button type="button" class="btn-close" data-bs-dismiss="toast"></button>
  </div>
  <div class="toast-body" id="runToastBody"></div>
</div>
"""
    topbar_actions = """
<a href="/export" class="btn btn-sm btn-outline-secondary">
  <i class="bi bi-download"></i> Export CSV
</a>
<a href="/log-stage" class="btn btn-sm btn-navy">
  <i class="bi bi-plus-circle"></i> Log Stage
</a>
<button class="btn btn-sm btn-warning" id="runBtn">
  <i class="bi bi-play-fill"></i> Run Check
</button>"""

    scripts = """<script>
function confirmDelete(url, appId){
  document.getElementById('deleteModalMsg').textContent = 'Delete case ' + appId + '? This cannot be undone.';
  document.getElementById('deleteConfirmBtn').href = url;
  new bootstrap.Modal(document.getElementById('deleteModal')).show();
}
function openStatusModal(caseId, appId, currentStatus){
  document.getElementById('sc_case_id').value = appId;
  document.getElementById('sc_app_id').textContent = appId;
  document.getElementById('sc_status_select').value = currentStatus;
  new bootstrap.Modal(document.getElementById('statusModal')).show();
}
function saveCurrentFilter(){
  new bootstrap.Modal(document.getElementById('saveFilterModal')).show();
}
function confirmSaveFilter(){
  var name = document.getElementById('filterNameInput').value.trim();
  if(!name){ alert('Please enter a name.'); return; }
  var form = document.getElementById('filterForm');
  var data = new FormData(form);
  var qs = new URLSearchParams(data).toString();
  fetch('/save-filter', {method:'POST',headers:{'Content-Type':'application/json'},
    body: JSON.stringify({name: name, qs: qs})
  }).then(r=>r.json()).then(d=>{
    if(d.ok){ bootstrap.Modal.getInstance(document.getElementById('saveFilterModal')).hide(); location.reload(); }
    else { alert(d.error || 'Error saving filter'); }
  });
}
function deleteFilter(id){
  if(!confirm('Delete this saved filter?')) return;
  fetch('/delete-filter/'+id, {method:'POST'}).then(()=>location.reload());
}
document.getElementById('runBtn').addEventListener('click', function(){
  var btn = this;
  btn.innerHTML = '<span class="spinner-border spinner-border-sm"></span>';
  btn.disabled = true;
  fetch('/run-check').then(r=>r.json()).then(d=>{
    var msg = 'R1: <strong>'+d.r1+'</strong> &nbsp; R2: <strong>'+d.r2+'</strong> &nbsp; Overdue: <strong>'+d.overdue+'</strong> &nbsp; Follow-ups: <strong>'+d.followup+'</strong>';
    if(d.errors && d.errors.length) msg += '<br><small class="text-danger">'+d.errors.length+' error(s)</small>';
    document.getElementById('runToastBody').innerHTML = msg;
    new bootstrap.Toast(document.getElementById('runToast')).show();
    btn.innerHTML = '<i class="bi bi-play-fill"></i> Run Check';
    btn.disabled = false;
    setTimeout(function(){ location.reload(); }, 2500);
  }).catch(function(err){
    btn.innerHTML = '<i class="bi bi-play-fill"></i> Run Check';
    btn.disabled = false;
  });
});
</script>"""
    return render_page(content, scripts, active_page="dashboard",
                       page_title="Dashboard", topbar_actions=topbar_actions)


@app.route("/log-stage", methods=["GET", "POST"])
@login_required
def log_stage():
    conn = get_db()
    bid = user_board_id()
    if bid is not None:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
    else:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes ORDER BY programme_name"
        ).fetchall()]
    conn.close()

    if request.method == "POST":
        data = {
            "application_id":     request.form["application_id"].strip(),
            "organisation_name":  request.form["organisation_name"].strip(),
            "programme_name":     request.form["programme_name"].strip(),
            "stage_name":         request.form["stage_name"].strip(),
            "stage_start_date":   request.form["stage_start_date"].strip(),
            "action_owner_name":  request.form["action_owner_name"].strip(),
            "action_owner_email": request.form["action_owner_email"].strip(),
            "program_officer_email": request.form["program_officer_email"].strip(),
        }
        try:
            data["_changed_by"] = session.get("full_name") or session.get("username", "")
            action = upsert_case(data)
            flash(f"Case {data['application_id']} {action} successfully.", "success")
            return redirect(url_for("log_stage"))
        except ValueError as e:
            flash(str(e), "error")

    opts = "".join(f'<option value="{p}">{p}</option>' for p in programmes)
    content = f"""
<div class="row g-4 justify-content-center" style="max-width:860px;margin:0 auto">

  <!-- Left column: Case details -->
  <div class="col-lg-7">
    <div class="card form-card">
      <div class="card-header">
        <i class="bi bi-file-earmark-text" style="color:#2563eb"></i> Case Details
      </div>
      <div class="card-body p-4">
        <form method="post" id="logForm">

          <div class="form-section-title">Programme &amp; Stage</div>
          <div class="mb-3">
            <label class="form-label">Programme</label>
            <select class="form-select" name="programme_name" id="progSelect" required>
              <option value="">Select a programme…</option>
              {opts}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label">Stage</label>
            <select class="form-select" name="stage_name" id="stageSelect" required>
              <option value="">Select programme first…</option>
            </select>
            <div class="stage-info-box" id="stageInfoBox">
              <div class="d-flex gap-3 flex-wrap">
                <div><span style="font-size:11px;color:#64748b">TAT</span><br><strong id="si_tat">—</strong> days</div>
                <div><span style="font-size:11px;color:#64748b">Reminder 1</span><br>Day <strong id="si_r1">—</strong></div>
                <div><span style="font-size:11px;color:#64748b">Reminder 2</span><br>Day <strong id="si_r2">—</strong></div>
                <div><span style="font-size:11px;color:#64748b">Owner</span><br><strong id="si_owner">—</strong></div>
                <div id="si_ms_div" style="display:none"><span style="font-size:11px;color:#64748b">Type</span><br>
                  <span class="pill pill-milestone" style="font-size:11px"><i class="bi bi-flag-fill"></i> Milestone</span></div>
                <div id="si_opt_div" style="display:none"><span style="font-size:11px;color:#64748b">Stage Type</span><br>
                  <span style="background:#e0f2fe;color:#0369a1;font-size:11px;padding:2px 8px;border-radius:4px"><i class="bi bi-skip-forward-fill"></i> Optional — can be skipped</span></div>
              </div>
            </div>
          </div>

          <hr style="border-color:#f1f5f9;margin:20px 0">
          <div class="form-section-title">Application Info</div>
          <div class="row g-3 mb-3">
            <div class="col-md-6">
              <label class="form-label">Application ID</label>
              <input type="text" class="form-control" name="application_id"
                     placeholder="e.g. NABH-2025-001" required>
            </div>
            <div class="col-md-6">
              <label class="form-label">Date of Stage Change</label>
              <input type="date" class="form-control" name="stage_start_date"
                     value="{date.today().isoformat()}" required>
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label">Organisation Name</label>
            <input type="text" class="form-control" name="organisation_name"
                   placeholder="Hospital / Organisation name" required>
          </div>

          <hr style="border-color:#f1f5f9;margin:20px 0">
          <div class="form-section-title">People</div>
          <div class="mb-3">
            <label class="form-label">Your Email (Program Officer)</label>
            <input type="email" class="form-control" name="program_officer_email"
                   placeholder="po@qci.org.in" required>
          </div>
          <div class="row g-3">
            <div class="col-md-6">
              <label class="form-label">Action Owner Name</label>
              <input type="text" class="form-control" name="action_owner_name" required>
            </div>
            <div class="col-md-6">
              <label class="form-label">Action Owner Email</label>
              <input type="email" class="form-control" name="action_owner_email" required>
            </div>
          </div>

          <button type="submit" class="btn btn-primary w-100 mt-4" style="padding:10px">
            <i class="bi bi-check2-circle"></i> Submit Stage Change
          </button>
        </form>
      </div>
    </div>
  </div>

  <!-- Right column: help -->
  <div class="col-lg-5 d-none d-lg-block">
    <div class="card" style="background:#f8fafc;border-style:dashed">
      <div class="card-body p-4">
        <div style="font-weight:600;font-size:14px;color:#1a3557;margin-bottom:12px">
          <i class="bi bi-info-circle"></i> How this works
        </div>
        <ul style="font-size:13px;color:#475569;line-height:1.9;padding-left:18px">
          <li>If the Application ID already exists, the row is <strong>updated</strong> and all sent-flags are reset.</li>
          <li>If it's new, a <strong>new case</strong> is created.</li>
          <li>TAT, reminders, and owner type are pulled from the selected programme configuration.</li>
          <li>Milestone stages receive <strong>no emails</strong>.</li>
          <li><i class="bi bi-skip-forward-fill" style="color:#0891b2"></i> <strong>Optional stages</strong> (marked ○) can be skipped — you may advance directly past them to the next required stage.</li>
          <li>The daily scheduler runs at the configured time (IST); change it in <strong>Settings → Scheduler Settings</strong>. Use Run Check on the dashboard to trigger immediately.</li>
        </ul>
        <div style="margin-top:16px;padding:12px 14px;background:#fff;border-radius:8px;border:1px solid #e2e8f0">
          <div style="font-size:11px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">Need bulk upload?</div>
          <a href="/bulk-upload" class="btn btn-sm btn-outline-primary w-100">
            <i class="bi bi-upload"></i> Go to Bulk Upload
          </a>
        </div>
      </div>
    </div>
  </div>

</div>
"""
    scripts = """<script>
var _stageData = {};
document.getElementById('progSelect').addEventListener('change', function(){
  var prog = this.value;
  var sel = document.getElementById('stageSelect');
  var box = document.getElementById('stageInfoBox');
  sel.innerHTML = '<option value="">Loading…</option>';
  box.classList.remove('show');
  _stageData = {};
  if(!prog){ sel.innerHTML='<option value="">Select programme first…</option>'; return; }
  fetch('/api/stages?programme='+encodeURIComponent(prog))
    .then(r=>r.json())
    .then(data=>{
      _stageData = {};
      data.forEach(function(s){ _stageData[s.stage_name] = s; });
      sel.innerHTML = '<option value="">Select a stage…</option>' +
        data.map(s=>'<option value="'+s.stage_name+'">'+(s.is_milestone ? '⬥ ' : '')+(s.is_optional ? '○ ' : '')+s.stage_name+(s.is_optional ? ' (optional)' : '')+'</option>').join('');
    });
});
document.getElementById('stageSelect').addEventListener('change', function(){
  var s = _stageData[this.value];
  var box = document.getElementById('stageInfoBox');
  if(!s){ box.classList.remove('show'); return; }
  document.getElementById('si_tat').textContent   = s.tat_days || '0';
  document.getElementById('si_r1').textContent    = s.reminder1_day || '—';
  document.getElementById('si_r2').textContent    = s.reminder2_day || '—';
  document.getElementById('si_owner').textContent = s.owner_type || '—';
  document.getElementById('si_ms_div').style.display  = s.is_milestone ? '' : 'none';
  document.getElementById('si_opt_div').style.display = s.is_optional  ? '' : 'none';
  box.classList.add('show');
});
</script>"""
    return render_page(content, scripts, active_page="log",
                       page_title="Log Stage Change",
                       page_crumb='<a href="/">Dashboard</a> / Log Stage Change')


@app.route("/api/stages")
@login_required
def api_stages():
    programme = request.args.get("programme", "")
    conn = get_db()
    bid = user_board_id()
    # Verify programme belongs to user's board (unless super_admin)
    if bid is not None:
        allowed = conn.execute(
            "SELECT id FROM programmes WHERE programme_name=? AND board_id=?", (programme, bid)
        ).fetchone()
        if not allowed:
            conn.close()
            return jsonify([])
    rows = conn.execute(
        "SELECT stage_name, stage_order, is_milestone, is_optional, tat_days, reminder1_day, reminder2_day, owner_type "
        "FROM programme_config WHERE programme_name=? ORDER BY stage_order",
        (programme,),
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/preview-tat-impact")
@board_admin_required
def preview_tat_impact():
    programme = request.args.get("programme", "")
    stage     = request.args.get("stage", "")
    try:
        new_tat = int(request.args.get("tat", "0"))
    except ValueError:
        return jsonify({"error": "Invalid TAT value"})
    if not programme or not stage:
        return jsonify({"error": "programme and stage required"})
    conn = get_db()
    bid = user_board_id()
    params = [programme, stage]
    q = "SELECT stage_start_date, tat_days FROM case_tracking WHERE programme_name=? AND current_stage=?"
    if bid is not None:
        q += " AND board_id=?"
        params.append(bid)
    cases = conn.execute(q, params).fetchall()
    conn.close()
    today = now_ist().date()
    total = len(cases)
    currently_overdue = sum(1 for c in cases if c["tat_days"] > 0 and working_days_elapsed(c["stage_start_date"], today) >= c["tat_days"])
    will_be_overdue   = sum(1 for c in cases if new_tat > 0 and working_days_elapsed(c["stage_start_date"], today) >= new_tat)
    return jsonify({
        "total_at_stage":     total,
        "currently_overdue":  currently_overdue,
        "will_be_overdue":    will_be_overdue,
        "newly_flagged":      max(0, will_be_overdue - currently_overdue),
        "newly_resolved":     max(0, currently_overdue - will_be_overdue),
    })


@app.route("/bulk-upload", methods=["GET", "POST"])
@login_required
def bulk_upload():
    REQUIRED_COLS = {"Application_ID", "Organisation_Name", "Programme_Name",
                     "Stage_Name", "Date_of_Stage_Change",
                     "Action_Owner_Name", "Action_Owner_Email", "Program_Officer_Email"}

    if request.method == "POST":
        f = request.files.get("upload_file")
        if not f or not f.filename:
            flash("No file uploaded.", "error")
            return redirect(url_for("bulk_upload"))

        fname = f.filename.lower()
        created = updated = failed = 0
        errors = []
        row_dicts = []

        try:
            if fname.endswith(".xlsx") or fname.endswith(".xls"):
                if not HAS_XLSX:
                    flash("openpyxl is not installed. Install it with: pip3 install openpyxl", "error")
                    return redirect(url_for("bulk_upload"))
                wb = openpyxl.load_workbook(io.BytesIO(f.stream.read()))
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    flash("The uploaded file is empty.", "error")
                    return redirect(url_for("bulk_upload"))
                headers = [str(h).strip() if h else "" for h in all_rows[0]]
                row_dicts = [
                    {headers[i]: (str(v).strip() if v is not None else "")
                     for i, v in enumerate(row) if i < len(headers)}
                    for row in all_rows[1:]
                ]
            else:
                raw = f.stream.read()
                stream = io.StringIO(raw.decode("utf-8-sig"))
                reader = csv.DictReader(stream)
                headers = list(reader.fieldnames or [])
                row_dicts = list(reader)

            if not REQUIRED_COLS.issubset(set(headers)):
                missing = REQUIRED_COLS - set(headers)
                flash(f"File missing columns: {', '.join(sorted(missing))}", "error")
                return redirect(url_for("bulk_upload"))

        except Exception as e:
            flash(f"Could not read file: {e}", "error")
            return redirect(url_for("bulk_upload"))

        for i, row in enumerate(row_dicts, start=2):
            try:
                data = {
                    "application_id":     row.get("Application_ID", "").strip(),
                    "organisation_name":  row.get("Organisation_Name", "").strip(),
                    "programme_name":     row.get("Programme_Name", "").strip(),
                    "stage_name":         row.get("Stage_Name", "").strip(),
                    "stage_start_date":   row.get("Date_of_Stage_Change", "").strip(),
                    "action_owner_name":  row.get("Action_Owner_Name", "").strip(),
                    "action_owner_email": row.get("Action_Owner_Email", "").strip(),
                    "program_officer_email": row.get("Program_Officer_Email", "").strip(),
                    "cc_emails":          row.get("CC_Emails", "").strip(),
                    "_changed_by":        session.get("full_name") or session.get("username", ""),
                }
                if not data["application_id"]:
                    raise ValueError("Application_ID is empty")
                # Validate date format (YYYY-MM-DD required)
                _raw_date = data["stage_start_date"]
                if not _raw_date:
                    data["stage_start_date"] = date.today().isoformat()
                else:
                    try:
                        datetime.strptime(_raw_date, "%Y-%m-%d")
                    except ValueError:
                        # Try common alternate formats: DD/MM/YYYY, MM/DD/YYYY
                        _parsed = None
                        for _fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                            try:
                                _parsed = datetime.strptime(_raw_date, _fmt).date().isoformat()
                                break
                            except ValueError:
                                pass
                        if _parsed:
                            data["stage_start_date"] = _parsed
                        else:
                            raise ValueError(f"Date_of_Stage_Change '{_raw_date}' must be YYYY-MM-DD")
                action = upsert_case(data)
                if action == "created":
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                failed += 1
                errors.append({"row": i, "app_id": row.get("Application_ID", ""),
                                "reason": str(e)})

        log_audit("bulk_upload", None,
                  f"Uploaded: {created} created, {updated} updated, {failed} failed",
                  session.get("full_name") or session.get("username", ""), user_board_id())

        if errors and request.form.get("download_errors"):
            # Download error report as CSV
            out = io.StringIO()
            w = csv.DictWriter(out, fieldnames=["Row", "Application_ID", "Reason"])
            w.writeheader()
            for e in errors:
                w.writerow({"Row": e["row"], "Application_ID": e["app_id"], "Reason": e["reason"]})
            out.seek(0)
            return Response(out.getvalue(), mimetype="text/csv",
                            headers={"Content-Disposition": "attachment;filename=upload_errors.csv"})

        msg = f"Upload complete — {created} created, {updated} updated, {failed} failed"
        if errors:
            msg += f". <strong>{failed} row(s) failed</strong> — resubmit with 'Download Error Report' checked."
        flash(msg, "success" if not failed else "warning")
        return redirect(url_for("bulk_upload"))

    content = """
<div class="row g-4 justify-content-center" style="max-width:860px;margin:0 auto">
  <div class="col-lg-7">
    <div class="card form-card">
      <div class="card-header"><i class="bi bi-upload" style="color:#059669"></i> Upload CSV File</div>
      <div class="card-body p-4">

        <!-- Step 1 -->
        <div style="display:flex;align-items:flex-start;gap:14px;margin-bottom:24px">
          <div style="min-width:32px;height:32px;border-radius:50%;background:#e8eef7;color:#1a3557;font-weight:700;font-size:14px;display:flex;align-items:center;justify-content:center">1</div>
          <div style="flex:1">
            <div style="font-weight:600;font-size:14px;margin-bottom:4px">Download the template</div>
            <div style="font-size:13px;color:#64748b;margin-bottom:10px">Fill in your data, then upload below.</div>
            <div class="d-flex gap-2">
              <a href="/csv-template" class="btn btn-sm btn-outline-primary">
                <i class="bi bi-filetype-csv"></i> CSV Template
              </a>
              <a href="/xlsx-template" class="btn btn-sm btn-outline-success">
                <i class="bi bi-file-earmark-excel"></i> Excel Template
              </a>
            </div>
          </div>
        </div>

        <hr style="border-color:#f1f5f9">

        <!-- Step 2 -->
        <div style="display:flex;align-items:flex-start;gap:14px;margin-top:24px">
          <div style="min-width:32px;height:32px;border-radius:50%;background:#e8eef7;color:#1a3557;font-weight:700;font-size:14px;display:flex;align-items:center;justify-content:center">2</div>
          <div style="flex:1">
            <div style="font-weight:600;font-size:14px;margin-bottom:4px">Upload your filled CSV</div>
            <form method="post" enctype="multipart/form-data" id="uploadForm">
              <label class="upload-zone" for="csvInput" id="uploadZone">
                <i class="bi bi-cloud-upload" style="font-size:32px;color:#94a3b8;display:block;margin-bottom:8px"></i>
                <div style="font-weight:500;color:#475569">Drop file here or click to browse</div>
                <div style="font-size:12px;color:#94a3b8;margin-top:4px">Accepts <strong>.csv</strong> and <strong>.xlsx</strong></div>
                <div style="font-size:12px;color:#64748b;margin-top:6px;font-weight:500" id="fileLabel">No file selected</div>
                <input type="file" id="csvInput" name="upload_file" accept=".csv,.xlsx,.xls" required style="display:none">
              </label>
              <div class="form-check mt-3 mb-1" style="font-size:12px">
                <input class="form-check-input" type="checkbox" name="download_errors" id="dlErrCheck" value="1">
                <label class="form-check-label" for="dlErrCheck">Download error report if any rows fail</label>
              </div>
              <button type="submit" class="btn btn-primary w-100 mt-2" style="padding:10px">
                <i class="bi bi-upload"></i> Upload &amp; Process
              </button>
            </form>
          </div>
        </div>

      </div>
    </div>
  </div>

  <div class="col-lg-5">
    <div class="card">
      <div class="card-header"><i class="bi bi-table" style="color:#7c3aed"></i> Required Columns</div>
      <div class="card-body p-0">
        <table class="data-table">
          <thead><tr><th>Column</th><th>Example</th></tr></thead>
          <tbody>
            <tr><td><code>Application_ID</code></td><td style="color:#64748b;font-size:12px">NABH-2025-001</td></tr>
            <tr><td><code>Organisation_Name</code></td><td style="color:#64748b;font-size:12px">City Hospital</td></tr>
            <tr><td><code>Programme_Name</code></td><td style="color:#64748b;font-size:12px">NABH Full Accreditation…</td></tr>
            <tr><td><code>Stage_Name</code></td><td style="color:#64748b;font-size:12px">Application In Progress</td></tr>
            <tr><td><code>Date_of_Stage_Change</code></td><td style="color:#64748b;font-size:12px">2026-03-20</td></tr>
            <tr><td><code>Action_Owner_Name</code></td><td style="color:#64748b;font-size:12px">Mr. Applicant</td></tr>
            <tr><td><code>Action_Owner_Email</code></td><td style="color:#64748b;font-size:12px">applicant@hospital.in</td></tr>
            <tr><td><code>Program_Officer_Email</code></td><td style="color:#64748b;font-size:12px">po@qci.org.in</td></tr>
          </tbody>
        </table>
      </div>
    </div>
    <div class="card mt-3" style="background:#fffbeb;border-color:#fde68a">
      <div class="card-body p-3" style="font-size:12.5px;color:#92400e">
        <i class="bi bi-lightbulb-fill" style="color:#d97706"></i>
        <strong>Tip:</strong> If an Application ID already exists, its record will be <em>updated</em> and all sent-flags reset. New IDs create fresh cases.
      </div>
    </div>
  </div>
</div>
"""
    scripts = """<script>
var inp = document.getElementById('csvInput');
var lbl = document.getElementById('fileLabel');
var zone = document.getElementById('uploadZone');
inp.addEventListener('change', function(){
  lbl.textContent = this.files[0] ? this.files[0].name : 'No file selected';
});
zone.addEventListener('dragover', function(e){ e.preventDefault(); zone.classList.add('drag'); });
zone.addEventListener('dragleave', function(){ zone.classList.remove('drag'); });
zone.addEventListener('drop', function(e){
  e.preventDefault(); zone.classList.remove('drag');
  var f = e.dataTransfer.files[0];
  if(f){ inp.files = e.dataTransfer.files; lbl.textContent = f.name; }
});
document.getElementById('uploadForm').addEventListener('submit', function(){
  var btn = this.querySelector('button[type=submit]');
  btn.innerHTML = '<span class="spinner-border spinner-border-sm"></span> Processing…';
  btn.disabled = true;
});
</script>"""
    return render_page(content, scripts, active_page="bulk",
                       page_title="Bulk Upload",
                       page_crumb='<a href="/">Dashboard</a> / Bulk Upload')


_TEMPLATE_COLS = ["Application_ID", "Organisation_Name", "Programme_Name", "Stage_Name",
                  "Date_of_Stage_Change", "Action_Owner_Name", "Action_Owner_Email", "Program_Officer_Email"]
_TEMPLATE_EXAMPLE = ["APP-001", "Sample Hospital", "NABH Full Accreditation Hospitals",
                     "Application In Progress", date.today().isoformat(),
                     "Mr. Applicant", "applicant@example.com", "po@qci.org.in"]


@app.route("/csv-template")
@login_required
def csv_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_TEMPLATE_COLS)
    writer.writerow(_TEMPLATE_EXAMPLE)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=case_upload_template.csv"},
    )


@app.route("/xlsx-template")
@login_required
def xlsx_template():
    if not HAS_XLSX:
        flash("openpyxl not installed. Run: pip3 install openpyxl", "error")
        return redirect(url_for("bulk_upload"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"
    ws.append(_TEMPLATE_COLS)
    ws.append(_TEMPLATE_EXAMPLE)
    # Style header row
    header_fill = PatternFill("solid", fgColor="1A3557")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    # Style example row
    for cell in ws[2]:
        cell.font = Font(italic=True, color="888888")
    # Column widths
    widths = [16, 24, 36, 30, 22, 22, 28, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=case_upload_template.xlsx"},
    )


@app.route("/edit-case/<int:case_id>", methods=["GET", "POST"])
@login_required
def edit_case(case_id):
    conn = get_db()
    case = conn.execute("SELECT * FROM case_tracking WHERE id=?", (case_id,)).fetchone()
    if not case:
        conn.close()
        flash("Case not found.", "error")
        return redirect(url_for("dashboard"))
    case = dict(case)

    bid = user_board_id()
    if bid is not None and case.get("board_id") != bid:
        conn.close()
        flash("Access denied.", "error")
        return redirect(url_for("dashboard"))
    if bid is not None:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
    else:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes ORDER BY programme_name"
        ).fetchall()]
    conn.close()

    if request.method == "POST":
        new_status = request.form.get("status", "Active")
        data = {
            "application_id":     case["application_id"],
            "organisation_name":  request.form["organisation_name"].strip(),
            "programme_name":     request.form["programme_name"].strip(),
            "stage_name":         request.form["stage_name"].strip(),
            "stage_start_date":   request.form["stage_start_date"].strip(),
            "action_owner_name":  request.form["action_owner_name"].strip(),
            "action_owner_email": request.form["action_owner_email"].strip(),
            "program_officer_email": request.form["program_officer_email"].strip(),
            "cc_emails":          request.form.get("cc_emails", "").strip(),
            "suppress_until":     request.form.get("suppress_until", "").strip() or None,
        }
        try:
            data["_changed_by"] = session.get("full_name") or session.get("username", "")
            upsert_case(data)
            # Update status separately (upsert_case doesn't manage status field)
            _sc = get_db()
            _sc.execute("UPDATE case_tracking SET status=? WHERE application_id=?",
                        (new_status, case["application_id"]))
            _sc.commit()
            _sc.close()
            flash(f"Case {case['application_id']} updated.", "success")
            return redirect(url_for("dashboard"))
        except ValueError as e:
            flash(str(e), "error")

    prog_opts = "".join(
        f'<option value="{p}" {"selected" if p==case["programme_name"] else ""}>{p}</option>'
        for p in programmes
    )
    content = f"""
<div style="max-width:720px;margin:0 auto">
  <div class="card form-card">
    <div class="card-header d-flex align-items-center gap-3">
      <i class="bi bi-pencil-square" style="color:#d97706;font-size:18px"></i>
      <div>
        <div>Edit Case</div>
        <div style="font-size:11px;font-weight:400;color:#94a3b8">{case['application_id']} · {case['organisation_name']}</div>
      </div>
    </div>
    <div class="card-body p-4">
      <form method="post" id="editForm">

        <div class="form-section-title">Programme &amp; Stage</div>
        <div class="row g-3 mb-3">
          <div class="col-md-6">
            <label class="form-label">Programme</label>
            <select class="form-select" name="programme_name" id="progSelect" required>
              {prog_opts}
            </select>
          </div>
          <div class="col-md-6">
            <label class="form-label">Stage</label>
            <select class="form-select" name="stage_name" id="stageSelect" required>
              <option value="{case['current_stage']}">{case['current_stage']}</option>
            </select>
          </div>
        </div>

        <hr style="border-color:#f1f5f9;margin:20px 0">
        <div class="form-section-title">Application Info</div>
        <div class="row g-3 mb-3">
          <div class="col-md-6">
            <label class="form-label">Organisation Name</label>
            <input type="text" class="form-control" name="organisation_name"
                   value="{case['organisation_name']}" required>
          </div>
          <div class="col-md-6">
            <label class="form-label">Date of Stage Change</label>
            <input type="date" class="form-control" name="stage_start_date"
                   value="{case['stage_start_date']}" required>
          </div>
        </div>

        <hr style="border-color:#f1f5f9;margin:20px 0">
        <div class="form-section-title">People</div>
        <div class="mb-3">
          <label class="form-label">Program Officer Email</label>
          <input type="email" class="form-control" name="program_officer_email"
                 value="{case['program_officer_email'] or ''}" required>
        </div>
        <div class="row g-3 mb-4">
          <div class="col-md-6">
            <label class="form-label">Action Owner Name</label>
            <input type="text" class="form-control" name="action_owner_name"
                   value="{case['action_owner_name'] or ''}" required>
          </div>
          <div class="col-md-6">
            <label class="form-label">Action Owner Email</label>
            <input type="email" class="form-control" name="action_owner_email"
                   value="{case['action_owner_email'] or ''}" required>
          </div>
        </div>

        <hr style="border-color:#f1f5f9;margin:20px 0">
        <div class="form-section-title">Notifications</div>
        <div class="mb-3">
          <label class="form-label">CC Emails <span style="color:#94a3b8;font-size:11px">(comma-separated — get copied on all notifications)</span></label>
          <input type="text" class="form-control" name="cc_emails"
                 value="{case.get('cc_emails') or ''}" placeholder="manager@org.com, supervisor@org.com">
        </div>
        <div class="mb-3">
          <label class="form-label">Suppress Notifications Until <span style="color:#94a3b8;font-size:11px">(optional — pause all emails until this date)</span></label>
          <input type="date" class="form-control" name="suppress_until"
                 value="{case.get('suppress_until') or ''}">
        </div>
        <div class="mb-3">
          <label class="form-label">Status</label>
          <select class="form-select" name="status">
            <option value="Active" {"selected" if (case.get("status") or "Active") == "Active" else ""}>Active</option>
            <option value="On Hold" {"selected" if case.get("status") == "On Hold" else ""}>On Hold</option>
            <option value="Closed" {"selected" if case.get("status") == "Closed" else ""}>Closed</option>
            <option value="Withdrawn" {"selected" if case.get("status") == "Withdrawn" else ""}>Withdrawn</option>
          </select>
        </div>

        <div class="alert" style="background:#fffbeb;border:1px solid #fde68a;border-radius:8px;font-size:13px;color:#92400e;padding:10px 14px">
          <i class="bi bi-arrow-clockwise"></i> Saving will <strong>reset</strong> all sent-flags (R1, R2, Overdue) and the overdue counter for this case.
        </div>

        <div class="d-flex gap-2 mt-3">
          <button type="submit" class="btn btn-warning flex-grow-1" style="padding:10px">
            <i class="bi bi-check2"></i> Save Changes
          </button>
          <a href="/" class="btn btn-outline-secondary" style="padding:10px 20px">Cancel</a>
        </div>
      </form>
    </div>
  </div>
</div>
"""
    scripts = f"""<script>
var _currentStage = "{case['current_stage']}";
function loadStages(prog, selectVal){{
  var sel = document.getElementById('stageSelect');
  fetch('/api/stages?programme='+encodeURIComponent(prog))
    .then(r=>r.json())
    .then(data=>{{
      sel.innerHTML = data.map(s=>'<option value="'+s.stage_name+'"'+(s.stage_name===selectVal?' selected':'')+'>'+s.stage_name+'</option>').join('');
    }});
}}
document.getElementById('progSelect').addEventListener('change', function(){{
  loadStages(this.value, '');
}});
loadStages("{case['programme_name']}", _currentStage);
</script>"""
    return render_page(content, scripts, active_page="",
                       page_title="Edit Case",
                       page_crumb=f'<a href="/">Dashboard</a> / Edit {case["application_id"]}')


@app.route("/delete-case/<int:case_id>")
@login_required
def delete_case(case_id):
    if session.get("role") not in ("super_admin", "board_admin"):
        flash("Permission denied.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db()
    row = conn.execute("SELECT application_id, board_id FROM case_tracking WHERE id=?", (case_id,)).fetchone()
    if not row:
        conn.close()
        flash("Case not found.", "error")
        return redirect(url_for("dashboard"))
    bid = user_board_id()
    if bid is not None and row["board_id"] != bid:
        conn.close()
        flash("Access denied.", "error")
        return redirect(url_for("dashboard"))
    conn.execute("DELETE FROM case_tracking WHERE id=?", (case_id,))
    conn.commit()
    flash(f"Case {row['application_id']} closed.", "success")
    conn.close()
    log_audit("case_closed", row["application_id"], "Case closed/removed",
              session.get("full_name") or session.get("username", ""),
              bid)
    return redirect(url_for("dashboard"))


@app.route("/update-case-status", methods=["POST"])
@login_required
def update_case_status():
    conn = get_db()
    app_id = request.form.get("application_id")
    new_status = request.form.get("status")
    valid = {"Active", "On Hold", "Closed", "Withdrawn", "Suspended"}
    if new_status not in valid:
        conn.close()
        flash("Invalid status.", "error")
        return redirect(url_for("dashboard"))

    case = conn.execute("SELECT * FROM case_tracking WHERE application_id=?", (app_id,)).fetchone()
    if not case:
        conn.close()
        flash("Case not found.", "error")
        return redirect(url_for("dashboard"))
    case = dict(case)

    bid = user_board_id()
    if bid is not None and case.get("board_id") != bid:
        conn.close()
        flash("Access denied.", "error")
        return redirect(url_for("dashboard"))

    if new_status == "On Hold":
        # Start hold — record hold_start_date
        conn.execute(
            "UPDATE case_tracking SET status=?, hold_start_date=? WHERE application_id=?",
            ("On Hold", date.today().isoformat(), app_id)
        )
    elif case.get("status") == "On Hold" and new_status != "On Hold":
        # Ending hold — calculate working days held and accumulate
        hold_start = case.get("hold_start_date")
        extra_hold = 0
        if hold_start:
            extra_hold = working_days_elapsed(hold_start, now_ist().date())
        new_hold_days = (case.get("hold_days") or 0) + extra_hold
        conn.execute(
            "UPDATE case_tracking SET status=?, hold_start_date=NULL, hold_days=? WHERE application_id=?",
            (new_status, new_hold_days, app_id)
        )
    else:
        conn.execute(
            "UPDATE case_tracking SET status=? WHERE application_id=?",
            (new_status, app_id)
        )

    conn.commit()
    log_audit("status_change", app_id,
              f"Status → {new_status}", session.get("full_name") or session.get("username", ""),
              case.get("board_id"))
    conn.close()
    flash(f"Case {app_id} status updated to {new_status}.", "success")
    return redirect(url_for("dashboard"))


@app.route("/save-filter", methods=["POST"])
@login_required
def save_filter():
    data = request.get_json()
    if not data or not data.get("name"):
        return jsonify({"ok": False, "error": "Name required"})
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO saved_filters (user_id, filter_name, filter_json, created_at) VALUES (?,?,?,?) ON CONFLICT (user_id, filter_name) DO UPDATE SET filter_json=EXCLUDED.filter_json, created_at=EXCLUDED.created_at",
            (session["user_id"], data["name"][:60],
             json.dumps({"qs": data.get("qs", "")}),
             now_ist().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "error": str(e)})


@app.route("/delete-filter/<int:filter_id>", methods=["POST"])
@login_required
def delete_filter(filter_id):
    conn = get_db()
    conn.execute("DELETE FROM saved_filters WHERE id=? AND user_id=?", (filter_id, session["user_id"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/settings", methods=["GET", "POST"])
@board_admin_required
def settings():
    conn = get_db()
    msg = ""

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_board":
            if session.get("role") != "super_admin":
                flash("Only Super Admin can add boards.", "error")
            else:
                bname = request.form.get("board_name", "").strip()
                if bname:
                    try:
                        conn.execute("INSERT INTO boards (board_name) VALUES (?)", (bname,))
                        conn.commit()
                        flash(f"Board '{bname}' added.", "success")
                    except Exception as e:
                        flash(f"Error: {e}", "error")

        elif action == "add_programme":
            pname = request.form.get("programme_name", "").strip()
            board_id = request.form.get("board_id", "")
            if not pname or not board_id:
                flash("Programme name and board are required.", "error")
            else:
                board_id = int(board_id)
                if session.get("role") == "board_admin" and board_id != session.get("board_id"):
                    flash("Cannot add programme to another board.", "error")
                else:
                    try:
                        conn.execute(
                            "INSERT INTO programmes (programme_name, board_id) VALUES (?,?)",
                            (pname, board_id)
                        )
                        conn.commit()
                        flash(f"Programme '{pname}' added.", "success")
                    except Exception as e:
                        flash(f"Error: {e}", "error")

        elif action == "edit_programme":
            ep_name = request.form.get("programme_name", "").strip()
            if not ep_name:
                flash("Programme name missing.", "error")
            else:
                ep_row = conn.execute(
                    "SELECT board_id FROM programmes WHERE programme_name=?", (ep_name,)
                ).fetchone()
                if not ep_row:
                    flash("Programme not found.", "error")
                elif session.get("role") == "board_admin" and ep_row["board_id"] != session.get("board_id"):
                    flash("Cannot edit a programme from another board.", "error")
                else:
                    try:
                        conn.execute(
                            """UPDATE programmes SET
                               tat_days=?, reminder1_days=?, reminder2_days=?,
                               overdue_days=?, notification_emails=?
                               WHERE programme_name=?""",
                            (
                                int(request.form.get("tat_days", 0) or 0),
                                int(request.form.get("reminder1_days", 0) or 0),
                                int(request.form.get("reminder2_days", 0) or 0),
                                int(request.form.get("overdue_days", 0) or 0),
                                request.form.get("notification_emails", "").strip() or None,
                                ep_name,
                            )
                        )
                        conn.commit()
                        flash(f"Programme '{ep_name}' updated.", "success")
                    except Exception as e:
                        flash(f"Error updating programme: {e}", "error")

        elif action == "add_stage":
            pname = request.form.get("programme_name", "").strip()
            sname = request.form.get("stage_name", "").strip()
            if not pname or not sname:
                flash("Programme and stage name are required.", "error")
            else:
                # Determine board_id from programme
                prog_row = conn.execute(
                    "SELECT board_id FROM programmes WHERE programme_name=?", (pname,)
                ).fetchone()
                prog_board_id = prog_row[0] if prog_row else None
                if session.get("role") == "board_admin" and prog_board_id != session.get("board_id"):
                    flash("Cannot add stages to a programme in another board.", "error")
                else:
                    _new_order = int(request.form.get("stage_order", 1))
                    _dup = conn.execute(
                        "SELECT id FROM programme_config WHERE programme_name=? AND stage_order=?",
                        (pname, _new_order)
                    ).fetchone()
                    if _dup:
                        flash(f"Stage order #{_new_order} already exists in '{pname}'. Choose a different order number.", "error")
                    else:
                        try:
                            conn.execute(
                                """INSERT INTO programme_config
                                   (programme_name,stage_name,stage_order,tat_days,reminder1_day,
                                    reminder2_day,owner_type,overdue_interval_days,is_milestone,is_optional,board_id)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                (pname, sname,
                                 _new_order,
                                 int(request.form.get("tat_days", 0)),
                                 int(request.form.get("reminder1_day", 0)),
                                 int(request.form.get("reminder2_day", 0)),
                                 request.form.get("owner_type") or None,
                                 int(request.form.get("overdue_interval_days", 3)),
                                 1 if request.form.get("is_milestone") else 0,
                                 1 if request.form.get("is_optional") else 0,
                                 prog_board_id),
                            )
                            conn.commit()
                            flash(f"Stage '{sname}' added to {pname}.", "success")
                        except Exception as e:
                            flash(f"Error: {e}", "error")

        elif action == "delete_board":
            if session.get("role") != "super_admin":
                flash("Only Super Admin can delete boards.", "error")
            else:
                bid_del = request.form.get("board_id")
                if bid_del:
                    prog_count = conn.execute(
                        "SELECT COUNT(*) FROM programmes WHERE board_id=?", (bid_del,)
                    ).fetchone()[0]
                    if prog_count > 0:
                        flash(f"Cannot delete board — {prog_count} programme(s) still exist. Delete them first.", "error")
                    else:
                        user_count = conn.execute(
                            "SELECT COUNT(*) FROM users WHERE board_id=?", (bid_del,)
                        ).fetchone()[0]
                        if user_count > 0:
                            flash(f"Cannot delete board — {user_count} user(s) are assigned to it. Reassign them first.", "error")
                        else:
                            conn.execute("DELETE FROM boards WHERE id=?", (bid_del,))
                            conn.commit()
                            flash("Board deleted.", "success")

        elif action == "delete_programme":
            pname_del = request.form.get("programme_name", "").strip()
            if not pname_del:
                flash("Programme name missing.", "error")
            else:
                prog_row = conn.execute(
                    "SELECT board_id FROM programmes WHERE programme_name=?", (pname_del,)
                ).fetchone()
                if not prog_row:
                    flash("Programme not found.", "error")
                elif session.get("role") == "board_admin" and prog_row["board_id"] != session.get("board_id"):
                    flash("Cannot delete a programme from another board.", "error")
                else:
                    case_count = conn.execute(
                        "SELECT COUNT(*) FROM case_tracking WHERE programme_name=?", (pname_del,)
                    ).fetchone()[0]
                    if case_count > 0:
                        flash(f"Cannot delete '{pname_del}' — {case_count} active case(s) exist. Close them first.", "error")
                    else:
                        conn.execute("DELETE FROM programme_config WHERE programme_name=?", (pname_del,))
                        conn.execute("DELETE FROM email_templates WHERE programme_name=?", (pname_del,))
                        conn.execute(
                            "DELETE FROM user_programme_map WHERE programme_id IN "
                            "(SELECT id FROM programmes WHERE programme_name=?)", (pname_del,)
                        )
                        conn.execute("DELETE FROM programmes WHERE programme_name=?", (pname_del,))
                        conn.commit()
                        flash(f"Programme '{pname_del}' and all its stages deleted.", "success")

        elif action == "delete_stage":
            pname_del = request.form.get("programme_name", "").strip()
            sname_del = request.form.get("stage_name", "").strip()
            if not pname_del or not sname_del:
                flash("Programme or stage name missing.", "error")
            else:
                prog_row = conn.execute(
                    "SELECT board_id FROM programmes WHERE programme_name=?", (pname_del,)
                ).fetchone()
                if session.get("role") == "board_admin" and prog_row and prog_row["board_id"] != session.get("board_id"):
                    flash("Cannot delete stages from another board's programme.", "error")
                else:
                    case_count = conn.execute(
                        "SELECT COUNT(*) FROM case_tracking WHERE programme_name=? AND current_stage=?",
                        (pname_del, sname_del)
                    ).fetchone()[0]
                    if case_count > 0:
                        flash(f"Cannot delete stage '{sname_del}' — {case_count} case(s) are currently at this stage.", "error")
                    else:
                        conn.execute(
                            "DELETE FROM programme_config WHERE programme_name=? AND stage_name=?",
                            (pname_del, sname_del)
                        )
                        conn.commit()
                        flash(f"Stage '{sname_del}' deleted.", "success")

        elif action == "update_stage":
            sid = request.form.get("stage_id", "").strip()
            if not sid:
                flash("Stage ID missing.", "error")
            else:
                stage_row = conn.execute(
                    "SELECT programme_name, board_id FROM programme_config WHERE id=?", (sid,)
                ).fetchone()
                if not stage_row:
                    flash("Stage not found.", "error")
                elif session.get("role") == "board_admin" and stage_row["board_id"] != session.get("board_id"):
                    flash("Cannot edit stages from another board.", "error")
                else:
                    _us_name = request.form.get("stage_name", "").strip()
                    _us_order = int(request.form.get("stage_order", 1))
                    # Check duplicate order (excluding self)
                    _dup2 = conn.execute(
                        "SELECT id FROM programme_config WHERE programme_name=? AND stage_order=? AND id!=?",
                        (stage_row["programme_name"], _us_order, sid)
                    ).fetchone()
                    if not _us_name:
                        flash("Stage name is required.", "error")
                    elif _dup2:
                        flash(f"Stage order #{_us_order} is already used by another stage in this programme.", "error")
                    else:
                        try:
                            conn.execute(
                                """UPDATE programme_config SET
                                   stage_name=?, stage_order=?, tat_days=?, reminder1_day=?,
                                   reminder2_day=?, owner_type=?, overdue_interval_days=?,
                                   is_milestone=?, is_optional=?
                                   WHERE id=?""",
                                (_us_name,
                                 _us_order,
                                 int(request.form.get("tat_days", 0)),
                                 int(request.form.get("reminder1_day", 0)),
                                 int(request.form.get("reminder2_day", 0)),
                                 request.form.get("owner_type") or None,
                                 int(request.form.get("overdue_interval_days", 3)),
                                 1 if request.form.get("is_milestone") else 0,
                                 1 if request.form.get("is_optional") else 0,
                                 sid)
                            )
                            conn.commit()
                            flash(f"Stage '{_us_name}' updated.", "success")
                        except Exception as e:
                            flash(f"Error updating stage: {e}", "error")

        elif action == "copy_stages":
            src_prog = request.form.get("source_programme", "").strip()
            dst_prog = request.form.get("dest_programme", "").strip()
            overwrite = request.form.get("overwrite_existing") == "1"
            if not src_prog or not dst_prog:
                flash("Source and destination programmes are required.", "error")
            elif src_prog == dst_prog:
                flash("Source and destination cannot be the same programme.", "error")
            else:
                dst_row = conn.execute(
                    "SELECT id, board_id FROM programmes WHERE programme_name=?", (dst_prog,)
                ).fetchone()
                if not dst_row:
                    flash("Destination programme not found.", "error")
                elif session.get("role") == "board_admin" and dst_row["board_id"] != session.get("board_id"):
                    flash("Cannot copy stages to a programme in another board.", "error")
                else:
                    src_stages = conn.execute(
                        "SELECT * FROM programme_config WHERE programme_name=? ORDER BY stage_order",
                        (src_prog,)
                    ).fetchall()
                    if not src_stages:
                        flash(f"Source programme '{src_prog}' has no stages to copy.", "error")
                    else:
                        if overwrite:
                            conn.execute(
                                "DELETE FROM programme_config WHERE programme_name=?", (dst_prog,)
                            )
                            conn.commit()
                        copied = 0
                        skipped = 0
                        for _ss in src_stages:
                            _exist = conn.execute(
                                "SELECT id FROM programme_config WHERE programme_name=? AND stage_order=?",
                                (dst_prog, _ss["stage_order"])
                            ).fetchone()
                            if _exist and not overwrite:
                                skipped += 1
                                continue
                            try:
                                conn.execute(
                                    """INSERT INTO programme_config
                                       (programme_name, stage_name, stage_order, tat_days,
                                        reminder1_day, reminder2_day, owner_type,
                                        overdue_interval_days, is_milestone, is_optional, board_id)
                                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                    (dst_prog, _ss["stage_name"], _ss["stage_order"],
                                     _ss["tat_days"], _ss["reminder1_day"], _ss["reminder2_day"],
                                     _ss["owner_type"], _ss["overdue_interval_days"],
                                     _ss["is_milestone"], _ss["is_optional"],
                                     dst_row["board_id"])
                                )
                                copied += 1
                            except Exception:
                                conn.rollback()  # PostgreSQL requires rollback after failed statement
                                skipped += 1
                        conn.commit()
                        msg = f"Copied {copied} stage(s) from '{src_prog}' to '{dst_prog}'."
                        if skipped:
                            msg += f" {skipped} skipped (order conflict)."
                        flash(msg, "success")

        elif action == "copy_stages_from":
            src_prog = request.form.get("source_programme_name", "").strip()
            dst_prog = request.form.get("current_programme_name", "").strip()
            confirm_replace = request.form.get("confirm_replace") == "1"
            if not src_prog or not dst_prog:
                flash("Source and current programme names are required.", "error")
            elif src_prog == dst_prog:
                flash("Source and destination cannot be the same programme.", "error")
            else:
                src_row = conn.execute(
                    "SELECT id, board_id FROM programmes WHERE programme_name=?", (src_prog,)
                ).fetchone()
                dst_row2 = conn.execute(
                    "SELECT id, board_id FROM programmes WHERE programme_name=?", (dst_prog,)
                ).fetchone()
                if not src_row or not dst_row2:
                    flash("Programme not found.", "error")
                elif src_row["board_id"] != dst_row2["board_id"]:
                    flash("Both programmes must belong to the same board.", "error")
                elif session.get("role") == "board_admin" and dst_row2["board_id"] != session.get("board_id"):
                    flash("Cannot copy stages to a programme in another board.", "error")
                else:
                    src_stages = conn.execute(
                        "SELECT * FROM programme_config WHERE programme_name=? ORDER BY stage_order",
                        (src_prog,)
                    ).fetchall()
                    if not src_stages:
                        flash(f"Source programme '{src_prog}' has no stages to copy.", "error")
                    else:
                        if confirm_replace:
                            conn.execute(
                                "DELETE FROM programme_config WHERE programme_name=?", (dst_prog,)
                            )
                            conn.commit()
                        copied2 = 0
                        skipped2 = 0
                        for _ss2 in src_stages:
                            _exist2 = conn.execute(
                                "SELECT id FROM programme_config WHERE programme_name=? AND stage_order=?",
                                (dst_prog, _ss2["stage_order"])
                            ).fetchone()
                            if _exist2 and not confirm_replace:
                                skipped2 += 1
                                continue
                            try:
                                conn.execute(
                                    """INSERT INTO programme_config
                                       (programme_name, stage_name, stage_order, tat_days,
                                        reminder1_day, reminder2_day, owner_type,
                                        overdue_interval_days, is_milestone, is_optional, board_id)
                                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                    (dst_prog, _ss2["stage_name"], _ss2["stage_order"],
                                     _ss2["tat_days"], _ss2["reminder1_day"], _ss2["reminder2_day"],
                                     _ss2["owner_type"], _ss2["overdue_interval_days"],
                                     _ss2["is_milestone"], _ss2["is_optional"],
                                     dst_row2["board_id"])
                                )
                                copied2 += 1
                            except Exception:
                                conn.rollback()  # PostgreSQL requires rollback after failed statement
                                skipped2 += 1
                        conn.commit()
                        msg2 = f"Copied {copied2} stage(s) from '{src_prog}' to '{dst_prog}'."
                        if skipped2:
                            msg2 += f" {skipped2} skipped (order conflict)."
                        flash(msg2, "success")

        elif action == "update_schedule":
            if session.get("role") != "super_admin":
                flash("Only Super Admin can change the schedule.", "error")
            else:
                raw = request.form.get("schedule_time", "08:00")
                try:
                    parts = raw.split(":")
                    sh, sm = int(parts[0]), int(parts[1])
                    set_app_setting("scheduler_hour",   str(sh))
                    set_app_setting("scheduler_minute", str(sm))
                    try:
                        scheduler.reschedule_job(
                            "daily_check", trigger="cron",
                            hour=sh, minute=sm
                        )
                    except Exception:
                        pass  # scheduler not running (Vercel serverless — Cron Job used instead)
                    flash(f"Daily check rescheduled to {sh:02d}:{sm:02d} IST.", "success")
                except Exception as e:
                    flash(f"Error updating schedule: {e}", "error")

        elif action == "update_sender":
            prog = request.form["programme_name"].strip()
            email = request.form["sender_email"].strip()
            password = request.form.get("sender_password", "").strip()
            smtp_host = request.form.get("smtp_host", "smtp.gmail.com").strip()
            smtp_port = int(request.form.get("smtp_port", 587) or 587)
            enc_pw = encrypt_str(password) if password else None
            if enc_pw:
                conn.execute(
                    "UPDATE programme_config SET sender_email=?, sender_password=?, smtp_host=?, smtp_port=? WHERE programme_name=?",
                    (email, enc_pw, smtp_host, smtp_port, prog),
                )
            else:
                conn.execute(
                    "UPDATE programme_config SET sender_email=?, smtp_host=?, smtp_port=? WHERE programme_name=?",
                    (email, smtp_host, smtp_port, prog),
                )
            conn.commit()
            flash(f"Sender credentials updated for {prog}.", "success")

        elif action == "add_board_holiday":
            _hbid = user_board_id()
            hdate = request.form.get("holiday_date", "").strip()
            hname = request.form.get("holiday_name", "").strip()
            if hdate and hname:
                try:
                    _existing = conn.execute(
                        "SELECT id FROM holidays WHERE holiday_date=? AND (board_id=? OR (board_id IS NULL AND ? IS NULL))",
                        (hdate, _hbid, _hbid)
                    ).fetchone()
                    if not _existing:
                        conn.execute(
                            "INSERT INTO holidays (holiday_date, name, board_id) VALUES (?,?,?)",
                            (hdate, hname, _hbid)
                        )
                    conn.commit()
                    flash(f"Holiday '{hname}' added.", "success")
                except Exception as e:
                    flash(f"Error: {e}", "error")

        elif action == "delete_board_holiday":
            _hbid = user_board_id()
            hid = request.form.get("holiday_id")
            if hid:
                conn.execute(
                    "DELETE FROM holidays WHERE id=? AND (board_id=? OR board_id IS NULL)",
                    (hid, _hbid)
                )
                conn.commit()
                flash("Holiday removed.", "success")

    # Build Board → Programme → Stage hierarchy
    bid = user_board_id()
    all_boards = [dict(r) for r in conn.execute(
        "SELECT * FROM boards ORDER BY board_name"
    ).fetchall()]

    # Filter boards visible to this user
    if bid is not None:
        visible_boards = [b for b in all_boards if b["id"] == bid]
    else:
        visible_boards = all_boards

    # Load programmes per board
    board_programmes = {}
    for b in visible_boards:
        board_programmes[b["id"]] = [dict(r) for r in conn.execute(
            "SELECT * FROM programmes WHERE board_id=? ORDER BY programme_name", (b["id"],)
        ).fetchall()]

    # Load stages per programme
    prog_stages = {}
    for b in visible_boards:
        for p in board_programmes[b["id"]]:
            rows = conn.execute(
                "SELECT * FROM programme_config WHERE programme_name=? ORDER BY stage_order",
                (p["programme_name"],)
            ).fetchall()
            prog_stages[p["programme_name"]] = [dict(r) for r in rows]
            # Merge sender info from first row
            first = prog_stages[p["programme_name"]]
            p["sender_email"] = first[0]["sender_email"] if first else None
            p["smtp_host"] = first[0].get("smtp_host", "smtp.gmail.com") if first else "smtp.gmail.com"
            p["smtp_port"] = first[0].get("smtp_port", 587) if first else 587

    # Load board-specific holidays
    if bid is not None:
        _board_holidays = [dict(r) for r in conn.execute(
            "SELECT * FROM holidays WHERE board_id=? OR board_id IS NULL ORDER BY holiday_date",
            (bid,)
        ).fetchall()]
    else:
        _board_holidays = [dict(r) for r in conn.execute(
            "SELECT * FROM holidays ORDER BY holiday_date"
        ).fetchall()]

    conn.close()

    # Programmes list for "Add Stage" dropdown (scoped to user's boards)
    all_prog_names = []
    for b in visible_boards:
        for p in board_programmes[b["id"]]:
            all_prog_names.append(p["programme_name"])

    # Build accordion HTML: Board → Programme → Stages
    board_sections = ""
    for b in visible_boards:
        bid_safe = str(b["id"])
        prog_count = len(board_programmes[b["id"]])
        prog_inner = ""
        for p in board_programmes[b["id"]]:
            pname = p["programme_name"]
            pid_safe = pname.replace(" ", "_").replace("/", "_")
            stages = prog_stages.get(pname, [])
            sender_status = p.get("sender_email") or ""
            credential_badge = (
                f'<span class="pill pill-ok" style="font-size:11px"><i class="bi bi-shield-check"></i> {sender_status}</span>'
                if sender_status else
                '<span class="pill pill-warn" style="font-size:11px"><i class="bi bi-exclamation-circle"></i> No sender configured</span>'
            )
            is_ba = session.get("role") in ("super_admin", "board_admin")
            stages_rows = ""
            for s in stages:
                stage_flags = ""
                if s["is_milestone"]:
                    stage_flags += '<span style="background:#ede9fe;color:#5b21b6;font-size:10px;padding:1px 6px;border-radius:4px;margin-left:4px"><i class="bi bi-flag-fill"></i> Milestone</span>'
                if s.get("is_optional"):
                    stage_flags += '<span style="background:#e0f2fe;color:#0369a1;font-size:10px;padding:1px 6px;border-radius:4px;margin-left:4px"><i class="bi bi-skip-forward-fill"></i> Optional</span>'
                sname_js = s["stage_name"].replace("'", "\\'").replace('"', '&quot;')
                if is_ba:
                    edit_stage_btn = (
                        f'<button class="btn btn-sm btn-outline-primary py-0 px-1 me-1" style="font-size:11px" '
                        f'title="Edit stage" '
                        f'data-sid="{s["id"]}" '
                        f'data-name="{h(s["stage_name"])}" '
                        f'data-order="{s["stage_order"]}" '
                        f'data-tat="{s["tat_days"]}" '
                        f'data-r1="{s["reminder1_day"]}" '
                        f'data-r2="{s["reminder2_day"]}" '
                        f'data-owner="{h(s["owner_type"] or "")}" '
                        f'data-odi="{s["overdue_interval_days"]}" '
                        f'data-ms="{s["is_milestone"]}" '
                        f'data-opt="{int(s.get("is_optional", 0))}" '
                        f'onclick="editStageFromBtn(this)">'
                        f'<i class="bi bi-pencil"></i></button>'
                    )
                    del_stage_btn = (
                        f'<form method="post" style="display:inline" '
                        f'onsubmit="return confirm(\'Delete stage &quot;{sname_js}&quot;?\')">'
                        f'<input type="hidden" name="action" value="delete_stage">'
                        f'<input type="hidden" name="programme_name" value="{pname}">'
                        f'<input type="hidden" name="stage_name" value="{s["stage_name"]}">'
                        f'<button class="btn btn-sm btn-outline-danger py-0 px-1" style="font-size:11px">'
                        f'<i class="bi bi-trash"></i></button></form>'
                    )
                else:
                    edit_stage_btn = ""
                    del_stage_btn = ""
                stages_rows += f"""<tr>
                  <td style="color:#94a3b8;font-size:12px">{s["stage_order"]}</td>
                  <td style="font-weight:500">{s["stage_name"]}{stage_flags}</td>
                  <td style="text-align:center">{s["tat_days"] or "—"}</td>
                  <td style="text-align:center;color:#2563eb">{s["reminder1_day"] or "—"}</td>
                  <td style="text-align:center;color:#d97706">{s["reminder2_day"] or "—"}</td>
                  <td>{s["owner_type"] or "—"}</td>
                  <td style="text-align:center">{s["overdue_interval_days"]}</td>
                  <td style="white-space:nowrap">{edit_stage_btn}{del_stage_btn}</td>
                </tr>"""
            th_center = 'style="text-align:center"'
            del_col_header = '<th style="width:72px"></th>' if is_ba else ''
            if stages:
                stages_table_html = (
                    '<div style="overflow-x:auto"><table class="data-table"><thead><tr>'
                    '<th>#</th><th>Stage Name</th>'
                    f'<th {th_center}>TAT</th><th {th_center}>R1 Day</th>'
                    f'<th {th_center}>R2 Day</th><th>Owner</th><th {th_center}>OD Interval</th>'
                    + del_col_header +
                    '</tr></thead><tbody>' + stages_rows + '</tbody></table></div>'
                )
            else:
                stages_table_html = '<div style="padding:16px 20px;color:#94a3b8;font-size:13px">No stages configured yet.</div>'

            del_prog_btn = ""
            edit_prog_btn = ""
            if is_ba:
                _p_tat = p.get("tat_days") or 0
                _p_r1 = p.get("reminder1_days") or 0
                _p_r2 = p.get("reminder2_days") or 0
                _p_od = p.get("overdue_days") or 0
                _p_emails = h(p.get("notification_emails") or "")
                edit_prog_btn = (
                    f'<button class="btn btn-sm btn-outline-secondary ms-2 flex-shrink-0" '
                    f'style="font-size:11px;padding:4px 10px;white-space:nowrap" '
                    f'onclick="editProgramme({json.dumps(pname)}, {_p_tat}, {_p_r1}, {_p_r2}, {_p_od}, {json.dumps(_p_emails)})">'
                    f'<i class="bi bi-gear"></i> Edit</button>'
                )
                del_prog_btn = f"""
    <form method="post" class="ms-2 flex-shrink-0"
          onsubmit="return confirm('Delete programme &quot;{pname}&quot; and ALL its stages? This cannot be undone.')">
      <input type="hidden" name="action" value="delete_programme">
      <input type="hidden" name="programme_name" value="{pname}">
      <button class="btn btn-sm btn-danger" style="font-size:11px;padding:4px 10px;white-space:nowrap">
        <i class="bi bi-trash"></i> Delete
      </button>
    </form>"""
            prog_inner += f"""
<div class="accordion-item" style="border:1px solid #e2e8f0;border-radius:8px;margin-bottom:8px;overflow:hidden">
  <div class="d-flex align-items-center" style="background:#f8fafc;padding-right:8px;border-bottom:1px solid #e2e8f0">
    <button class="accordion-button collapsed flex-grow-1 py-2" type="button" data-bs-toggle="collapse"
            data-bs-target="#prog_{pid_safe}"
            style="background:transparent;font-size:14px;border:none;box-shadow:none">
      <div class="d-flex align-items-center gap-3 w-100">
        <i class="bi bi-list-task" style="color:#7c3aed"></i>
        <span style="font-weight:600">{pname}</span>
        <span>{credential_badge}</span>
        <span class="ms-auto" style="font-size:12px;color:#94a3b8">{len(stages)} stages</span>
      </div>
    </button>
    {edit_prog_btn}
    {del_prog_btn}
  </div>
  <div id="prog_{pid_safe}" class="accordion-collapse collapse">
    <div class="accordion-body p-0">
      {stages_table_html}
      <div style="padding:12px 20px;background:#f0fdf4;border-top:1px solid #e2e8f0">
        <div style="font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">
          <i class="bi bi-copy"></i> Copy Stages From Another Programme
        </div>
        <form method="post" class="row g-2 align-items-end">
          <input type="hidden" name="action" value="copy_stages_from">
          <input type="hidden" name="current_programme_name" value="{pname}">
          <div class="col">
            <label class="form-label" style="font-size:11px">Source Programme (same board)</label>
            <select class="form-select form-select-sm" name="source_programme_name" required>
              <option value="">— select source —</option>
              {''.join(f'<option value="{h(op["programme_name"])}">{h(op["programme_name"])}</option>' for op in board_programmes[b["id"]] if op["programme_name"] != pname)}
            </select>
          </div>
          <div class="col-auto d-flex align-items-end gap-2">
            <div class="form-check mb-0" style="white-space:nowrap">
              <input class="form-check-input" type="checkbox" name="confirm_replace" value="1"
                     id="cr_{pid_safe}" onchange="return this.checked ? confirm('This will DELETE all existing stages in &quot;{pname}&quot; first. Are you sure?') : true">
              <label class="form-check-label" style="font-size:11px" for="cr_{pid_safe}">Replace existing</label>
            </div>
            <button class="btn btn-sm btn-outline-success" type="submit"
                    onclick="return confirm('Copy stages from selected programme into &quot;{pname}&quot;?')">
              <i class="bi bi-copy"></i> Copy
            </button>
          </div>
        </form>
      </div>
      <div style="padding:16px 20px;background:#f8fafc;border-top:1px solid #e2e8f0">
        <div style="font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px">
          <i class="bi bi-envelope-at"></i> Outbound Email Settings
        </div>
        <form method="post" class="row g-2 align-items-end">
          <input type="hidden" name="action" value="update_sender">
          <input type="hidden" name="programme_name" value="{pname}">
          <div class="col-md-4">
            <label class="form-label" style="font-size:12px">Sender Email</label>
            <input type="email" class="form-control form-control-sm" name="sender_email"
                   value="{p.get('sender_email') or ''}" placeholder="notifications@yourdomain.com">
          </div>
          <div class="col-md-3">
            <label class="form-label" style="font-size:12px">Password <span style="color:#94a3b8">(blank = keep)</span></label>
            <input type="password" class="form-control form-control-sm" name="sender_password"
                   placeholder="••••••••" autocomplete="new-password">
          </div>
          <div class="col-md-3">
            <label class="form-label" style="font-size:12px">SMTP Host</label>
            <input type="text" class="form-control form-control-sm" name="smtp_host"
                   value="{p.get('smtp_host','smtp.gmail.com')}" placeholder="smtp.gmail.com">
          </div>
          <div class="col-md-1">
            <label class="form-label" style="font-size:12px">Port</label>
            <input type="number" class="form-control form-control-sm" name="smtp_port"
                   value="{p.get('smtp_port', 587)}" placeholder="587">
          </div>
          <div class="col-auto">
            <button class="btn btn-sm btn-primary" type="submit">Save</button>
          </div>
        </form>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px">
          Gmail: smtp.gmail.com:587 · Outlook: smtp.office365.com:587 · Port 465 = SSL
        </div>
      </div>
    </div>
  </div>
</div>"""

        del_board_btn = ""
        if session.get("role") == "super_admin":
            del_board_btn = f"""
    <form method="post" class="ms-2 flex-shrink-0"
          onsubmit="return confirm('Delete board &quot;{b['board_name']}&quot;? All programmes must be deleted first.')">
      <input type="hidden" name="action" value="delete_board">
      <input type="hidden" name="board_id" value="{b['id']}">
      <button class="btn btn-sm btn-danger" style="font-size:11px;padding:4px 10px;white-space:nowrap">
        <i class="bi bi-trash"></i> Delete Board
      </button>
    </form>"""
        board_sections += f"""
<div class="accordion-item" style="border:2px solid #e2e8f0;border-radius:12px;margin-bottom:16px;overflow:hidden">
  <div class="d-flex align-items-center" style="background:linear-gradient(135deg,#003356,#0094ca);padding-right:12px">
    <button class="accordion-button flex-grow-1" type="button" data-bs-toggle="collapse"
            data-bs-target="#board_{bid_safe}"
            style="background:transparent;color:#fff;font-weight:700;font-size:15px;border:none;box-shadow:none">
      <div class="d-flex align-items-center gap-3 w-100">
        <i class="bi bi-building" style="font-size:18px"></i>
        <span>{b['board_name']}</span>
        <span class="ms-auto" style="font-size:12px;opacity:.8">{prog_count} programme{"s" if prog_count != 1 else ""}</span>
      </div>
    </button>
    {del_board_btn}
  </div>
  <div id="board_{bid_safe}" class="accordion-collapse collapse show">
    <div class="accordion-body" style="background:#fafbfc;padding:16px">
      <div class="accordion" id="progAccordion_{bid_safe}">
        {prog_inner if prog_inner else '<div style="color:#94a3b8;font-size:13px;padding:8px">No programmes in this board yet.</div>'}
      </div>
    </div>
  </div>
</div>"""

    # Board options for add_programme form
    board_opts = "".join(
        f'<option value="{b["id"]}">{b["board_name"]}</option>'
        for b in visible_boards
    )
    # Programme options for add_stage form
    prog_opts = "".join(
        f'<option value="{pn}">{pn}</option>'
        for pn in all_prog_names
    )

    is_super = session.get("role") == "super_admin"
    is_board_admin = session.get("role") in ("board_admin", "super_admin")
    _settings_bid = user_board_id()

    # Per-board scheduler setting (used by board admins)
    if request.method == "POST" and request.form.get("action") == "save_board_schedule":
        if _settings_bid and is_board_admin:
            _bh = request.form.get("board_sched_hour", "8")
            _bm = request.form.get("board_sched_minute", "0")
            set_app_setting(f"sched_hour_board_{_settings_bid}", _bh)
            set_app_setting(f"sched_minute_board_{_settings_bid}", _bm)
            flash("Board notification schedule updated.", "success")

    _board_sched_hour = int(get_app_setting(
        f"sched_hour_board_{_settings_bid}" if _settings_bid else "scheduler_hour",
        get_app_setting("scheduler_hour", "8")
    ))
    _board_sched_minute = int(get_app_setting(
        f"sched_minute_board_{_settings_bid}" if _settings_bid else "scheduler_minute",
        get_app_setting("scheduler_minute", "0")
    ))
    _bsh_opts = "".join(
        f'<option value="{h}" {"selected" if h==_board_sched_hour else ""}>{h:02d}:00</option>'
        for h in range(24)
    )
    _bsm_opts = "".join(
        f'<option value="{m}" {"selected" if m==_board_sched_minute else ""}>{m:02d}</option>'
        for m in [0, 15, 30, 45]
    )

    board_schedule_html = f"""
      <div class="card mb-3">
        <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
             onclick="togglePanel('boardSchedBody')">
          <span><i class="bi bi-clock" style="color:#0891b2"></i> Board Notification Schedule</span>
          <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
        </div>
        <div id="boardSchedBody" style="display:none">
          <div class="card-body p-3">
            <form method="post">
              <input type="hidden" name="action" value="save_board_schedule">
              <div class="row g-2 mb-2">
                <div class="col-6">
                  <label class="form-label" style="font-size:12px">Hour (IST)</label>
                  <select class="form-select form-select-sm" name="board_sched_hour">{_bsh_opts}</select>
                </div>
                <div class="col-6">
                  <label class="form-label" style="font-size:12px">Minute</label>
                  <select class="form-select form-select-sm" name="board_sched_minute">{_bsm_opts}</select>
                </div>
              </div>
              <div style="font-size:11px;color:#94a3b8;margin-bottom:10px">
                Currently: <strong>{_board_sched_hour:02d}:{_board_sched_minute:02d} IST</strong>.
                Notifications for this board will run at this time daily.
              </div>
              <button class="btn btn-sm btn-primary w-100" type="submit">Save Schedule</button>
            </form>
          </div>
        </div>
      </div>""" if is_board_admin else ""

    scheduler_html = ""  # Removed — managed in System Settings (super admin only)

    # Board holiday calendar card
    _hol_rows = ""
    for _hol in _board_holidays:
        scope = '<span style="font-size:10px;color:#94a3b8">(global)</span>' if not _hol.get("board_id") else ""
        _hol_rows += f"""<tr>
  <td style="font-size:12px">{_hol['holiday_date']}</td>
  <td style="font-size:12px">{_hol['name']} {scope}</td>
  <td>
    <form method="post" style="display:inline">
      <input type="hidden" name="action" value="delete_board_holiday">
      <input type="hidden" name="holiday_id" value="{_hol['id']}">
      <button type="submit" class="btn btn-xs btn-outline-danger" style="font-size:10px;padding:1px 6px"
              onclick="return confirm('Remove holiday?')"><i class="bi bi-trash"></i></button>
    </form>
  </td>
</tr>"""
    if not _hol_rows:
        _hol_rows = '<tr><td colspan="3" style="text-align:center;color:#94a3b8;padding:16px;font-size:12px">No holidays added yet.</td></tr>'

    board_holiday_html = f"""
      <div class="card mb-3">
        <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
             onclick="togglePanel('boardHolBody')">
          <span><i class="bi bi-calendar3-event" style="color:#d97706"></i> Holiday Calendar</span>
          <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
        </div>
        <div id="boardHolBody" style="display:none">
          <div class="card-body p-3">
            <form method="post" class="row g-2 mb-3">
              <input type="hidden" name="action" value="add_board_holiday">
              <div class="col-5">
                <input type="date" class="form-control form-control-sm" name="holiday_date" required>
              </div>
              <div class="col-5">
                <input type="text" class="form-control form-control-sm" name="holiday_name"
                       placeholder="e.g. Diwali" required>
              </div>
              <div class="col-2">
                <button class="btn btn-sm btn-success w-100" type="submit">Add</button>
              </div>
            </form>
            <div style="max-height:180px;overflow-y:auto">
              <table class="data-table">
                <thead><tr><th>Date</th><th>Name</th><th style="width:40px"></th></tr></thead>
                <tbody>{_hol_rows}</tbody>
              </table>
            </div>
            <div style="font-size:10px;color:#94a3b8;margin-top:6px">
              TAT calculation skips these dates + all weekends.
            </div>
          </div>
        </div>
      </div>""" if is_board_admin else ""

    add_board_html = f"""
      <div class="card mb-3">
        <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
             onclick="togglePanel('addBoardBody')">
          <span><i class="bi bi-building-add" style="color:#7c3aed"></i> Add New Board</span>
          <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
        </div>
        <div id="addBoardBody" style="display:none">
          <div class="card-body p-3">
            <form method="post" class="row g-2 align-items-end">
              <input type="hidden" name="action" value="add_board">
              <div class="col">
                <label class="form-label" style="font-size:12px">Board Name</label>
                <input type="text" class="form-control form-control-sm" name="board_name"
                       required placeholder="e.g. NABH">
              </div>
              <div class="col-auto">
                <button class="btn btn-sm btn-primary" type="submit">Create Board</button>
              </div>
            </form>
          </div>
        </div>
      </div>""" if is_super else ""

    content = f"""
<div class="row g-4">
  <div class="col-lg-4">
    <!-- Action panels -->
    {board_schedule_html}
    {board_holiday_html}
    {add_board_html}

    <div class="card mb-3">
      <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
           onclick="togglePanel('addProgBody')">
        <span><i class="bi bi-folder-plus" style="color:#2563eb"></i> Add New Programme</span>
        <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
      </div>
      <div id="addProgBody" style="display:none">
        <div class="card-body p-3">
          <form method="post">
            <input type="hidden" name="action" value="add_programme">
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">Board</label>
              <select class="form-select form-select-sm" name="board_id" required>
                <option value="">— select —</option>
                {board_opts}
              </select>
            </div>
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">Programme Name</label>
              <input type="text" class="form-control form-control-sm" name="programme_name"
                     required placeholder="e.g. NABH Full Accreditation Hospitals">
            </div>
            <button class="btn btn-sm btn-primary w-100" type="submit">Create Programme</button>
          </form>
        </div>
      </div>
    </div>

    <div class="card">
      <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
           onclick="togglePanel('addStageBody')">
        <span><i class="bi bi-plus-circle" style="color:#059669"></i> Add New Stage</span>
        <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
      </div>
      <div id="addStageBody" style="display:none">
        <div class="card-body p-3">
          <form method="post">
            <input type="hidden" name="action" value="add_stage">
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">Programme</label>
              <select class="form-select form-select-sm" name="programme_name" required>
                <option value="">— select —</option>
                {prog_opts}
              </select>
            </div>
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">Stage Name</label>
              <input type="text" class="form-control form-control-sm" name="stage_name" required>
            </div>
            <div class="row g-2 mb-2">
              <div class="col-4">
                <label class="form-label" style="font-size:12px">Order #</label>
                <input type="number" class="form-control form-control-sm" name="stage_order" value="1" min="1">
              </div>
              <div class="col-4">
                <label class="form-label" style="font-size:12px">TAT Days</label>
                <input type="number" class="form-control form-control-sm" name="tat_days" value="0" min="0">
              </div>
              <div class="col-4">
                <label class="form-label" style="font-size:12px">OD Interval</label>
                <input type="number" class="form-control form-control-sm" name="overdue_interval_days" value="3" min="1">
              </div>
            </div>
            <div class="row g-2 mb-2">
              <div class="col-6">
                <label class="form-label" style="font-size:12px">R1 Day</label>
                <input type="number" class="form-control form-control-sm" name="reminder1_day" value="0" min="0">
              </div>
              <div class="col-6">
                <label class="form-label" style="font-size:12px">R2 Day</label>
                <input type="number" class="form-control form-control-sm" name="reminder2_day" value="0" min="0">
              </div>
            </div>
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">Owner Type</label>
              <select class="form-select form-select-sm" name="owner_type">
                <option value="">—</option>
                <option>Applicant</option>
                <option>Assessor</option>
                <option>Program Officer</option>
              </select>
            </div>
            <div class="d-flex gap-3 mb-3">
              <div class="form-check">
                <input class="form-check-input" type="checkbox" name="is_milestone" id="msCheck">
                <label class="form-check-label" style="font-size:12px" for="msCheck">
                  <i class="bi bi-flag-fill" style="color:#7c3aed;font-size:11px"></i> Milestone <span style="color:#94a3b8">(no emails)</span>
                </label>
              </div>
              <div class="form-check">
                <input class="form-check-input" type="checkbox" name="is_optional" id="optCheck">
                <label class="form-check-label" style="font-size:12px" for="optCheck">
                  <i class="bi bi-skip-forward-fill" style="color:#0891b2;font-size:11px"></i> Optional <span style="color:#94a3b8">(can skip)</span>
                </label>
              </div>
            </div>
            <button class="btn btn-sm btn-success w-100" type="submit">Add Stage</button>
          </form>
        </div>
      </div>
    </div>

    <div class="card mt-3">
      <div class="card-header d-flex justify-content-between align-items-center" style="cursor:pointer"
           onclick="togglePanel('copyStageBody')">
        <span><i class="bi bi-copy" style="color:#7c3aed"></i> Copy Stages from Programme</span>
        <i class="bi bi-chevron-down" style="color:#94a3b8"></i>
      </div>
      <div id="copyStageBody" style="display:none">
        <div class="card-body p-3">
          <p style="font-size:12px;color:#64748b;margin-bottom:12px">
            Clone all stages (TAT, reminders, owner) from one programme into another.
            Useful when creating a new programme with a similar workflow.
          </p>
          <form method="post">
            <input type="hidden" name="action" value="copy_stages">
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">
                <i class="bi bi-box-arrow-right" style="color:#059669"></i> Source Programme <span style="color:#94a3b8">(copy from)</span>
              </label>
              <select class="form-select form-select-sm" name="source_programme" required>
                <option value="">— select —</option>
                {prog_opts}
              </select>
            </div>
            <div class="mb-2">
              <label class="form-label" style="font-size:12px">
                <i class="bi bi-box-arrow-in-right" style="color:#2563eb"></i> Destination Programme <span style="color:#94a3b8">(copy to)</span>
              </label>
              <select class="form-select form-select-sm" name="dest_programme" required>
                <option value="">— select —</option>
                {prog_opts}
              </select>
            </div>
            <div class="form-check mb-3">
              <input class="form-check-input" type="checkbox" name="overwrite_existing" value="1" id="overwriteChk">
              <label class="form-check-label" style="font-size:12px" for="overwriteChk">
                Replace existing stages in destination
              </label>
            </div>
            <button class="btn btn-sm btn-primary w-100" type="submit"
                    onclick="return confirm('Copy all stages from the source? If overwrite is checked, existing stages in the destination will be replaced.')">
              <i class="bi bi-copy"></i> Copy Stages
            </button>
          </form>
        </div>
      </div>
    </div>
  </div>

  <div class="col-lg-8">
    <!-- Board → Programme → Stage accordion -->
    <div class="accordion" id="boardAccordion">
      {board_sections if board_sections else '<div class="card"><div class="card-body text-center" style="color:#94a3b8;padding:40px">No boards configured yet.</div></div>'}
    </div>
  </div>
</div>
"""
    scripts = """
<!-- Edit Stage Modal -->
<div class="modal fade" id="editStageModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header" style="background:#f8fafc;border-bottom:1px solid #e2e8f0">
        <h6 class="modal-title" style="font-weight:600">
          <i class="bi bi-pencil-square" style="color:#2563eb"></i> Edit Stage
        </h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <form method="post" id="editStageForm">
        <input type="hidden" name="action" value="update_stage">
        <input type="hidden" name="stage_id" id="es_id">
        <div class="modal-body">
          <div class="mb-3">
            <label class="form-label" style="font-size:12px;font-weight:600">Stage Name</label>
            <input type="text" class="form-control form-control-sm" name="stage_name" id="es_name" required>
          </div>
          <div class="row g-2 mb-3">
            <div class="col-4">
              <label class="form-label" style="font-size:12px;font-weight:600">Order #</label>
              <input type="number" class="form-control form-control-sm" name="stage_order" id="es_order" min="1">
            </div>
            <div class="col-4">
              <label class="form-label" style="font-size:12px;font-weight:600">TAT Days</label>
              <input type="number" class="form-control form-control-sm" name="tat_days" id="es_tat" min="0">
            </div>
            <div class="col-4">
              <label class="form-label" style="font-size:12px;font-weight:600">OD Interval</label>
              <input type="number" class="form-control form-control-sm" name="overdue_interval_days" id="es_odi" min="1">
            </div>
          </div>
          <div class="row g-2 mb-3">
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">R1 Reminder Day</label>
              <input type="number" class="form-control form-control-sm" name="reminder1_day" id="es_r1" min="0">
            </div>
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">R2 Reminder Day</label>
              <input type="number" class="form-control form-control-sm" name="reminder2_day" id="es_r2" min="0">
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label" style="font-size:12px;font-weight:600">Owner Type</label>
            <select class="form-select form-select-sm" name="owner_type" id="es_owner">
              <option value="">—</option>
              <option>Applicant</option>
              <option>Assessor</option>
              <option>Program Officer</option>
            </select>
          </div>
          <div class="d-flex gap-4">
            <div class="form-check">
              <input class="form-check-input" type="checkbox" name="is_milestone" id="es_ms">
              <label class="form-check-label" style="font-size:12px" for="es_ms">
                <i class="bi bi-flag-fill" style="color:#7c3aed;font-size:11px"></i> Milestone <span style="color:#94a3b8">(no emails)</span>
              </label>
            </div>
            <div class="form-check">
              <input class="form-check-input" type="checkbox" name="is_optional" id="es_opt">
              <label class="form-check-label" style="font-size:12px" for="es_opt">
                <i class="bi bi-skip-forward-fill" style="color:#0891b2;font-size:11px"></i> Optional <span style="color:#94a3b8">(can skip)</span>
              </label>
            </div>
          </div>
        </div>
        <div class="modal-footer" style="background:#f8fafc;border-top:1px solid #e2e8f0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button type="submit" class="btn btn-sm btn-primary">
            <i class="bi bi-check-lg"></i> Save Changes
          </button>
        </div>
      </form>
    </div>
  </div>
</div>
<script>
function togglePanel(id){
  var el = document.getElementById(id);
  el.style.display = el.style.display === 'none' ? '' : 'none';
}
function editStageFromBtn(btn) {
  var d = btn.dataset;
  editStage(d.sid, d.name, d.order, d.tat, d.r1, d.r2, d.owner, d.odi, d.ms, d.opt);
}
function editStage(id, name, order, tat, r1, r2, owner, odi, ms, opt) {
  document.getElementById('es_id').value = id;
  document.getElementById('es_name').value = name;
  document.getElementById('es_order').value = order;
  document.getElementById('es_tat').value = tat;
  document.getElementById('es_r1').value = r1;
  document.getElementById('es_r2').value = r2;
  document.getElementById('es_odi').value = odi;
  var ownerSel = document.getElementById('es_owner');
  for (var i = 0; i < ownerSel.options.length; i++) {
    ownerSel.options[i].selected = (ownerSel.options[i].value === owner);
  }
  document.getElementById('es_ms').checked = (ms == 1);
  document.getElementById('es_opt').checked = (opt == 1);
  new bootstrap.Modal(document.getElementById('editStageModal')).show();
}
function editProgramme(name, tat, r1, r2, od, emails) {
  document.getElementById('ep_name').value = name;
  document.getElementById('ep_tat').value = tat;
  document.getElementById('ep_r1').value = r1;
  document.getElementById('ep_r2').value = r2;
  document.getElementById('ep_od').value = od;
  document.getElementById('ep_emails').value = emails;
  new bootstrap.Modal(document.getElementById('editProgModal')).show();
}
</script>
<!-- Edit Programme Modal -->
<div class="modal fade" id="editProgModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header" style="background:#f8fafc;border-bottom:1px solid #e2e8f0">
        <h6 class="modal-title" style="font-weight:600">
          <i class="bi bi-gear" style="color:#7c3aed"></i> Edit Programme Settings
        </h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <form method="post" id="editProgForm">
        <input type="hidden" name="action" value="edit_programme">
        <input type="hidden" name="programme_name" id="ep_name">
        <div class="modal-body">
          <div class="row g-2 mb-3">
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">Default TAT Days</label>
              <input type="number" class="form-control form-control-sm" name="tat_days" id="ep_tat" min="0" value="0">
            </div>
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">Overdue Days</label>
              <input type="number" class="form-control form-control-sm" name="overdue_days" id="ep_od" min="0" value="0">
            </div>
          </div>
          <div class="row g-2 mb-3">
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">Reminder 1 Days</label>
              <input type="number" class="form-control form-control-sm" name="reminder1_days" id="ep_r1" min="0" value="0">
            </div>
            <div class="col-6">
              <label class="form-label" style="font-size:12px;font-weight:600">Reminder 2 Days</label>
              <input type="number" class="form-control form-control-sm" name="reminder2_days" id="ep_r2" min="0" value="0">
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label" style="font-size:12px;font-weight:600">Notification Emails <span style="color:#94a3b8;font-weight:400">(comma-separated)</span></label>
            <input type="text" class="form-control form-control-sm" name="notification_emails" id="ep_emails"
                   placeholder="admin@org.com, manager@org.com">
            <div style="font-size:11px;color:#94a3b8;margin-top:4px">These addresses receive programme-level notification copies.</div>
          </div>
        </div>
        <div class="modal-footer" style="background:#f8fafc;border-top:1px solid #e2e8f0">
          <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
          <button type="submit" class="btn btn-sm btn-primary">
            <i class="bi bi-check-lg"></i> Save Changes
          </button>
        </div>
      </form>
    </div>
  </div>
</div>"""
    scripts += """
<script>
(function() {
  document.addEventListener('DOMContentLoaded', function() {
    document.querySelectorAll('input[name="tat_days"]').forEach(function(inp) {
      inp.addEventListener('change', function() {
        var tr = this.closest('tr');
        var stage = tr ? (tr.querySelector('[name="stage_name"]') || {}).value : '';
        var prog  = (document.querySelector('[name="programme_name"]') || {}).value || '';
        if (!stage || !prog || !this.value) return;
        fetch('/api/preview-tat-impact?programme='+encodeURIComponent(prog)+'&stage='+encodeURIComponent(stage)+'&tat='+encodeURIComponent(this.value))
          .then(function(r){return r.json();}).then(function(d) {
            var wrap = inp.parentElement;
            var prev = wrap.querySelector('.tat-impact-preview');
            if (!prev) { prev = document.createElement('div'); prev.className='tat-impact-preview'; prev.style.cssText='font-size:11px;margin-top:2px'; wrap.appendChild(prev); }
            if (d.error) { prev.innerHTML=''; return; }
            if (d.newly_flagged > 0) prev.innerHTML='<span style="color:#dc2626">&#9888; '+d.newly_flagged+' case(s) become overdue</span>';
            else if (d.newly_resolved > 0) prev.innerHTML='<span style="color:#00984C">&#10003; '+d.newly_resolved+' case(s) no longer overdue</span>';
            else prev.innerHTML='<span style="color:#64748b">No impact on '+d.total_at_stage+' current case(s)</span>';
          }).catch(function(){});
      });
    });
  });
})();
</script>"""
    return render_page(content, scripts, active_page="settings",
                       page_title="Programme Settings")


@app.route("/templates", methods=["GET", "POST"])
@board_admin_required
def email_templates_page():
    conn = get_db()

    if request.method == "POST":
        action = request.form.get("action")
        if action == "save":
            tmpl_id = request.form.get("id")
            if tmpl_id:
                conn.execute(
                    "UPDATE email_templates SET subject_line=?, email_body=? WHERE id=?",
                    (request.form["subject_line"], request.form["email_body"], tmpl_id),
                )
            else:
                conn.execute(
                    "INSERT INTO email_templates "
                    "(programme_name, notification_type, subject_line, email_body) VALUES (?,?,?,?) "
                    "ON CONFLICT(programme_name, notification_type) "
                    "DO UPDATE SET subject_line=EXCLUDED.subject_line, email_body=EXCLUDED.email_body",
                    (request.form["programme_name"], request.form["notification_type"],
                     request.form["subject_line"], request.form["email_body"]),
                )
            conn.commit()
            flash("Template saved.", "success")
        elif action == "save_stage_override":
            conn.execute(
                "INSERT INTO stage_email_override "
                "(programme_name, stage_name, notification_type, subject_line, email_body) VALUES (?,?,?,?,?) "
                "ON CONFLICT (programme_name, stage_name, notification_type) DO UPDATE SET subject_line=EXCLUDED.subject_line, email_body=EXCLUDED.email_body",
                (request.form["programme_name"], request.form["stage_name"],
                 request.form["notification_type"],
                 request.form["subject_line"], request.form["email_body"]),
            )
            conn.commit()
            flash(f"Stage override saved for '{request.form['stage_name']}'.", "success")
        elif action == "delete_stage_override":
            conn.execute("DELETE FROM stage_email_override WHERE id=?", (request.form["override_id"],))
            conn.commit()
            flash("Stage override deleted.", "success")

    et_bid = user_board_id()
    if et_bid is not None:
        templates = [dict(r) for r in conn.execute(
            "SELECT * FROM email_templates WHERE board_id=? ORDER BY programme_name, notification_type",
            (et_bid,)
        ).fetchall()]
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name",
            (et_bid,)
        ).fetchall()]
        stage_overrides = [dict(r) for r in conn.execute(
            "SELECT seo.* FROM stage_email_override seo "
            "JOIN programmes p ON seo.programme_name = p.programme_name "
            "WHERE p.board_id=? ORDER BY seo.programme_name, seo.stage_name",
            (et_bid,)
        ).fetchall()]
        stages_rows = conn.execute(
            "SELECT DISTINCT programme_name, stage_name, stage_order FROM programme_config "
            "WHERE programme_name IN (SELECT programme_name FROM programmes WHERE board_id=?) "
            "ORDER BY programme_name, stage_order",
            (et_bid,)
        ).fetchall()
    else:
        templates = [dict(r) for r in conn.execute(
            "SELECT * FROM email_templates ORDER BY programme_name, notification_type"
        ).fetchall()]
        programmes = [r[0] for r in conn.execute(
            "SELECT DISTINCT programme_name FROM programme_config ORDER BY programme_name"
        ).fetchall()]
        stage_overrides = [dict(r) for r in conn.execute(
            "SELECT * FROM stage_email_override ORDER BY programme_name, stage_name"
        ).fetchall()]
        stages_rows = conn.execute(
            "SELECT DISTINCT programme_name, stage_name, stage_order FROM programme_config ORDER BY programme_name, stage_order"
        ).fetchall()
    conn.close()

    stages_by_prog = {}
    for _r in stages_rows:
        stages_by_prog.setdefault(_r[0], []).append(_r[1])
    stages_json = json.dumps(stages_by_prog)

    PLACEHOLDER_HELP = (
        "{{Organisation_Name}}, {{Stage_Name}}, {{Action_Owner_Name}}, "
        "{{Days_Remaining}}, {{TAT_Days}}, {{Stage_Start_Date}}, "
        "{{Programme_Name}}, {{Followup_Count}}, {{PO_Name}}"
    )

    TYPE_META = {
        "R1":      ("bi-bell",            "#2563eb", "#dbeafe", "Friendly reminder — sent at R1 day"),
        "R2":      ("bi-clock-history",   "#d97706", "#fef3c7", "Urgent reminder — sent at R2 day"),
        "Overdue": ("bi-exclamation-octagon","#dc2626","#fee2e2","TAT breached — sent when overdue"),
        "Followup":("bi-arrow-clockwise", "#7c3aed", "#ede9fe", "Repeated follow-up — every overdue interval"),
    }

    tmpl_forms = ""
    for t in templates:
        icon, color, bg, desc = TYPE_META.get(t['notification_type'], ("bi-envelope","#64748b","#f1f5f9",""))
        safe_body = t['email_body'].replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        safe_subject = t['subject_line'].replace('"','&quot;')
        tmpl_forms += f"""
<div class="card mb-3">
  <div class="card-header d-flex align-items-center gap-3">
    <div style="width:34px;height:34px;border-radius:8px;background:{bg};color:{color};
                display:flex;align-items:center;justify-content:center;font-size:16px">
      <i class="bi {icon}"></i>
    </div>
    <div>
      <div style="font-weight:600">{t['notification_type']} — {t['programme_name']}</div>
      <div style="font-size:12px;color:#94a3b8">{desc}</div>
    </div>
  </div>
  <div class="card-body p-4">
    <form method="post">
      <input type="hidden" name="action" value="save">
      <input type="hidden" name="id" value="{t['id']}">
      <div class="mb-3">
        <label class="form-label">Subject Line</label>
        <input type="text" class="form-control" name="subject_line" value="{safe_subject}">
      </div>
      <div class="mb-3">
        <label class="form-label">Email Body</label>
        <textarea class="form-control" name="email_body" rows="10"
                  style="font-family:monospace;font-size:13px">{safe_body}</textarea>
      </div>
      <button type="submit" class="btn btn-sm btn-primary">
        <i class="bi bi-save"></i> Save Template
      </button>
    </form>
  </div>
</div>"""

    prog_opts = "".join(f'<option value="{p}">{p}</option>' for p in programmes)

    # ── Stage override rows ────────────────────────────────────────────────────
    sov_rows = ""
    for sov in stage_overrides:
        safe_subj = sov['subject_line'].replace('"', '&quot;')
        sov_rows += f"""<tr>
  <td style="font-weight:500">{sov['programme_name']}</td>
  <td>{sov['stage_name']}</td>
  <td><span class="badge" style="background:#ede9fe;color:#7c3aed">{sov['notification_type']}</span></td>
  <td style="font-size:12px;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{sov['subject_line']}</td>
  <td>
    <button class="btn btn-xs btn-outline-primary" style="font-size:11px;padding:2px 8px"
            onclick="editOverride({sov['id']},'{sov['programme_name'].replace(chr(39),'&#39;')}','{sov['stage_name'].replace(chr(39),'&#39;')}','{sov['notification_type']}','{safe_subj}',this)">
      <i class="bi bi-pencil"></i>
    </button>
    <form method="post" style="display:inline" onsubmit="return confirm('Delete this stage override?')">
      <input type="hidden" name="action" value="delete_stage_override">
      <input type="hidden" name="override_id" value="{sov['id']}">
      <button type="submit" class="btn btn-xs btn-outline-danger" style="font-size:11px;padding:2px 8px">
        <i class="bi bi-trash"></i>
      </button>
    </form>
  </td>
</tr>"""
    if not sov_rows:
        sov_rows = '<tr><td colspan="5" style="text-align:center;color:#94a3b8;padding:24px">No stage-specific overrides yet.</td></tr>'

    ph_chips = " ".join(
        f'<span class="ph-chip" onclick="insertPH(this.textContent)">{{{{{p}}}}}</span>'
        for p in ["Organisation_Name","Stage_Name","Action_Owner_Name","Days_Remaining",
                  "TAT_Days","Stage_Start_Date","Programme_Name","Followup_Count","PO_Name"]
    )

    content = f"""
<!-- ── Stage-level overrides ─────────────────────────────────────────────── -->
<div class="card mb-4">
  <div class="card-header d-flex justify-content-between align-items-center">
    <span><i class="bi bi-layers" style="color:#7c3aed"></i> Per-Stage Template Overrides
      <span style="font-size:12px;color:#94a3b8;font-weight:400;margin-left:8px">
        — override a programme template for a specific stage</span>
    </span>
    <button class="btn btn-sm btn-outline-primary" data-bs-toggle="collapse"
            data-bs-target="#stageOverrideForm">
      <i class="bi bi-plus-circle"></i> Add Override
    </button>
  </div>

  <!-- Add/Edit form (collapsed by default) -->
  <div id="stageOverrideForm" class="collapse">
    <div class="card-body border-bottom" style="background:#f8fafc">
      <form method="post" id="overrideForm">
        <input type="hidden" name="action" value="save_stage_override">
        <input type="hidden" name="override_edit_id" id="overrideEditId" value="">
        <div class="row g-3">
          <div class="col-md-3">
            <label class="form-label">Programme</label>
            <select class="form-select form-select-sm" name="programme_name" id="sovProg" onchange="loadSovStages()" required>
              <option value="">— select —</option>
              {prog_opts}
            </select>
          </div>
          <div class="col-md-3">
            <label class="form-label">Stage</label>
            <select class="form-select form-select-sm" name="stage_name" id="sovStage" required>
              <option value="">— select programme first —</option>
            </select>
          </div>
          <div class="col-md-2">
            <label class="form-label">Notification Type</label>
            <select class="form-select form-select-sm" name="notification_type" required>
              <option value="R1">R1 — Reminder 1</option>
              <option value="R2">R2 — Reminder 2</option>
              <option value="Overdue">Overdue</option>
              <option value="Followup">Followup</option>
            </select>
          </div>
          <div class="col-md-4">
            <label class="form-label">Subject Line</label>
            <input type="text" class="form-control form-control-sm" name="subject_line" id="sovSubject" required>
          </div>
          <div class="col-12">
            <label class="form-label">Email Body</label>
            <textarea class="form-control form-control-sm" name="email_body" id="sovBody"
                      rows="8" style="font-family:monospace;font-size:12px" required></textarea>
          </div>
          <div class="col-12">
            <button type="submit" class="btn btn-sm btn-primary">
              <i class="bi bi-save"></i> Save Stage Override
            </button>
            <button type="button" class="btn btn-sm btn-outline-secondary ms-2"
                    onclick="document.getElementById('overrideForm').reset();document.getElementById('stageOverrideForm').classList.remove('show')">
              Cancel
            </button>
          </div>
        </div>
      </form>
    </div>
  </div>

  <!-- Existing overrides table -->
  <div style="overflow-x:auto">
    <table class="data-table">
      <thead><tr>
        <th>Programme</th><th>Stage</th><th>Type</th><th>Subject</th><th>Actions</th>
      </tr></thead>
      <tbody>{sov_rows}</tbody>
    </table>
  </div>
</div>

<div class="row g-4">
  <div class="col-xl-8">

    <!-- Placeholder reference -->
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:14px 18px;margin-bottom:20px">
      <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#94a3b8;margin-bottom:8px">
        <i class="bi bi-braces"></i> Available Placeholders — click to copy
      </div>
      <div>{ph_chips}</div>
    </div>

    <!-- Existing templates -->
    {tmpl_forms}
  </div>

  <div class="col-xl-4">
    <div class="card" style="position:sticky;top:72px">
      <div class="card-header"><i class="bi bi-plus-circle" style="color:#059669"></i> Add New Template</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="save">
          <div class="mb-3">
            <label class="form-label">Programme</label>
            <select class="form-select" name="programme_name" required>
              {prog_opts}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label">Notification Type</label>
            <select class="form-select" name="notification_type" required>
              <option value="R1">R1 — Friendly Reminder</option>
              <option value="R2">R2 — Deadline Approaching</option>
              <option value="Overdue">Overdue — TAT Breached</option>
              <option value="Followup">Followup — Repeat Notice</option>
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label">Subject Line</label>
            <input type="text" class="form-control" name="subject_line" required>
          </div>
          <div class="mb-3">
            <label class="form-label">Email Body</label>
            <textarea class="form-control" name="email_body" rows="10"
                      style="font-family:monospace;font-size:12px" required></textarea>
          </div>
          <button type="submit" class="btn btn-success w-100">Create Template</button>
        </form>
      </div>
    </div>
  </div>
</div>
"""
    scripts = f"""<script>
var _STAGES_BY_PROG = {stages_json};

function insertPH(text){{
  navigator.clipboard.writeText(text).then(function(){{
    var t = document.createElement('div');
    t.className = 'toast qci-toast show align-items-center text-white bg-primary border-0';
    t.style.cssText = 'position:fixed;bottom:24px;right:24px;z-index:9999;min-width:220px';
    t.innerHTML = '<div class="d-flex"><div class="toast-body">Copied: '+text+'</div></div>';
    document.body.appendChild(t);
    setTimeout(function(){{ t.remove(); }}, 2000);
  }});
}}

function loadSovStages(){{
  var prog = document.getElementById('sovProg').value;
  var sel = document.getElementById('sovStage');
  sel.innerHTML = '';
  var stages = _STAGES_BY_PROG[prog] || [];
  if(!stages.length){{ sel.innerHTML='<option value="">— no stages —</option>'; return; }}
  stages.forEach(function(s){{
    var o = document.createElement('option'); o.value=s; o.textContent=s; sel.appendChild(o);
  }});
}}

function editOverride(id, prog, stage, notifType, subject, btn){{
  // populate form and expand
  document.getElementById('overrideEditId').value = id;
  document.getElementById('sovProg').value = prog;
  loadSovStages();
  document.getElementById('sovStage').value = stage;
  document.querySelector('#overrideForm select[name=notification_type]').value = notifType;
  document.getElementById('sovSubject').value = subject;
  // expand panel
  var panel = document.getElementById('stageOverrideForm');
  panel.classList.add('show');
  panel.scrollIntoView({{behavior:'smooth', block:'nearest'}});
}}
</script>"""
    return render_page(content, scripts, active_page="templates",
                       page_title="Email Templates")


# ── Audit Log page ────────────────────────────────────────────────────────────
@app.route("/audit-log")
@board_admin_required
def audit_log_page():
    conn = get_db()
    bid = user_board_id()
    if bid is not None:
        rows = conn.execute(
            "SELECT * FROM audit_log WHERE board_id=? OR board_id IS NULL ORDER BY id DESC LIMIT 500",
            (bid,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 500").fetchall()
    conn.close()

    EVENT_ICONS = {
        "case_created": ("bi-plus-circle-fill", "#00984C"),
        "case_updated": ("bi-pencil-fill", "#0094ca"),
        "case_closed":  ("bi-x-circle-fill", "#dc2626"),
        "stage_change": ("bi-arrow-right-circle-fill", "#7c3aed"),
        "stage_skipped":("bi-skip-forward-fill",       "#0891b2"),
        "email_sent":   ("bi-envelope-check-fill", "#00984C"),
        "email_error":  ("bi-envelope-x-fill", "#dc2626"),
        "bulk_upload":  ("bi-upload", "#0094ca"),
        "bulk_advance": ("bi-fast-forward-fill", "#d97706"),
    }

    tbl_rows = ""
    for r in rows:
        r = dict(r)
        icon, clr = EVENT_ICONS.get(r["event_type"], ("bi-circle", "#94a3b8"))
        tbl_rows += f"""<tr>
  <td style="white-space:nowrap;font-size:12px;color:#64748b">{h(r['timestamp'])}</td>
  <td><i class="bi {icon}" style="color:{clr};margin-right:4px"></i>
    <span style="font-size:12px;font-weight:600">{h(r['event_type'].replace('_',' ').title())}</span></td>
  <td style="font-weight:500">{h(r['application_id']) if r['application_id'] else '—'}</td>
  <td style="font-size:12.5px">{h(r['detail']) if r['detail'] else ''}</td>
  <td style="font-size:12px;color:#94a3b8">{h(r['user_name']) if r['user_name'] else 'system'}</td>
</tr>"""

    content = f"""
<div class="card">
  <div class="card-header d-flex justify-content-between align-items-center">
    <span><i class="bi bi-journal-text" style="color:var(--accent)"></i> Audit Log</span>
    <span style="font-size:12px;color:#94a3b8">Last 500 events</span>
  </div>
  <div style="overflow-x:auto">
    <table class="data-table">
      <thead><tr><th>Timestamp</th><th>Event</th><th>Application</th><th>Detail</th><th>User</th></tr></thead>
      <tbody>{tbl_rows if tbl_rows else '<tr><td colspan="5" style="text-align:center;color:#94a3b8;padding:40px">No audit events yet.</td></tr>'}</tbody>
    </table>
  </div>
</div>"""
    return render_page(content, active_page="audit", page_title="Audit Log")


# ── Case History page ─────────────────────────────────────────────────────────
@app.route("/case-history/<app_id>")
@login_required
def case_history(app_id):
    conn = get_db()
    # IDOR: verify the case belongs to the caller's board
    _ch_case = conn.execute(
        "SELECT board_id FROM case_tracking WHERE application_id=?", (app_id,)
    ).fetchone()
    if not _ch_case:
        conn.close()
        flash("Case not found.", "error")
        return redirect(url_for("dashboard"))
    _ch_bid = user_board_id()
    if _ch_bid is not None and _ch_case["board_id"] != _ch_bid:
        conn.close()
        flash("Access denied.", "error")
        return redirect(url_for("dashboard"))
    transitions = [dict(r) for r in conn.execute(
        "SELECT * FROM stage_history WHERE application_id=? ORDER BY id ASC", (app_id,)
    ).fetchall()]
    audits = [dict(r) for r in conn.execute(
        "SELECT * FROM audit_log WHERE application_id=? ORDER BY id ASC", (app_id,)
    ).fetchall()]
    conn.close()

    timeline_html = ""
    for t in transitions:
        from_lbl = h(t["from_stage"]) if t["from_stage"] else '<em style="color:#94a3b8">New Case</em>'
        timeline_html += f"""
<div class="d-flex align-items-start gap-3 mb-3">
  <div style="width:12px;height:12px;border-radius:50%;background:#7c3aed;margin-top:4px;flex-shrink:0"></div>
  <div>
    <div style="font-size:13px"><span style="color:#94a3b8">{from_lbl}</span>
      <i class="bi bi-arrow-right" style="margin:0 6px;color:#7c3aed"></i>
      <strong>{h(t['to_stage'])}</strong></div>
    <div style="font-size:11px;color:#94a3b8">{h(t['timestamp'])} · by {h(t['changed_by']) if t['changed_by'] else 'system'}</div>
  </div>
</div>"""

    _AH_ICONS = {
        "stage_change":  ("bi-arrow-right-circle-fill", "#7c3aed"),
        "stage_skipped": ("bi-skip-forward-fill",       "#0891b2"),
        "case_created":  ("bi-plus-circle-fill",        "#00984C"),
        "case_updated":  ("bi-pencil-fill",             "#0094ca"),
        "email_sent":    ("bi-envelope-check-fill",     "#00984C"),
        "email_error":   ("bi-envelope-x-fill",         "#dc2626"),
        "bulk_upload":   ("bi-upload",                  "#0094ca"),
    }
    audit_rows = ""
    for a in audits:
        _ah_icon, _ah_clr = _AH_ICONS.get(a["event_type"], ("bi-circle", "#94a3b8"))
        _ah_label = a['event_type'].replace('_', ' ').title()
        _ah_bg = "background:#f0f9ff" if a["event_type"] == "stage_skipped" else ""
        audit_rows += f"""<tr style="{_ah_bg}">
  <td style="font-size:12px;color:#64748b">{a['timestamp']}</td>
  <td style="font-size:12.5px;font-weight:500">
    <i class="bi {_ah_icon}" style="color:{_ah_clr};margin-right:4px"></i>{_ah_label}
  </td>
  <td style="font-size:12.5px">{a['detail'] or ''}</td>
  <td style="font-size:12px;color:#94a3b8">{a['user_name'] or 'system'}</td>
</tr>"""

    content = f"""
<div class="row g-4" style="max-width:960px;margin:0 auto">
  <div class="col-lg-5">
    <div class="card">
      <div class="card-header"><i class="bi bi-signpost-split-fill" style="color:#7c3aed"></i> Stage Timeline</div>
      <div class="card-body">
        {timeline_html if timeline_html else '<div style="color:#94a3b8;font-size:13px">No stage transitions recorded yet.</div>'}
      </div>
    </div>
  </div>
  <div class="col-lg-7">
    <div class="card">
      <div class="card-header"><i class="bi bi-clock-history" style="color:var(--accent)"></i> Activity Log</div>
      <div style="overflow-x:auto">
        <table class="data-table">
          <thead><tr><th>Time</th><th>Event</th><th>Detail</th><th>User</th></tr></thead>
          <tbody>{audit_rows if audit_rows else '<tr><td colspan="4" style="text-align:center;color:#94a3b8;padding:20px">No events.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="dashboard", page_title=f"Case History — {app_id}",
                       page_crumb=f'<a href="/">Dashboard</a> / {app_id}')


# ── Email Preview / Test Send ─────────────────────────────────────────────────
@app.route("/email-preview", methods=["GET", "POST"])
@board_admin_required
def email_preview():
    conn = get_db()
    bid = user_board_id()
    if bid is not None:
        progs = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
    else:
        progs = [r[0] for r in conn.execute(
            "SELECT DISTINCT programme_name FROM programme_config ORDER BY programme_name"
        ).fetchall()]

    preview_html = ""
    if request.method == "POST":
        action = request.form.get("action")
        programme = request.form.get("programme_name", "")
        ntype = request.form.get("notification_type", "R1")

        tmpl = conn.execute(
            "SELECT * FROM email_templates WHERE programme_name=? AND notification_type=?",
            (programme, ntype)
        ).fetchone()

        sample_ph = {
            "Organisation_Name": "Sample Hospital Ltd.",
            "Stage_Name": "DR In Progress",
            "Action_Owner_Name": "Dr. Sharma",
            "Days_Remaining": "5",
            "TAT_Days": "10",
            "Stage_Start_Date": date.today().strftime("%Y-%m-%d"),
            "Programme_Name": programme,
            "Followup_Count": "1",
            "PO_Name": session.get("full_name", "Program Officer"),
        }

        if tmpl:
            tmpl = dict(tmpl)
            subj = tmpl["subject_line"]
            body = tmpl["email_body"]
            for k, v in sample_ph.items():
                subj = subj.replace("{{" + k + "}}", str(v))
                body = body.replace("{{" + k + "}}", str(v))
        else:
            subj = f"[{ntype}] Reminder — Sample"
            body = "No template found for this programme/type."

        preview_html = f"""
<div class="card mt-3">
  <div class="card-header d-flex justify-content-between align-items-center">
    <span><i class="bi bi-eye" style="color:#7c3aed"></i> Email Preview</span>
    <span class="pill pill-muted">{ntype}</span>
  </div>
  <div class="card-body">
    <div style="font-size:12px;color:#64748b;margin-bottom:4px">Subject:</div>
    <div style="font-weight:600;font-size:14px;margin-bottom:16px;padding:8px 12px;background:#f8fafc;border-radius:6px;border:1px solid var(--border)">{subj}</div>
    <div style="font-size:12px;color:#64748b;margin-bottom:4px">Body:</div>
    <div style="white-space:pre-wrap;font-size:13px;line-height:1.6;padding:16px;background:#fafbfc;border-radius:8px;border:1px solid var(--border)">{body}</div>
  </div>
</div>"""

        if action == "test_send":
            test_email = request.form.get("test_email", "").strip()
            if test_email and programme:
                cfg = conn.execute(
                    "SELECT sender_email, sender_password, smtp_host, smtp_port "
                    "FROM programme_config WHERE programme_name=? AND sender_email IS NOT NULL LIMIT 1",
                    (programme,)
                ).fetchone()
                if cfg and cfg["sender_email"]:
                    ok, err = send_notification(
                        programme, ntype, test_email, "",
                        cfg["sender_email"], decrypt_str(cfg["sender_password"]),
                        sample_ph, cfg["smtp_host"] or "smtp.gmail.com",
                        cfg["smtp_port"] or 587
                    )
                    if ok:
                        flash(f"Test email sent to {test_email}!", "success")
                    else:
                        flash(f"Send failed: {err}", "error")
                else:
                    flash("No sender credentials configured for this programme.", "error")
            else:
                flash("Enter a test email address.", "error")

    conn.close()

    prog_opts = "".join(f'<option value="{p}">{p}</option>' for p in progs)
    content = f"""
<div style="max-width:800px;margin:0 auto">
  <div class="card">
    <div class="card-header"><i class="bi bi-envelope-open" style="color:var(--accent)"></i> Email Preview &amp; Test Send</div>
    <div class="card-body p-4">
      <form method="post">
        <div class="row g-3 mb-3">
          <div class="col-md-5">
            <label class="form-label">Programme</label>
            <select class="form-select" name="programme_name" required>
              <option value="">— select —</option>
              {prog_opts}
            </select>
          </div>
          <div class="col-md-3">
            <label class="form-label">Notification Type</label>
            <select class="form-select" name="notification_type">
              <option value="R1">R1 — Reminder 1</option>
              <option value="R2">R2 — Reminder 2</option>
              <option value="OVERDUE">Overdue</option>
              <option value="FOLLOWUP">Follow-up</option>
            </select>
          </div>
          <div class="col-md-4">
            <label class="form-label">Test Email <span style="color:#94a3b8;font-size:11px">(for test send)</span></label>
            <input type="email" class="form-control" name="test_email" placeholder="your@email.com">
          </div>
        </div>
        <div class="d-flex gap-2">
          <button type="submit" name="action" value="preview" class="btn btn-primary">
            <i class="bi bi-eye"></i> Preview
          </button>
          <button type="submit" name="action" value="test_send" class="btn btn-success">
            <i class="bi bi-send"></i> Send Test Email
          </button>
        </div>
      </form>
      {preview_html}
      <div style="margin-top:16px;font-size:12px;color:#94a3b8">
        <i class="bi bi-info-circle"></i> Preview uses sample data. Test Send uses the programme's configured SMTP credentials.
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="templates", page_title="Email Preview & Test")


# ── Bulk Stage Advance ────────────────────────────────────────────────────────
@app.route("/bulk-advance", methods=["GET", "POST"])
@login_required
def bulk_advance():
    conn = get_db()
    bid = user_board_id()

    if bid is not None:
        cases = [dict(r) for r in conn.execute(
            "SELECT * FROM case_tracking WHERE board_id=? ORDER BY programme_name, application_id", (bid,)
        ).fetchall()]
    else:
        cases = [dict(r) for r in conn.execute(
            "SELECT * FROM case_tracking ORDER BY programme_name, application_id"
        ).fetchall()]

    if request.method == "POST":
        selected_ids = request.form.getlist("case_ids")
        target_stage = request.form.get("target_stage", "").strip()
        new_start_date = request.form.get("new_start_date", date.today().strftime("%Y-%m-%d"))

        if not selected_ids or not target_stage:
            flash("Select cases and a target stage.", "error")
        else:
            advanced = 0
            errors = []
            user_name = session.get("full_name") or session.get("username", "")
            for cid in selected_ids:
                case = conn.execute("SELECT * FROM case_tracking WHERE id=?", (cid,)).fetchone()
                if not case:
                    continue
                case = dict(case)
                upsert_data = {
                    "application_id":      case["application_id"],
                    "organisation_name":   case["organisation_name"],
                    "programme_name":      case["programme_name"],
                    "stage_name":          target_stage,
                    "stage_start_date":    new_start_date,
                    "action_owner_name":   case.get("action_owner_name") or "",
                    "action_owner_email":  case.get("action_owner_email") or "",
                    "program_officer_email": case.get("program_officer_email") or "",
                    "cc_emails":           case.get("cc_emails") or None,
                    "suppress_until":      case.get("suppress_until") or None,
                    "_changed_by":         user_name,
                    "_force_advance":      True,
                    "_suppress_notifications": True,  # Bulk advance: suppress SC emails to avoid flooding
                }
                try:
                    upsert_case(upsert_data)
                    advanced += 1
                except Exception as e:
                    errors.append(f"{case['application_id']}: {e}")
            conn.close()
            log_audit("bulk_advance", None,
                      f"Advanced {advanced} cases to '{target_stage}'", user_name, bid)
            err_count = len(errors)
            flash(f"{advanced} case(s) advanced to '{target_stage}'. {err_count} error(s).", "success")
            return redirect(url_for("bulk_advance"))

    conn.close()

    # Build case checkboxes grouped by programme
    prog_groups = {}
    for c in cases:
        pn = c["programme_name"]
        if pn not in prog_groups:
            prog_groups[pn] = []
        prog_groups[pn].append(c)

    case_list_html = ""
    for pn, pcases in prog_groups.items():
        rows = ""
        for c in pcases:
            rows += f"""<tr>
  <td><input type="checkbox" name="case_ids" value="{c['id']}" class="form-check-input case-cb"></td>
  <td class="id-cell">{h(c['application_id'])}</td>
  <td>{h(c['organisation_name'])}</td>
  <td>{h(c['current_stage'])}</td>
  <td style="font-size:12px;color:#94a3b8">{h(c['stage_start_date'])}</td>
</tr>"""
        case_list_html += f"""
<div class="mb-3">
  <div style="font-weight:600;font-size:13px;color:var(--navy);margin-bottom:6px">
    <i class="bi bi-folder2" style="color:var(--accent)"></i> {pn}
    <span style="font-size:11px;color:#94a3b8;font-weight:400;margin-left:6px">{len(pcases)} cases</span>
  </div>
  <table class="data-table">
    <thead><tr><th style="width:30px"><input type="checkbox" class="form-check-input" onclick="toggleGroup(this)"></th><th>App ID</th><th>Organisation</th><th>Current Stage</th><th>Start Date</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
</div>"""

    content = f"""
<form method="post">
<div class="row g-4">
  <div class="col-lg-8">
    <div class="card">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span><i class="bi bi-fast-forward-fill" style="color:#d97706"></i> Select Cases to Advance</span>
        <span style="font-size:12px;color:#94a3b8" id="selCount">0 selected</span>
      </div>
      <div class="card-body p-3" style="max-height:600px;overflow-y:auto">
        {case_list_html if case_list_html else '<div style="text-align:center;color:#94a3b8;padding:40px">No cases found.</div>'}
      </div>
    </div>
  </div>
  <div class="col-lg-4">
    <div class="card" style="position:sticky;top:72px">
      <div class="card-header">Advance To</div>
      <div class="card-body p-3">
        <div class="mb-3">
          <label class="form-label" style="font-size:12px">Programme (for stage lookup)</label>
          <select class="form-select form-select-sm" id="baProgSelect">
            <option value="">— select programme —</option>
            {''.join('<option value="' + h(pn) + '">' + h(pn) + '</option>' for pn in sorted(prog_groups.keys()))}
          </select>
        </div>
        <div class="mb-3">
          <label class="form-label" style="font-size:12px">Target Stage</label>
          <select class="form-select" name="target_stage" id="baStageSelect" required>
            <option value="">— select programme first —</option>
          </select>
        </div>
        <div class="mb-3">
          <label class="form-label" style="font-size:12px">New Start Date</label>
          <input type="date" class="form-control" name="new_start_date"
                 value="{now_ist().strftime('%Y-%m-%d')}">
        </div>
        <button type="submit" class="btn btn-warning w-100">
          <i class="bi bi-fast-forward-fill"></i> Advance Selected Cases
        </button>
      </div>
    </div>
  </div>
</div>
</form>"""
    scripts = """<script>
function toggleGroup(el){
  var tbody = el.closest('table').querySelector('tbody');
  tbody.querySelectorAll('.case-cb').forEach(function(cb){ cb.checked = el.checked; });
  updateCount();
}
document.addEventListener('change', function(e){
  if(e.target.classList.contains('case-cb')) updateCount();
});
function updateCount(){
  var n = document.querySelectorAll('.case-cb:checked').length;
  document.getElementById('selCount').textContent = n + ' selected';
}
document.getElementById('baProgSelect').addEventListener('change', function(){
  var prog = this.value;
  var sel = document.getElementById('baStageSelect');
  sel.innerHTML = '<option value="">Loading…</option>';
  if(!prog){ sel.innerHTML='<option value="">— select programme first —</option>'; return; }
  fetch('/api/stages?programme='+encodeURIComponent(prog))
    .then(function(r){return r.json();})
    .then(function(data){
      sel.innerHTML = '<option value="">— select stage —</option>';
      (data.stages||[]).forEach(function(s){
        var o = document.createElement('option');
        o.value = s; o.textContent = s;
        sel.appendChild(o);
      });
    })
    .catch(function(){ sel.innerHTML='<option value="">Error loading stages</option>'; });
});
</script>"""
    return render_page(content, scripts, active_page="bulk_advance",
                       page_title="Bulk Stage Advance")


# ── Quick Stage Advance (POST from dashboard modal) ──────────────────────────
@app.route("/quick-advance", methods=["POST"])
@login_required
def quick_advance_post():
    case_id = request.form.get("case_id")
    target_stage = request.form.get("target_stage", "").strip()
    new_start_date = request.form.get("new_start_date", date.today().strftime("%Y-%m-%d"))
    suppress_until = request.form.get("suppress_until", "").strip() or None

    conn = get_db()
    case = conn.execute("SELECT * FROM case_tracking WHERE id=?", (case_id,)).fetchone()
    if not case:
        conn.close()
        flash("Case not found.", "error")
        return redirect(url_for("dashboard"))
    case = dict(case)
    cfg = conn.execute(
        "SELECT * FROM programme_config WHERE programme_name=? AND stage_name=?",
        (case["programme_name"], target_stage)
    ).fetchone()
    if not cfg:
        conn.close()
        flash(f"Stage '{target_stage}' not found in this programme.", "error")
        return redirect(url_for("dashboard"))

    old_stage = case["current_stage"]
    conn.execute(
        """UPDATE case_tracking SET current_stage=?, stage_start_date=?,
           tat_days=?, reminder1_day=?, reminder2_day=?, owner_type=?, is_milestone=?,
           r1_sent=0, r2_sent=0, overdue_sent=0, overdue_count=0, last_overdue_date=NULL,
           suppress_until=? WHERE id=?""",
        (target_stage, new_start_date, cfg["tat_days"], cfg["reminder1_day"],
         cfg["reminder2_day"], cfg["owner_type"], cfg["is_milestone"],
         suppress_until, case_id)
    )
    conn.commit()
    conn.close()
    user_name = session.get("full_name") or session.get("username", "")
    log_stage_transition(case["application_id"], old_stage, target_stage, user_name, case.get("board_id"))
    log_audit("stage_change", case["application_id"],
              f"Quick advance: {old_stage} → {target_stage}", user_name, case.get("board_id"))
    flash(f"{case['application_id']} advanced to '{target_stage}'.", "success")
    return redirect(url_for("dashboard"))


# ── Global Search ─────────────────────────────────────────────────────────────
@app.route("/search")
@login_required
def search():
    q = request.args.get("q", "").strip()
    bid = user_board_id()
    ph_progs = user_programme_names()
    results = []
    if q:
        conn = get_db()
        pattern = f"%{q}%"
        base = "SELECT * FROM case_tracking WHERE (application_id LIKE ? OR organisation_name LIKE ?)"
        params = [pattern, pattern]
        if bid is not None:
            base += " AND board_id=?"
            params.append(bid)
        if ph_progs is not None:
            if ph_progs:
                placeholders = ",".join("?" * len(ph_progs))
                base += f" AND programme_name IN ({placeholders})"
                params.extend(ph_progs)
            else:
                base += " AND 1=0"
        base += " LIMIT 50"
        results = [dict(r) for r in conn.execute(base, params).fetchall()]
        try:
            _search_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                               for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
        except Exception:
            _search_db_hols = set()
        conn.close()
    else:
        _search_db_hols = set()

    today = now_ist().date()
    for c in results:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, extra_holidays=_search_db_hols)

    rows = ""
    for c in results:
        tat = c["tat_days"]
        elapsed = c["days_elapsed"]
        if tat > 0 and elapsed >= tat:
            badge = f'<span class="pill pill-danger">Overdue·{elapsed}d</span>'
        elif tat > 0 and elapsed >= c.get("reminder2_day", 0):
            badge = f'<span class="pill pill-warn">At Risk·{elapsed}d</span>'
        else:
            badge = f'<span class="pill pill-ok">On Track·{elapsed}d</span>'
        rows += f"""<tr>
  <td class="id-cell">{h(c['application_id'])}</td>
  <td>{h(c['organisation_name'])}</td>
  <td style="font-size:12px;color:#64748b">{h(c['programme_name'])}</td>
  <td>{h(c['current_stage'])}</td>
  <td>{badge}</td>
  <td><a href="/edit-case/{c['id']}" class="btn btn-sm btn-action btn-outline-primary me-1">Edit</a>
      <a href="/case-history/{h(c['application_id'])}" class="btn btn-sm btn-action btn-outline-secondary">History</a></td>
</tr>"""

    content = f"""
<div class="card">
  <div class="card-header">
    <form method="get" class="d-flex gap-2 align-items-center">
      <input type="text" class="form-control" name="q" value="{h(q)}" placeholder="Search by Application ID or Organisation…" style="max-width:400px">
      <button class="btn btn-primary" type="submit"><i class="bi bi-search"></i> Search</button>
    </form>
  </div>
  <div style="overflow-x:auto">
    <table class="data-table">
      <thead><tr><th>App ID</th><th>Organisation</th><th>Programme</th><th>Stage</th><th>Status</th><th>Actions</th></tr></thead>
      <tbody>
        {rows if rows else ('<tr><td colspan="6" style="text-align:center;color:#94a3b8;padding:40px">' +
          ('No results for &quot;' + str(h(q)) + '&quot;' if q else 'Enter a search term above.') + '</td></tr>')}
      </tbody>
    </table>
  </div>
</div>"""
    return render_page(content, active_page="search", page_title="Search Cases",
                       page_crumb=f"Search: {h(q)}" if q else "Search")


# ── Analytics / Reports Hub ───────────────────────────────────────────────────
@app.route("/reports")
@login_required
def reports():
    conn = get_db()
    bid = user_board_id()
    ph_progs = user_programme_names()
    today = now_ist().date()

    if bid is not None:
        if ph_progs is not None:
            # program_head: filter by mapped programmes within board
            if ph_progs:
                placeholders = ",".join("?" * len(ph_progs))
                cases = [dict(r) for r in conn.execute(
                    f"SELECT * FROM case_tracking WHERE board_id=? AND programme_name IN ({placeholders})",
                    [bid] + ph_progs
                ).fetchall()]
                history = [dict(r) for r in conn.execute(
                    "SELECT * FROM stage_history WHERE board_id=? ORDER BY timestamp DESC", (bid,)
                ).fetchall()]
            else:
                cases, history = [], []
        else:
            cases = [dict(r) for r in conn.execute(
                "SELECT * FROM case_tracking WHERE board_id=?", (bid,)
            ).fetchall()]
            history = [dict(r) for r in conn.execute(
                "SELECT * FROM stage_history WHERE board_id=? ORDER BY timestamp DESC", (bid,)
            ).fetchall()]
    else:
        cases = [dict(r) for r in conn.execute("SELECT * FROM case_tracking").fetchall()]
        history = [dict(r) for r in conn.execute(
            "SELECT * FROM stage_history ORDER BY timestamp DESC"
        ).fetchall()]
    try:
        _rep_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                        for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _rep_db_hols = set()
    conn.close()

    for c in cases:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, extra_holidays=_rep_db_hols)

    # ── TAT Breach Report ──
    breached = [c for c in cases if not c["is_milestone"] and c["tat_days"] > 0
                and c["days_elapsed"] >= c["tat_days"]]
    breach_rows = ""
    for c in sorted(breached, key=lambda x: x["days_elapsed"] - x["tat_days"], reverse=True):
        breach_days = c["days_elapsed"] - c["tat_days"]
        breach_rows += f"""<tr>
  <td class="id-cell">{c['application_id']}</td>
  <td>{c['organisation_name']}</td>
  <td style="font-size:12px">{c['programme_name']}</td>
  <td>{c['current_stage']}</td>
  <td style="text-align:center">{c['tat_days']}</td>
  <td style="text-align:center">{c['days_elapsed']}</td>
  <td style="text-align:center;font-weight:700;color:#dc2626">+{breach_days}d</td>
  <td style="font-size:12px;color:#94a3b8">{c['stage_start_date']}</td>
</tr>"""

    # ── Stage Funnel ──
    stage_counts = {}
    for c in cases:
        if not c["is_milestone"]:
            stage_counts[c["current_stage"]] = stage_counts.get(c["current_stage"], 0) + 1
    max_cnt = max(stage_counts.values()) if stage_counts else 1
    funnel_html = ""
    for stage, cnt in sorted(stage_counts.items(), key=lambda x: -x[1])[:20]:
        pct = int(cnt / max_cnt * 100)
        funnel_html += f"""
<div class="mb-2">
  <div class="d-flex justify-content-between" style="font-size:12px;margin-bottom:3px">
    <span style="font-weight:500;color:#1e293b">{stage}</span>
    <span style="color:#64748b">{cnt}</span>
  </div>
  <div class="tat-bar" style="width:100%;height:8px">
    <div class="tat-bar-fill" style="width:{pct}%;background:var(--accent)"></div>
  </div>
</div>"""

    # ── Programme Throughput (transitions per programme this month) ──
    this_month = today.strftime("%Y-%m")
    throughput = {}
    for h in history:
        if h["timestamp"][:7] == this_month:
            key = h.get("application_id", "")
            # Get programme from case
            for c in cases:
                if c["application_id"] == key:
                    prog = c["programme_name"]
                    throughput[prog] = throughput.get(prog, 0) + 1
                    break

    tp_html = ""
    if throughput:
        max_tp = max(throughput.values())
        for prog, cnt in sorted(throughput.items(), key=lambda x: -x[1]):
            pct = int(cnt / max_tp * 100)
            tp_html += f"""
<div class="mb-2">
  <div class="d-flex justify-content-between" style="font-size:12px;margin-bottom:3px">
    <span style="font-weight:500">{prog}</span>
    <span style="color:#64748b">{cnt} transitions</span>
  </div>
  <div class="tat-bar" style="width:100%;height:8px">
    <div class="tat-bar-fill" style="width:{pct}%;background:#7c3aed"></div>
  </div>
</div>"""
    else:
        tp_html = '<div style="color:#94a3b8;font-size:13px">No transitions recorded this month.</div>'

    # ── Average TAT per Stage (from stage_history) ──
    stage_durations = {}
    # For each transition in history, look up the previous transition to compute dwell time
    app_timelines = {}
    for h in sorted(history, key=lambda x: x["timestamp"]):
        app = h["application_id"]
        if app not in app_timelines:
            app_timelines[app] = []
        app_timelines[app].append(h)

    for app, transitions in app_timelines.items():
        for i in range(1, len(transitions)):
            prev = transitions[i-1]
            curr = transitions[i]
            try:
                d1 = datetime.strptime(prev["timestamp"], "%Y-%m-%d %H:%M:%S").date()
                d2 = datetime.strptime(curr["timestamp"], "%Y-%m-%d %H:%M:%S").date()
                days = working_days_elapsed(d1.isoformat(), d2)
                stage = prev["to_stage"]
                if stage not in stage_durations:
                    stage_durations[stage] = []
                stage_durations[stage].append(days)
            except Exception:
                pass

    avg_tat_rows = ""
    if stage_durations:
        for stage, durations in sorted(stage_durations.items(), key=lambda x: -sum(x[1])/len(x[1])):
            avg = sum(durations) / len(durations)
            avg_tat_rows += f"""<tr>
  <td style="font-weight:500">{stage}</td>
  <td style="text-align:center">{len(durations)}</td>
  <td style="text-align:center;font-weight:600;color:var(--accent)">{avg:.1f}d</td>
  <td style="text-align:center;color:#94a3b8">{min(durations)}d</td>
  <td style="text-align:center;color:#dc2626">{max(durations)}d</td>
</tr>"""
    else:
        avg_tat_rows = '<tr><td colspan="5" style="text-align:center;color:#94a3b8;padding:20px">Not enough history yet. TAT averages will populate as cases move through stages.</td></tr>'

    content = f"""
<div class="row g-4 mb-4">
  <div class="col-lg-6">
    <div class="card h-100">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span><i class="bi bi-exclamation-triangle-fill" style="color:#dc2626"></i> TAT Breach Report</span>
        <span class="pill pill-danger">{len(breached)} cases</span>
      </div>
      <div style="overflow-x:auto;max-height:360px">
        <table class="data-table">
          <thead><tr><th>App ID</th><th>Org</th><th>Programme</th><th>Stage</th>
            <th style="text-align:center">TAT</th><th style="text-align:center">Elapsed</th>
            <th style="text-align:center">Breach</th><th>Since</th></tr></thead>
          <tbody>{breach_rows if breach_rows else '<tr><td colspan="8" style="text-align:center;color:#94a3b8;padding:30px">No TAT breaches</td></tr>'}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="col-lg-6">
    <div class="card h-100">
      <div class="card-header"><i class="bi bi-funnel-fill" style="color:var(--accent)"></i> Stage Pipeline (Where Cases Are Now)</div>
      <div class="card-body" style="max-height:360px;overflow-y:auto">
        {funnel_html if funnel_html else '<div style="color:#94a3b8;font-size:13px">No cases.</div>'}
      </div>
    </div>
  </div>
</div>

<div class="row g-4">
  <div class="col-lg-5">
    <div class="card">
      <div class="card-header"><i class="bi bi-bar-chart-steps" style="color:#7c3aed"></i> Programme Throughput — {this_month}</div>
      <div class="card-body">
        {tp_html}
      </div>
    </div>
  </div>
  <div class="col-lg-7">
    <div class="card">
      <div class="card-header"><i class="bi bi-clock-history" style="color:#d97706"></i> Average Working Days per Stage</div>
      <div style="overflow-x:auto">
        <table class="data-table">
          <thead><tr><th>Stage</th><th style="text-align:center">Cases</th>
            <th style="text-align:center">Avg Days</th>
            <th style="text-align:center">Min</th><th style="text-align:center">Max</th></tr></thead>
          <tbody>{avg_tat_rows}</tbody>
        </table>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="reports", page_title="Analytics & Reports")


@app.route("/assessor-scorecard")
@board_admin_required
def assessor_scorecard():
    conn = get_db()
    bid = user_board_id()
    q = """SELECT action_owner_email, action_owner_name,
               COUNT(*) AS total_cases,
               SUM(r1_sent) AS r1_count,
               SUM(r2_sent) AS r2_count,
               SUM(overdue_sent) AS overdue_count,
               ROUND(100.0 * SUM(overdue_sent) / NULLIF(COUNT(*),0), 1) AS overdue_rate
               FROM case_tracking WHERE owner_type='Assessor'"""
    if bid is not None:
        rows = conn.execute(q + " AND board_id=? GROUP BY action_owner_email, action_owner_name ORDER BY overdue_rate DESC NULLS LAST", (bid,)).fetchall()
    else:
        rows = conn.execute(q + " GROUP BY action_owner_email, action_owner_name ORDER BY overdue_rate DESC NULLS LAST").fetchall()
    conn.close()
    rows = [dict(r) for r in rows]

    trs = ""
    for r in rows:
        rate = float(r["overdue_rate"] or 0)
        badge_color = "#dc2626" if rate >= 50 else "#d97706" if rate >= 25 else "#00984C"
        trs += (f'<tr><td style="font-weight:500">{h(r["action_owner_name"] or "—")}</td>'
                f'<td style="font-size:12px;color:#64748b">{h(r["action_owner_email"])}</td>'
                f'<td style="text-align:center">{r["total_cases"]}</td>'
                f'<td style="text-align:center;color:#d97706">{r["r1_count"]}</td>'
                f'<td style="text-align:center;color:#dc2626">{r["r2_count"]}</td>'
                f'<td style="text-align:center;color:#7c3aed">{r["overdue_count"]}</td>'
                f'<td style="text-align:center;font-weight:700;color:{badge_color}">{rate}%</td></tr>')
    if not trs:
        trs = '<tr><td colspan="7" style="text-align:center;color:#94a3b8;padding:32px">No assessor data yet. Cases appear here once owner_type=Assessor cases are tracked.</td></tr>'

    content = f"""
<div style="max-width:1100px;margin:0 auto">
  <div class="card">
    <div class="card-header"><i class="bi bi-person-badge" style="color:var(--accent)"></i> Assessor Performance Scorecard</div>
    <div class="card-body p-0">
      <table class="table table-hover mb-0" style="font-size:13px">
        <thead><tr style="background:#f8fafc">
          <th style="padding:10px 16px">Name</th><th>Email</th>
          <th style="text-align:center">Cases</th><th style="text-align:center">R1 Sent</th>
          <th style="text-align:center">R2 Sent</th><th style="text-align:center">Overdue</th>
          <th style="text-align:center">Overdue %</th>
        </tr></thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="assessor_scorecard", page_title="Assessor Scorecard")


def _export_csv_filtered():
    """CSV export helper called from export_excel when fmt=csv."""
    conn = get_db()
    bid = user_board_id()
    ph_progs = user_programme_names()
    today = now_ist().date()
    prog_filter = request.form.getlist("prog_filter")
    case_status_f = request.form.get("case_status", "")
    date_from = request.form.get("date_from", "")
    date_to = request.form.get("date_to", "")
    q = "SELECT * FROM case_tracking WHERE 1=1"
    params = []
    if bid is not None:
        q += " AND board_id=?"; params.append(bid)
    if case_status_f:
        q += " AND (status=? OR (status IS NULL AND ?='Active'))"; params += [case_status_f, case_status_f]
    if date_from:
        q += " AND stage_start_date >= ?"; params.append(date_from)
    if date_to:
        q += " AND stage_start_date <= ?"; params.append(date_to)
    cases = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    if prog_filter:
        cases = [c for c in cases if c["programme_name"] in prog_filter]
    if ph_progs is not None:
        cases = [c for c in cases if c["programme_name"] in ph_progs]
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Application ID","Organisation","Programme","Stage","Owner Type",
                "Action Owner","Email","Stage Start","TAT Days","Days Elapsed",
                "SLA Status","Case Status","R1 Sent","R2 Sent","Overdue Sent","Follow-ups"])
    for c in cases:
        elapsed = working_days_elapsed(c["stage_start_date"], today)
        tat = c["tat_days"]
        if c["is_milestone"]: sla = "Milestone"
        elif tat > 0 and elapsed >= tat: sla = "Overdue"
        elif tat > 0 and elapsed >= c.get("reminder2_day", 0): sla = "At Risk"
        else: sla = "On Track"
        w.writerow([c["application_id"], c["organisation_name"], c["programme_name"],
                    c["current_stage"], c["owner_type"] or "", c["action_owner_name"] or "",
                    c["action_owner_email"] or "", c["stage_start_date"], tat, elapsed,
                    sla, c.get("status","Active"),
                    "Yes" if c["r1_sent"] else "No", "Yes" if c["r2_sent"] else "No",
                    "Yes" if c["overdue_sent"] else "No", c["overdue_count"]])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=qci_report_{today}.csv"})


# ── Export Report Config Page ─────────────────────────────────────────────────
@app.route("/export-excel", methods=["GET"])
@login_required
def export_excel_page():
    """Show export configuration form."""
    conn = get_db()
    bid = user_board_id()
    ph_progs = user_programme_names()
    if bid is not None:
        programmes = [r[0] for r in conn.execute(
            "SELECT DISTINCT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
        stages = [r[0] for r in conn.execute(
            "SELECT DISTINCT stage_name FROM programme_config WHERE board_id=? ORDER BY stage_name", (bid,)
        ).fetchall()]
    else:
        programmes = [r[0] for r in conn.execute(
            "SELECT DISTINCT programme_name FROM programmes ORDER BY programme_name"
        ).fetchall()]
        stages = [r[0] for r in conn.execute(
            "SELECT DISTINCT stage_name FROM programme_config ORDER BY stage_name"
        ).fetchall()]
    if ph_progs is not None:
        programmes = [p for p in programmes if p in ph_progs]
    conn.close()

    prog_checkboxes = "".join(
        f'<div class="form-check"><input class="form-check-input" type="checkbox" name="prog_filter" value="{p}" id="p_{i}" checked>'
        f'<label class="form-check-label" for="p_{i}" style="font-size:12px">{p}</label></div>'
        for i, p in enumerate(programmes)
    )
    status_opts = "".join(
        f'<div class="form-check"><input class="form-check-input" type="checkbox" name="status_filter" value="{s}" checked>'
        f'<label class="form-check-label" style="font-size:12px">{s}</label></div>'
        for s in ["Active", "Closed", "Withdrawn", "Suspended", "All"]
    )

    col_opts = ""
    for col_id, col_label in [
        ("app_id","Application ID"),("org","Organisation"),("programme","Programme"),
        ("stage","Current Stage"),("owner_type","Owner Type"),("owner_name","Action Owner"),
        ("owner_email","Owner Email"),("po_email","PO Email"),("stage_start","Stage Start"),
        ("tat","TAT Days"),("elapsed","Days Elapsed"),("sla_status","SLA Status"),
        ("r1","R1 Sent"),("r2","R2 Sent"),("overdue","Overdue Sent"),("followups","Follow-ups"),
        ("case_status","Case Status"),
    ]:
        col_opts += (f'<div class="form-check form-check-inline">'
                     f'<input class="form-check-input" type="checkbox" name="cols" value="{col_id}" id="c_{col_id}" checked>'
                     f'<label class="form-check-label" for="c_{col_id}" style="font-size:12px">{col_label}</label></div>')

    content = f"""
<div class="row g-4">
  <div class="col-lg-8">
    <form method="post" action="/export-excel/download">
      <div class="card mb-4">
        <div class="card-header"><i class="bi bi-funnel" style="color:var(--accent)"></i> Filter Data</div>
        <div class="card-body p-4">
          <div class="row g-3">
            <div class="col-md-6">
              <label class="form-label">Date Range (Stage Start)</label>
              <div class="d-flex gap-2">
                <input type="date" class="form-control form-control-sm" name="date_from" placeholder="From">
                <input type="date" class="form-control form-control-sm" name="date_to" placeholder="To">
              </div>
            </div>
            <div class="col-md-3">
              <label class="form-label">Case Status</label>
              <select class="form-select form-select-sm" name="case_status">
                <option value="">All</option>
                <option>Active</option><option>Closed</option>
                <option>Withdrawn</option><option>Suspended</option>
              </select>
            </div>
            <div class="col-md-3">
              <label class="form-label">SLA Status</label>
              <select class="form-select form-select-sm" name="sla_status">
                <option value="">All</option>
                <option>On Track</option><option>At Risk</option>
                <option>Overdue</option><option>Milestone</option>
              </select>
            </div>
          </div>
          <div class="mt-3">
            <label class="form-label">Programmes
              <button type="button" class="btn btn-xs ms-2" style="font-size:10px;padding:1px 6px;border:1px solid #cbd5e1"
                      onclick="document.querySelectorAll('[name=prog_filter]').forEach(c=>c.checked=true)">All</button>
              <button type="button" class="btn btn-xs ms-1" style="font-size:10px;padding:1px 6px;border:1px solid #cbd5e1"
                      onclick="document.querySelectorAll('[name=prog_filter]').forEach(c=>c.checked=false)">None</button>
            </label>
            <div class="row g-1">{prog_checkboxes or '<span style="color:#94a3b8;font-size:12px">No programmes available.</span>'}</div>
          </div>
        </div>
      </div>

      <div class="card mb-4">
        <div class="card-header"><i class="bi bi-table" style="color:#7c3aed"></i> Columns to Include</div>
        <div class="card-body p-4">
          <button type="button" class="btn btn-xs mb-2" style="font-size:11px;padding:2px 8px;border:1px solid #cbd5e1"
                  onclick="document.querySelectorAll('[name=cols]').forEach(c=>c.checked=true)">Select All</button>
          <button type="button" class="btn btn-xs mb-2 ms-1" style="font-size:11px;padding:2px 8px;border:1px solid #cbd5e1"
                  onclick="document.querySelectorAll('[name=cols]').forEach(c=>c.checked=false)">Deselect All</button>
          <div>{col_opts}</div>
        </div>
      </div>

      <div class="card mb-4">
        <div class="card-header"><i class="bi bi-file-earmark-excel" style="color:#059669"></i> Export Format</div>
        <div class="card-body p-4">
          <div class="d-flex gap-3">
            <div class="form-check">
              <input class="form-check-input" type="radio" name="fmt" value="xlsx" id="fmtXlsx" checked>
              <label class="form-check-label" for="fmtXlsx">Excel (.xlsx) — multi-sheet with formatting</label>
            </div>
            <div class="form-check">
              <input class="form-check-input" type="radio" name="fmt" value="csv" id="fmtCsv">
              <label class="form-check-label" for="fmtCsv">CSV — simple flat file</label>
            </div>
          </div>
        </div>
      </div>

      <button type="submit" class="btn btn-primary px-5">
        <i class="bi bi-download"></i> Generate &amp; Download Report
      </button>
    </form>
  </div>

  <div class="col-lg-4">
    <div class="card" style="position:sticky;top:72px">
      <div class="card-header"><i class="bi bi-info-circle" style="color:#0891b2"></i> About This Export</div>
      <div class="card-body p-4" style="font-size:13px;color:#475569">
        <p>Select the filters and columns you need, then click <strong>Generate &amp; Download</strong>.</p>
        <ul style="font-size:12px">
          <li><strong>Excel (.xlsx)</strong> — includes Active Cases, Stage History, and Audit Log sheets with colour-coded headers.</li>
          <li><strong>CSV</strong> — flat file, easy to open in any spreadsheet tool.</li>
          <li>Data is scoped to your board/programmes.</li>
        </ul>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="export_excel", page_title="Export Report")


# ── Multi-sheet Excel Export (download) ───────────────────────────────────────
@app.route("/export-excel/download", methods=["POST"])
@login_required
def export_excel():
    """Handles parameterised export — filters applied from POST form."""
    fmt = request.form.get("fmt", "xlsx")
    if fmt == "csv":
        return _export_csv_filtered()

    if not HAS_XLSX:
        flash("openpyxl not installed.", "error")
        return redirect(url_for("export_excel_page"))
    conn = get_db()
    bid = user_board_id()
    ph_progs = user_programme_names()
    today = now_ist().date()

    # Build filter conditions from POST params
    prog_filter = request.form.getlist("prog_filter")
    case_status_f = request.form.get("case_status", "")
    sla_status_f = request.form.get("sla_status", "")
    date_from = request.form.get("date_from", "")
    date_to = request.form.get("date_to", "")
    selected_cols = set(request.form.getlist("cols")) or {
        "app_id","org","programme","stage","owner_type","owner_name",
        "stage_start","tat","elapsed","sla_status","case_status"
    }

    q = "SELECT * FROM case_tracking WHERE 1=1"
    params = []
    if bid is not None:
        q += " AND board_id=?"; params.append(bid)
    if case_status_f:
        q += " AND (status=? OR (status IS NULL AND ?='Active'))"; params += [case_status_f, case_status_f]
    if date_from:
        q += " AND stage_start_date >= ?"; params.append(date_from)
    if date_to:
        q += " AND stage_start_date <= ?"; params.append(date_to)

    cases = [dict(r) for r in conn.execute(q, params).fetchall()]

    # Filter by programme
    if prog_filter:
        cases = [c for c in cases if c["programme_name"] in prog_filter]
    if ph_progs is not None:
        cases = [c for c in cases if c["programme_name"] in ph_progs]

    if bid is not None:
        history = [dict(r) for r in conn.execute(
            "SELECT * FROM stage_history WHERE board_id=? ORDER BY timestamp DESC LIMIT 2000", (bid,)
        ).fetchall()]
        audit = [dict(r) for r in conn.execute(
            "SELECT * FROM audit_log WHERE board_id=? ORDER BY id DESC LIMIT 2000", (bid,)
        ).fetchall()]
    else:
        history = [dict(r) for r in conn.execute(
            "SELECT * FROM stage_history ORDER BY timestamp DESC LIMIT 2000"
        ).fetchall()]
        audit = [dict(r) for r in conn.execute(
            "SELECT * FROM audit_log ORDER BY id DESC LIMIT 2000"
        ).fetchall()]
    try:
        _excl_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                         for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _excl_db_hols = set()
    conn.close()

    for c in cases:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, extra_holidays=_excl_db_hols)
        tat = c["tat_days"]
        elapsed = c["days_elapsed"]
        if c["is_milestone"]:
            c["status"] = "Milestone"
        elif tat > 0 and elapsed >= tat:
            c["status"] = "Overdue"
        elif tat > 0 and elapsed >= c.get("reminder2_day", 0):
            c["status"] = "At Risk"
        else:
            c["status"] = "On Track"

    # Apply SLA status filter
    if sla_status_f:
        cases = [c for c in cases if c.get("status") == sla_status_f]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Active Cases ──
    ws1 = wb.active
    ws1.title = "Active Cases"
    hdr_fill = PatternFill("solid", fgColor="003356")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers1 = ["Application ID", "Organisation", "Programme", "Stage", "Status",
                "Days Elapsed", "TAT Days", "Start Date", "Owner Type",
                "Action Owner", "Action Email", "PO Email",
                "R1 Sent", "R2 Sent", "Overdue Sent", "Follow-ups",
                "CC Emails", "Suppress Until"]
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for ri, c in enumerate(cases, 2):
        ws1.cell(ri, 1, c["application_id"])
        ws1.cell(ri, 2, c["organisation_name"])
        ws1.cell(ri, 3, c["programme_name"])
        ws1.cell(ri, 4, c["current_stage"])
        ws1.cell(ri, 5, c["status"])
        ws1.cell(ri, 6, c["days_elapsed"])
        ws1.cell(ri, 7, c["tat_days"])
        ws1.cell(ri, 8, c["stage_start_date"])
        ws1.cell(ri, 9, c.get("owner_type") or "")
        ws1.cell(ri, 10, c.get("action_owner_name") or "")
        ws1.cell(ri, 11, c.get("action_owner_email") or "")
        ws1.cell(ri, 12, c.get("program_officer_email") or "")
        ws1.cell(ri, 13, "Yes" if c.get("r1_sent") else "No")
        ws1.cell(ri, 14, "Yes" if c.get("r2_sent") else "No")
        ws1.cell(ri, 15, "Yes" if c.get("overdue_sent") else "No")
        ws1.cell(ri, 16, c.get("overdue_count") or 0)
        ws1.cell(ri, 17, c.get("cc_emails") or "")
        ws1.cell(ri, 18, c.get("suppress_until") or "")
        # Colour overdue rows
        if c["status"] == "Overdue":
            red_fill = PatternFill("solid", fgColor="FEE2E2")
            for ci2 in range(1, len(headers1)+1):
                ws1.cell(ri, ci2).fill = red_fill

    # ── Sheet 2: Stage Transition History ──
    ws2 = wb.create_sheet("Stage History")
    for ci, h in enumerate(["Timestamp", "Application ID", "From Stage", "To Stage", "Changed By"], 1):
        cell = ws2.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for ri, h in enumerate(history, 2):
        ws2.cell(ri, 1, h["timestamp"])
        ws2.cell(ri, 2, h["application_id"])
        ws2.cell(ri, 3, h.get("from_stage") or "")
        ws2.cell(ri, 4, h["to_stage"])
        ws2.cell(ri, 5, h.get("changed_by") or "")

    # ── Sheet 3: Audit Log ──
    ws3 = wb.create_sheet("Audit Log")
    for ci, h in enumerate(["Timestamp", "Event", "Application ID", "Detail", "User"], 1):
        cell = ws3.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for ri, a in enumerate(audit, 2):
        ws3.cell(ri, 1, a["timestamp"])
        ws3.cell(ri, 2, a["event_type"])
        ws3.cell(ri, 3, a.get("application_id") or "")
        ws3.cell(ri, 4, a.get("detail") or "")
        ws3.cell(ri, 5, a.get("user_name") or "")

    # ── Sheet 4: TAT Summary ──
    ws4 = wb.create_sheet("TAT Summary")
    for ci, h in enumerate(["Programme", "Total Cases", "Overdue", "At Risk", "On Track", "SLA %"], 1):
        cell = ws4.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    prog_summary = {}
    for c in cases:
        pn = c["programme_name"]
        if pn not in prog_summary:
            prog_summary[pn] = {"total": 0, "overdue": 0, "at_risk": 0, "on_track": 0}
        prog_summary[pn]["total"] += 1
        s = c["status"]
        if s == "Overdue": prog_summary[pn]["overdue"] += 1
        elif s == "At Risk": prog_summary[pn]["at_risk"] += 1
        elif s == "On Track": prog_summary[pn]["on_track"] += 1
    for ri, (pn, d) in enumerate(prog_summary.items(), 2):
        sla = int((d["on_track"] + d["at_risk"]) / d["total"] * 100) if d["total"] else 100
        ws4.cell(ri, 1, pn); ws4.cell(ri, 2, d["total"])
        ws4.cell(ri, 3, d["overdue"]); ws4.cell(ri, 4, d["at_risk"])
        ws4.cell(ri, 5, d["on_track"]); ws4.cell(ri, 6, f"{sla}%")

    # Auto-width
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_w = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"qci_export_{today.strftime('%Y%m%d')}.xlsx"
    return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})


# ── Email Queue Monitor ───────────────────────────────────────────────────────
@app.route("/email-queue")
@board_admin_required
def email_queue_page():
    conn = get_db()
    bid = user_board_id()
    if bid is not None:
        items = [dict(r) for r in conn.execute(
            "SELECT * FROM email_queue WHERE board_id=? OR board_id IS NULL ORDER BY id DESC LIMIT 200",
            (bid,)
        ).fetchall()]
    else:
        items = [dict(r) for r in conn.execute(
            "SELECT * FROM email_queue ORDER BY id DESC LIMIT 200"
        ).fetchall()]
    conn.close()

    STATUS_COLORS = {"sent": "#00984C", "failed": "#dc2626", "pending": "#d97706"}
    rows = ""
    for item in items:
        clr = STATUS_COLORS.get(item["status"], "#64748b")
        rows += f"""<tr>
  <td style="font-size:11px;color:#64748b">{item['queued_at']}</td>
  <td><span class="pill" style="background:{clr}22;color:{clr}">{item['status']}</span></td>
  <td style="font-size:12px">{item.get('application_id') or '—'}</td>
  <td style="font-size:12px">{item.get('notification_type', '')}</td>
  <td style="font-size:12px">{item.get('to_email', '')}</td>
  <td style="font-size:11px;color:#64748b">{item.get('subject', '')[:60]}</td>
  <td style="text-align:center">{item.get('attempts', 0)}</td>
  <td style="font-size:11px;color:#dc2626">{item.get('error_msg', '') or ''}</td>
</tr>"""

    content = f"""
<div class="card">
  <div class="card-header d-flex justify-content-between align-items-center">
    <span><i class="bi bi-send-check" style="color:var(--accent)"></i> Email Queue</span>
    <form method="post" action="/retry-queue">
      <button class="btn btn-sm btn-outline-primary" type="submit">
        <i class="bi bi-arrow-repeat"></i> Retry Failed
      </button>
    </form>
  </div>
  <div style="overflow-x:auto">
    <table class="data-table">
      <thead><tr><th>Queued At</th><th>Status</th><th>App ID</th><th>Type</th>
        <th>To</th><th>Subject</th><th style="text-align:center">Attempts</th><th>Error</th></tr></thead>
      <tbody>{rows if rows else '<tr><td colspan="8" style="text-align:center;color:#94a3b8;padding:40px">Queue is empty.</td></tr>'}</tbody>
    </table>
  </div>
</div>"""
    return render_page(content, active_page="queue", page_title="Email Queue")


@app.route("/retry-queue", methods=["POST"])
@board_admin_required
def retry_queue():
    """Reset failed emails to pending so next scheduler run retries them."""
    conn = get_db()
    conn.execute("UPDATE email_queue SET status='pending', attempts=0 WHERE status='failed'")
    conn.commit()
    conn.close()
    flash("Failed emails reset to pending. They will be retried on next scheduler run.", "success")
    return redirect(url_for("email_queue_page"))


# ── System Settings (super_admin) ────────────────────────────────────────────
@app.route("/system-settings", methods=["GET", "POST"])
@admin_required
def system_settings():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "save_scheduler":
            hour = int(request.form.get("sched_hour", 8))
            minute = int(request.form.get("sched_minute", 0))
            set_app_setting("scheduler_hour", str(hour))    # key must match startup read
            set_app_setting("scheduler_minute", str(minute))
            # Reschedule live
            try:
                scheduler.reschedule_job("daily_check", trigger="cron",
                                         hour=hour, minute=minute)
            except Exception:
                pass  # scheduler not running on Vercel; setting saved to DB is enough
            flash(f"Scheduler updated to {hour:02d}:{minute:02d} IST. Effective on next restart.", "success")

        elif action == "save_webhook":
            set_app_setting("webhook_url", request.form.get("webhook_url", "").strip())
            flash("Webhook URL saved.", "success")

        elif action == "save_2fa":
            uid = request.form.get("user_id")
            enable = request.form.get("enable_2fa")
            conn = get_db()
            if enable:
                secret = totp_generate_secret()
                conn.execute("UPDATE users SET totp_secret=? WHERE id=?", (secret, uid))
                conn.commit()
                user = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
                conn.close()
                uri = totp_provisioning_uri(secret, user["username"])
                import urllib.parse
                qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={urllib.parse.quote(uri)}"
                flash(f"2FA enabled for {user['username']}. Secret: {secret}", "success")
            else:
                conn.execute("UPDATE users SET totp_secret=NULL WHERE id=?", (uid,))
                conn.commit()
                conn.close()
                flash("2FA disabled.", "success")

        elif action == "force_reset":
            uid = request.form.get("user_id")
            conn = get_db()
            conn.execute("UPDATE users SET force_password_reset=1 WHERE id=?", (uid,))
            conn.commit()
            conn.close()
            flash("User will be required to reset password on next login.", "success")

        elif action == "add_holiday":
            hdate = request.form.get("holiday_date", "").strip()
            hname = request.form.get("holiday_name", "").strip()
            if hdate and hname:
                conn = get_db()
                try:
                    conn.execute("INSERT INTO holidays (holiday_date, name) VALUES (?,?)", (hdate, hname))
                    conn.commit()
                    flash(f"Holiday '{hname}' on {hdate} added.", "success")
                except Exception as e:
                    flash(f"Error: {e}", "error")
                conn.close()

        elif action == "delete_holiday":
            hid = request.form.get("holiday_id")
            conn = get_db()
            conn.execute("DELETE FROM holidays WHERE id=?", (hid,))
            conn.commit()
            conn.close()
            flash("Holiday removed.", "success")

        elif action == "save_digest":
            set_app_setting("digest_enabled", "1" if request.form.get("digest_enabled") else "0")
            set_app_setting("ph_escalation_days", request.form.get("ph_escalation_days", "5"))
            flash("Notification settings saved.", "success")
            try:
                scheduler.reschedule_job("weekly_digest", trigger="cron",
                                         day_of_week="mon", hour=8, minute=0)
            except Exception:
                pass  # scheduler not running on Vercel

    sched_hour   = int(get_app_setting("scheduler_hour", "8"))
    sched_minute = int(get_app_setting("scheduler_minute", "0"))
    webhook_url  = get_app_setting("webhook_url", "")
    digest_enabled = get_app_setting("digest_enabled", "1") == "1"
    ph_escalation_days = get_app_setting("ph_escalation_days", "5")

    conn = get_db()
    users = [dict(r) for r in conn.execute("SELECT id, username, totp_secret, force_password_reset FROM users ORDER BY username").fetchall()]
    holidays_list = [dict(r) for r in conn.execute("SELECT * FROM holidays ORDER BY holiday_date").fetchall()]
    conn.close()

    user_rows = ""
    for u in users:
        has_2fa = bool(u["totp_secret"])
        needs_reset = bool(u["force_password_reset"])
        user_rows += f"""<tr>
  <td style="font-weight:600">{u['username']}</td>
  <td><span class="pill {'pill-ok' if has_2fa else 'pill-muted'}">{('Enabled' if has_2fa else 'Off')}</span></td>
  <td>{'<span class="pill pill-warn">Required</span>' if needs_reset else '<span style="color:#94a3b8;font-size:12px">—</span>'}</td>
  <td style="white-space:nowrap">
    <form method="post" class="d-inline">
      <input type="hidden" name="user_id" value="{u['id']}">
      <input type="hidden" name="action" value="save_2fa">
      <input type="hidden" name="enable_2fa" value="{'0' if has_2fa else '1'}">
      <button class="btn btn-sm btn-action {'btn-outline-danger' if has_2fa else 'btn-outline-success'}" type="submit">
        {'Disable 2FA' if has_2fa else 'Enable 2FA'}</button>
    </form>
    <form method="post" class="d-inline ms-1">
      <input type="hidden" name="user_id" value="{u['id']}">
      <input type="hidden" name="action" value="force_reset">
      <button class="btn btn-sm btn-action btn-outline-warning" type="submit">Force PW Reset</button>
    </form>
  </td>
</tr>"""

    minute_opts = "".join(
        f'<option value="{m:02d}" {"selected" if m == sched_minute else ""}>{m:02d}</option>'
        for m in [0, 15, 30, 45]
    )
    hour_opts = "".join(
        f'<option value="{h}" {"selected" if h == sched_hour else ""}>{h:02d}:00</option>'
        for h in range(24)
    )

    holiday_rows = "".join(
        f'<tr><td style="font-size:13px">{h["holiday_date"]}</td><td style="font-size:13px">{h["name"]}</td>'
        f'<td><form method="post" class="d-inline">'
        f'<input type="hidden" name="action" value="delete_holiday">'
        f'<input type="hidden" name="holiday_id" value="{h["id"]}">'
        f'<button class="btn btn-sm btn-action btn-outline-danger" type="submit"><i class="bi bi-trash"></i></button>'
        f'</form></td></tr>'
        for h in holidays_list
    ) or '<tr><td colspan="3" style="text-align:center;color:#94a3b8;padding:16px;font-size:13px">No custom holidays added (built-in Indian holidays are always applied)</td></tr>'

    content = f"""
<div class="row g-4">
  <div class="col-lg-5">
    <div class="card mb-4">
      <div class="card-header"><i class="bi bi-alarm" style="color:var(--accent)"></i> Scheduler Time (IST)</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="save_scheduler">
          <div class="row g-2 mb-3">
            <div class="col-6">
              <label class="form-label" style="font-size:12px">Hour (24h)</label>
              <select class="form-select" name="sched_hour">{hour_opts}</select>
            </div>
            <div class="col-6">
              <label class="form-label" style="font-size:12px">Minute</label>
              <select class="form-select" name="sched_minute">{minute_opts}</select>
            </div>
          </div>
          <button class="btn btn-primary w-100" type="submit">Update Schedule</button>
        </form>
        <div style="font-size:11px;color:#94a3b8;margin-top:8px">
          Currently: <strong>{sched_hour:02d}:{sched_minute:02d} IST</strong>. Applies to all programmes.
        </div>
      </div>
    </div>

    <div class="card mb-4">
      <div class="card-header"><i class="bi bi-envelope-at" style="color:#00984C"></i> Notifications &amp; Digest</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="save_digest">
          <div class="mb-3 d-flex align-items-center gap-3">
            <div class="form-check form-switch mb-0">
              <input class="form-check-input" type="checkbox" name="digest_enabled"
                     {"checked" if digest_enabled else ""} id="digestSwitch">
              <label class="form-check-label" for="digestSwitch" style="font-size:13px">
                Weekly Digest Email (Board CEO &amp; Admin — every Monday)
              </label>
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label" style="font-size:12px">
              Programme Head Escalation Threshold (days overdue)
            </label>
            <input type="number" class="form-control" name="ph_escalation_days"
                   value="{ph_escalation_days}" min="1" max="30"
                   style="width:100px">
            <div style="font-size:11px;color:#94a3b8;margin-top:4px">
              Email Programme Head when a case is this many days past TAT.
            </div>
          </div>
          <button class="btn btn-primary w-100" type="submit">Save Settings</button>
        </form>
      </div>
    </div>

    <div class="card">
      <div class="card-header"><i class="bi bi-send-arrow-up" style="color:#7c3aed"></i> Outbound Webhook</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="save_webhook">
          <div class="mb-3">
            <label class="form-label" style="font-size:12px">Webhook URL</label>
            <input type="url" class="form-control" name="webhook_url"
                   value="{webhook_url}" placeholder="https://your-gateway/hook">
          </div>
          <button class="btn btn-primary w-100" type="submit">Save Webhook</button>
        </form>
        <div style="font-size:11px;color:#94a3b8;margin-top:8px">
          POST fired on every notification. JSON: <code style="font-size:10px">{{"event":"notification_r1","data":{{...}}}}</code><br>
          Use with MSG91, Gupshup, Twilio, etc. Leave blank to disable.
        </div>
      </div>
    </div>
  </div>

  <div class="col-lg-7">
    <div class="card mb-4">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span><i class="bi bi-calendar3-event" style="color:#d97706"></i> Holiday Calendar</span>
        <span style="font-size:11px;color:#94a3b8">TAT calculation skips these + weekends</span>
      </div>
      <div class="card-body p-4">
        <form method="post" class="row g-2 mb-3">
          <input type="hidden" name="action" value="add_holiday">
          <div class="col-5">
            <input type="date" class="form-control form-control-sm" name="holiday_date" required>
          </div>
          <div class="col-5">
            <input type="text" class="form-control form-control-sm" name="holiday_name"
                   placeholder="e.g. Diwali" required>
          </div>
          <div class="col-2">
            <button class="btn btn-sm btn-success w-100" type="submit">Add</button>
          </div>
        </form>
        <div style="max-height:220px;overflow-y:auto">
          <table class="data-table">
            <thead><tr><th>Date</th><th>Name</th><th style="width:40px"></th></tr></thead>
            <tbody>{holiday_rows}</tbody>
          </table>
        </div>
      </div>
    </div>

    <div class="card mb-4">
      <div class="card-header"><i class="bi bi-plug-fill" style="color:#0094ca"></i> SMTP Connection Test</div>
      <div class="card-body p-4">
        <div class="row g-2 mb-2">
          <div class="col-6"><input type="email" class="form-control form-control-sm" id="test_smtp_user" placeholder="sender@gmail.com"></div>
          <div class="col-6"><input type="password" class="form-control form-control-sm" id="test_smtp_pass" placeholder="App password"></div>
        </div>
        <div class="row g-2 mb-3">
          <div class="col-5"><input type="text" class="form-control form-control-sm" id="test_smtp_host" value="smtp.gmail.com"></div>
          <div class="col-3"><input type="number" class="form-control form-control-sm" id="test_smtp_port" value="587"></div>
          <div class="col-4"><input type="email" class="form-control form-control-sm" id="test_smtp_to" placeholder="Test recipient"></div>
        </div>
        <button class="btn btn-primary btn-sm" onclick="testSmtp()" id="smtpTestBtn">
          <i class="bi bi-send"></i> Send Test Email
        </button>
        <div id="smtpTestResult" style="font-size:13px;margin-top:10px"></div>
      </div>
    </div>

    <div class="card">
      <div class="card-header"><i class="bi bi-shield-lock-fill" style="color:#7c3aed"></i> User Security</div>
      <div style="overflow-x:auto">
        <table class="data-table">
          <thead><tr><th>Username</th><th>2FA</th><th>PW Reset</th><th>Actions</th></tr></thead>
          <tbody>{user_rows}</tbody>
        </table>
      </div>
    </div>
  </div>
</div>"""
    scripts = """<script>
function testSmtp(){
  var btn = document.getElementById('smtpTestBtn');
  btn.innerHTML = '<span class="spinner-border spinner-border-sm"></span> Testing…';
  btn.disabled = true;
  fetch('/test-smtp', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({
      email: document.getElementById('test_smtp_user').value,
      password: document.getElementById('test_smtp_pass').value,
      host: document.getElementById('test_smtp_host').value,
      port: parseInt(document.getElementById('test_smtp_port').value),
      to: document.getElementById('test_smtp_to').value
    })
  }).then(r=>r.json()).then(d=>{
    var res = document.getElementById('smtpTestResult');
    if(d.ok){
      res.innerHTML = '<span style="color:#00984C"><i class="bi bi-check-circle-fill"></i> Test email sent successfully!</span>';
    } else {
      res.innerHTML = '<span style="color:#dc2626"><i class="bi bi-x-circle-fill"></i> Failed: ' + (d.error||'Unknown error') + '</span>';
    }
    btn.innerHTML = '<i class="bi bi-send"></i> Send Test Email';
    btn.disabled = false;
  }).catch(e=>{
    document.getElementById('smtpTestResult').innerHTML = '<span style="color:#dc2626">Network error: '+e+'</span>';
    btn.innerHTML = '<i class="bi bi-send"></i> Send Test Email';
    btn.disabled = false;
  });
}
</script>"""
    return render_page(content, scripts, active_page="system", page_title="System Settings")


@app.route("/test-smtp", methods=["POST"])
@admin_required
def test_smtp():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data"})
    sender = data.get("email", "").strip()
    password = data.get("password", "").strip()
    host = data.get("host", "smtp.gmail.com").strip()
    port = int(data.get("port", 587))
    to = data.get("to", "").strip() or sender
    if not sender or not password:
        return jsonify({"ok": False, "error": "Sender email and password required"})
    try:
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = to
        msg["Subject"] = "QCI Notify — SMTP Test"
        msg.attach(MIMEText("This is a test email from QCI Notification Engine. SMTP is working correctly.", "plain"))
        if port == 465:
            with smtplib.SMTP_SSL(host, 465, timeout=15) as s:
                s.login(sender, password)
                s.sendmail(sender, [to], msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.starttls()
                s.login(sender, password)
                s.sendmail(sender, [to], msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]})


# ── Database Backup ───────────────────────────────────────────────────────────
@app.route("/backup")
@admin_required
def backup_db():
    flash(
        "Database backup is not available in PostgreSQL mode. "
        "Use pg_dump via your Railway database dashboard or CLI to export data.",
        "info",
    )
    return redirect(url_for("system_settings"))


@app.route("/export", methods=["GET", "POST"])
@login_required
def export_report():
    """Parameterised export — choose filters then download."""
    conn = get_db()
    bid = user_board_id()
    ph_progs = user_programme_names()

    if bid is not None:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes WHERE board_id=? ORDER BY programme_name", (bid,)
        ).fetchall()]
    else:
        programmes = [r[0] for r in conn.execute(
            "SELECT programme_name FROM programmes ORDER BY programme_name"
        ).fetchall()]
    if ph_progs is not None:
        programmes = [p for p in programmes if p in ph_progs]

    if request.method == "POST":
        status_f   = request.form.get("status", "")
        prog_f     = request.form.get("programme", "")
        owner_f    = request.form.get("owner_type", "")
        date_from  = request.form.get("date_from", "")
        date_to    = request.form.get("date_to", "")

        q = "SELECT * FROM case_tracking WHERE 1=1"
        params = []
        if bid is not None:
            q += " AND board_id=?"
            params.append(bid)
        if ph_progs is not None and ph_progs:
            placeholders = ",".join("?" * len(ph_progs))
            q += f" AND programme_name IN ({placeholders})"
            params.extend(ph_progs)
        if status_f:
            q += " AND status=?"
            params.append(status_f)
        if prog_f:
            q += " AND programme_name=?"
            params.append(prog_f)
        if owner_f:
            q += " AND owner_type=?"
            params.append(owner_f)
        if date_from:
            q += " AND stage_start_date >= ?"
            params.append(date_from)
        if date_to:
            q += " AND stage_start_date <= ?"
            params.append(date_to)

        cases = [dict(r) for r in conn.execute(q, params).fetchall()]
        try:
            _export_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                               for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
        except Exception:
            _export_db_hols = set()
        conn.close()
        today = now_ist().date()

        headers_row = [
            "Application ID", "Organisation", "Programme", "Stage", "Owner Type",
            "Action Owner", "Email", "Stage Start", "TAT Days", "Days Elapsed",
            "Case Status", "TAT Status", "Hold Days", "R1 Sent", "R2 Sent", "Overdue Sent", "Follow-ups"
        ]
        rows = []
        for c in cases:
            elapsed = working_days_elapsed(c["stage_start_date"], today, hold_days=c.get("hold_days", 0), extra_holidays=_export_db_hols)
            if c["is_milestone"]:
                tat_status = "Milestone"
            elif c.get("status") == "On Hold":
                tat_status = "On Hold"
            elif c["tat_days"] > 0 and elapsed >= c["tat_days"]:
                tat_status = "Overdue"
            elif c["tat_days"] > 0 and elapsed >= c.get("reminder2_day", 0):
                tat_status = "At Risk"
            else:
                tat_status = "On Track"
            rows.append([
                c["application_id"], c["organisation_name"], c["programme_name"],
                c["current_stage"], c["owner_type"] or "",
                c["action_owner_name"] or "", c["action_owner_email"] or "",
                c["stage_start_date"], c["tat_days"], elapsed,
                c.get("status", "Active"), tat_status, c.get("hold_days", 0),
                "Yes" if c["r1_sent"] else "No",
                "Yes" if c["r2_sent"] else "No",
                "Yes" if c["overdue_sent"] else "No", c["overdue_count"]
            ])

        fname_base = f"qci_export_{date.today()}"
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(headers_row)
        w.writerows(rows)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={fname_base}.csv"}
        )

    conn.close()
    prog_opts = '<option value="">All Programmes</option>' + "".join(
        f'<option value="{p}">{p}</option>' for p in programmes
    )
    content = f"""
<div class="row g-4 justify-content-center" style="max-width:700px;margin:0 auto">
  <div class="col-12">
    <div class="card">
      <div class="card-header">
        <i class="bi bi-file-earmark-spreadsheet" style="color:#059669"></i> Export Report
        <span style="font-size:12px;color:#94a3b8;margin-left:8px">Configure filters then download</span>
      </div>
      <div class="card-body p-4">
        <form method="post">
          <div class="row g-3">
            <div class="col-md-6">
              <label class="form-label">Programme</label>
              <select class="form-select" name="programme">{prog_opts}</select>
            </div>
            <div class="col-md-6">
              <label class="form-label">Case Status</label>
              <select class="form-select" name="status">
                <option value="">All Statuses</option>
                <option value="Active">Active</option>
                <option value="On Hold">On Hold</option>
                <option value="Closed">Closed</option>
                <option value="Withdrawn">Withdrawn</option>
                <option value="Suspended">Suspended</option>
              </select>
            </div>
            <div class="col-md-6">
              <label class="form-label">Owner Type</label>
              <select class="form-select" name="owner_type">
                <option value="">All Types</option>
                <option value="Applicant">Applicant</option>
                <option value="QCI">QCI</option>
                <option value="Assessor">Assessor</option>
                <option value="Hospital">Hospital</option>
              </select>
            </div>
            <div class="col-md-3">
              <label class="form-label">Stage Start From</label>
              <input type="date" class="form-control" name="date_from">
            </div>
            <div class="col-md-3">
              <label class="form-label">Stage Start To</label>
              <input type="date" class="form-control" name="date_to">
            </div>
            <div class="col-12 pt-2">
              <button type="submit" class="btn btn-success">
                <i class="bi bi-download"></i> Download CSV
              </button>
            </div>
          </div>
        </form>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="export", page_title="Export Report")


@app.route("/export-dashboard")
@login_required
def export_dashboard():
    return redirect(url_for("export_report"))


# ── Inbound REST API ──────────────────────────────────────────────────────────
@app.route("/api/v1/cases/advance", methods=["POST"])
def api_advance_case():
    """Inbound API: advance a case to a new stage.
    Auth: X-API-Key header.
    Body JSON: {application_id, stage_name, stage_start_date, action_owner_name,
                action_owner_email, organisation_name, programme_name, changed_by}
    """
    api_key = _verify_api_key(request)
    if not api_key:
        return jsonify({"ok": False, "error": "Invalid or missing API key"}), 401
    data = request.get_json(silent=True) or {}
    required = ["application_id", "stage_name", "programme_name"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"ok": False, "error": f"Missing fields: {', '.join(missing)}"}), 400

    # ── IDOR fix: verify programme belongs to the API key's board ──────────────
    if api_key.get("board_id"):
        _conn = get_db()
        _prog_row = _conn.execute(
            "SELECT id FROM programmes WHERE programme_name=? AND board_id=?",
            (data["programme_name"].strip(), api_key["board_id"])
        ).fetchone()
        _conn.close()
        if not _prog_row:
            return jsonify({"ok": False,
                            "error": "Programme not found or not authorised for this API key"}), 403

    try:
        upsert_data = {
            "application_id":     data["application_id"].strip(),
            "organisation_name":  data.get("organisation_name", "").strip(),
            "programme_name":     data["programme_name"].strip(),
            "stage_name":         data["stage_name"].strip(),
            "stage_start_date":   data.get("stage_start_date", date.today().isoformat()),
            "action_owner_name":  data.get("action_owner_name", "").strip(),
            "action_owner_email": data.get("action_owner_email", "").strip(),
            "program_officer_email": data.get("program_officer_email", "").strip(),
            "_changed_by":        data.get("changed_by", "API"),
            "_force_advance":     True,
        }
        action = upsert_case(upsert_data)
        return jsonify({"ok": True, "action": action, "application_id": data["application_id"]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/v1/cases/<app_id>", methods=["GET"])
def api_get_case(app_id):
    """Inbound API: get case details."""
    api_key = _verify_api_key(request)
    if not api_key:
        return jsonify({"ok": False, "error": "Invalid or missing API key"}), 401
    conn = get_db()
    case = conn.execute(
        "SELECT * FROM case_tracking WHERE application_id=?", (app_id,)
    ).fetchone()
    conn.close()
    if not case:
        return jsonify({"ok": False, "error": "Case not found"}), 404
    c = dict(case)
    c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], hold_days=c.get("hold_days", 0))
    return jsonify({"ok": True, "case": c})


@app.route("/api/v1/cases", methods=["GET"])
def api_list_cases():
    """Inbound API: list cases, optionally filtered by programme."""
    api_key = _verify_api_key(request)
    if not api_key:
        return jsonify({"ok": False, "error": "Invalid or missing API key"}), 401
    conn = get_db()
    q = "SELECT * FROM case_tracking WHERE 1=1"
    params = []
    if api_key.get("board_id"):
        q += " AND board_id=?"
        params.append(api_key["board_id"])
    if request.args.get("programme"):
        q += " AND programme_name=?"
        params.append(request.args["programme"])
    if request.args.get("status"):
        q += " AND status=?"
        params.append(request.args["status"])
    q += " LIMIT 200"
    cases = [dict(r) for r in conn.execute(q, params).fetchall()]
    try:
        _api_db_hols = {datetime.strptime(r[0][:10], "%Y-%m-%d").date()
                        for r in conn.execute("SELECT holiday_date FROM holidays").fetchall()}
    except Exception:
        _api_db_hols = set()
    conn.close()
    today = now_ist().date()
    for c in cases:
        c["days_elapsed"] = working_days_elapsed(c["stage_start_date"], today, hold_days=c.get("hold_days", 0), extra_holidays=_api_db_hols)
    return jsonify({"ok": True, "count": len(cases), "cases": cases})


@app.route("/api-keys", methods=["GET", "POST"])
@admin_required
def manage_api_keys():
    conn = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            raw_key = secrets.token_urlsafe(32)
            key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
            key_prefix = "qci_" + raw_key[:8]
            name = request.form.get("name", "Unnamed Key").strip()
            board_id = request.form.get("board_id") or None
            conn.execute(
                "INSERT INTO api_keys (key_hash, key_prefix, name, board_id, created_at) VALUES (?,?,?,?,?)",
                (key_hash, key_prefix, name, board_id, now_ist().strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
            flash(f"API key created. Copy it now — it won't be shown again: {raw_key}", "success")
        elif action == "revoke":
            conn.execute("UPDATE api_keys SET is_active=0 WHERE id=?", (request.form["key_id"],))
            conn.commit()
            flash("API key revoked.", "success")

    keys = [dict(r) for r in conn.execute(
        "SELECT id, name, key_prefix, board_id, created_at, last_used, is_active FROM api_keys ORDER BY id DESC"
    ).fetchall()]
    boards = [dict(r) for r in conn.execute("SELECT id, board_name FROM boards ORDER BY board_name").fetchall()]
    conn.close()

    board_opts = "".join(f'<option value="{b["id"]}">{b["board_name"]}</option>' for b in boards)
    key_rows = ""
    for k in keys:
        status_badge = '<span class="badge bg-success">Active</span>' if k["is_active"] else '<span class="badge bg-secondary">Revoked</span>'
        revoke_btn = f'''<form method="post" style="display:inline">
          <input type="hidden" name="action" value="revoke">
          <input type="hidden" name="key_id" value="{k["id"]}">
          <button class="btn btn-sm btn-outline-danger" style="font-size:11px"
                  {"disabled" if not k["is_active"] else ""}
                  onclick="return confirm('Revoke this key?')">Revoke</button>
        </form>''' if k["is_active"] else ""
        prefix_cell = f'<code style="font-size:11px;background:#f1f5f9;padding:1px 5px;border-radius:3px">{k["key_prefix"] or "—"}…</code>' if k.get("key_prefix") else '<span style="color:#94a3b8;font-size:11px">—</span>'
        key_rows += f"""<tr>
          <td style="font-weight:600">{k["name"]}</td>
          <td>{prefix_cell}</td>
          <td style="font-size:12px;color:#64748b">{k["created_at"]}</td>
          <td style="font-size:12px;color:#64748b">{k["last_used"] or "Never"}</td>
          <td>{status_badge}</td>
          <td>{revoke_btn}</td>
        </tr>"""

    content = f"""
<div class="row g-4">
  <div class="col-lg-8">
    <div class="card">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span><i class="bi bi-key-fill" style="color:#2563eb"></i> API Keys</span>
        <span style="font-size:12px;color:#94a3b8">Use X-API-Key header for all requests</span>
      </div>
      <div class="card-body p-0">
        <table class="data-table">
          <thead><tr><th>Name</th><th>Prefix</th><th>Created</th><th>Last Used</th><th>Status</th><th>Action</th></tr></thead>
          <tbody>{key_rows if key_rows else '<tr><td colspan="6" style="text-align:center;padding:30px;color:#94a3b8">No API keys yet.</td></tr>'}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="col-lg-4">
    <div class="card">
      <div class="card-header"><i class="bi bi-plus-circle" style="color:#059669"></i> Generate New Key</div>
      <div class="card-body p-4">
        <form method="post">
          <input type="hidden" name="action" value="create">
          <div class="mb-3">
            <label class="form-label">Key Name / Description</label>
            <input type="text" class="form-control" name="name" placeholder="e.g. NABH Portal Integration" required>
          </div>
          <div class="mb-3">
            <label class="form-label">Board Scope (optional)</label>
            <select class="form-select" name="board_id">
              <option value="">All Boards</option>
              {board_opts}
            </select>
          </div>
          <div style="background:#fef9c3;border:1px solid #fde68a;border-radius:8px;padding:10px 12px;font-size:11px;color:#92400e;margin-bottom:16px">
            <i class="bi bi-exclamation-triangle-fill"></i> The key is shown ONCE after creation. Store it securely.
          </div>
          <button class="btn btn-primary w-100">Generate Key</button>
        </form>
      </div>
    </div>
    <div class="card mt-3">
      <div class="card-header"><i class="bi bi-code-slash" style="color:#7c3aed"></i> API Reference</div>
      <div class="card-body p-3" style="font-size:12px">
        <strong>Advance a case:</strong>
        <pre style="background:#f8fafc;padding:8px;border-radius:6px;font-size:10px;overflow-x:auto">POST /api/v1/cases/advance
X-API-Key: your-key-here
Content-Type: application/json

{{
  "application_id": "NABH-001",
  "programme_name": "NABH Full...",
  "stage_name": "Document Review",
  "organisation_name": "ABC Hospital",
  "changed_by": "NABH Portal"
}}</pre>
        <strong>Get case:</strong>
        <pre style="background:#f8fafc;padding:8px;border-radius:6px;font-size:10px">GET /api/v1/cases/NABH-001
X-API-Key: your-key-here</pre>
      </div>
    </div>
  </div>
</div>"""
    return render_page(content, active_page="api_keys", page_title="API Keys")


@app.route("/healthz")
def healthz():
    """Public health-check — shows DB path and user count for diagnosis."""
    try:
        conn = get_db()
        user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        admin_exists = conn.execute("SELECT 1 FROM users WHERE username='admin'").fetchone() is not None
        conn.close()
        return jsonify({
            "status": "ok",
            "db_path": DATABASE_URL[:30] + "..." if DATABASE_URL else "not set",
            "user_count": user_count,
            "admin_exists": admin_exists,
        })
    except Exception as e:
        return jsonify({"status": "error", "detail": str(e)}), 500


@app.route("/run-check")
@admin_required
def run_check():
    # Called synchronously by Vercel Cron (schedule defined in vercel.json).
    # Vercel allows up to 5 minutes for cron functions, enough for email processing.
    summary = _scheduled_job()
    return jsonify(summary or {"status": "ok"})


# ── Scheduler ─────────────────────────────────────────────────────────────────
def _scheduled_job():
    """Run the daily check with an atomic DB-level lock so only one worker fires it."""
    worker_id = secrets.token_hex(8)
    now_str = now_ist().strftime("%Y-%m-%d %H:%M:%S")
    # Stale lock threshold: if another worker crashed, release locks older than 10 min
    stale_cutoff = (now_ist() - timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        # Delete any stale locks first
        conn.execute(
            "DELETE FROM scheduler_locks WHERE lock_name='daily_check' AND locked_at < ?",
            (stale_cutoff,),
        )
        conn.commit()
        # Atomically claim the lock — fails silently if another worker already holds it
        conn.execute(
            "INSERT INTO scheduler_locks (lock_name, locked_at, worker_id) VALUES (?,?,?) "
            "ON CONFLICT (lock_name) DO NOTHING",
            ("daily_check", now_str, worker_id),
        )
        conn.commit()
        # Check whether we won the lock
        row = conn.execute(
            "SELECT worker_id FROM scheduler_locks WHERE lock_name='daily_check'",
        ).fetchone()
        if not row or row["worker_id"] != worker_id:
            log.info("Scheduler lock held by another worker — skipping.")
            return {"status": "skipped", "reason": "lock held by another worker"}
    finally:
        conn.close()

    try:
        log.info("Scheduled daily check running… (worker %s)", worker_id)
        result = run_daily_check()
        log.info("Daily check complete: %s", result)
        return result
    finally:
        # Release the lock
        rel = get_db()
        try:
            rel.execute(
                "DELETE FROM scheduler_locks WHERE lock_name='daily_check' AND worker_id=?",
                (worker_id,),
            )
            rel.commit()
        finally:
            rel.close()


def _weekly_digest_job():
    """Wrapper called by APScheduler for the weekly digest."""
    try:
        run_weekly_digest()
    except Exception as e:
        log.error("Weekly digest job error: %s", e)


# ── Scheduler (disabled on Vercel — Cron Jobs call /run-check instead) ───────
class _NoOpScheduler:
    """Stub used on Vercel so scheduler.reschedule_job() calls don't crash."""
    def reschedule_job(self, *a, **kw): pass
    def add_job(self, *a, **kw): pass
    def start(self): pass

_IS_VERCEL = bool(os.environ.get("VERCEL"))
if _IS_VERCEL:
    scheduler = _NoOpScheduler()
    log.info("Vercel detected — APScheduler disabled; using Vercel Cron Jobs.")
else:
    try:
        from apscheduler.schedulers.background import BackgroundScheduler as _BGS
        scheduler = _BGS(timezone="Asia/Kolkata")
        with app.app_context():
            _sched_hour   = int(get_app_setting("scheduler_hour",   "8"))
            _sched_minute = int(get_app_setting("scheduler_minute", "0"))
        scheduler.add_job(_scheduled_job,     "cron",
                          hour=_sched_hour, minute=_sched_minute, id="daily_check")
        scheduler.add_job(_weekly_digest_job, "cron",
                          day_of_week="mon", hour=8, minute=0, id="weekly_digest")
        scheduler.add_job(_drain_webhook_queue, "interval", minutes=5, id="webhook_drain")
        scheduler.start()
    except Exception as _sched_err:
        log.warning("APScheduler not available: %s — using no-op stub.", _sched_err)
        scheduler = _NoOpScheduler()


# ── App startup ───────────────────────────────────────────────────────────────
# Vercel runs this at cold-start (module import). Tables are created if absent.
# Skipped when TESTING=1 so pytest can import without a live database.
if not os.environ.get("TESTING"):
    with app.app_context():
        init_db()
        migrate_data()
        seed_data()

if __name__ == "__main__":
    app.run(debug=True, use_reloader=False, port=5050)
